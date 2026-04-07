"""parse-quotes.py — 견적 엑셀 파일을 파싱하여 csagent.quotes / csagent.quote_items / csagent.item_price_history에 적재.

실행:
  python parse-quotes.py                # 전체 파싱
  python parse-quotes.py --dry-run      # DB 저장 없이 파싱 결과만 출력
  python parse-quotes.py --year 2026    # 특정 연도만
"""

import argparse
import logging
import os
import re
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation

import psycopg2
from psycopg2.extras import execute_values

# ── 설정 ──
QUOTES_DIR = r"C:\MES\wta-agents\data\cs_quotes"
DB_CONFIG = {
    "host": "localhost",
    "port": 55432,
    "user": "postgres",
    "password": "your-super-secret-and-long-postgres-password",
    "dbname": "postgres",
}

# xlsx: 헤더 Row 21, 데이터 Row 22+
# xls: 헤더 Row 22, 데이터 Row 23+ (일관된 +1 오프셋)
XLSX_DATA_START = 22
XLS_DATA_START = 23
ITEM_MAX_ROW = 50  # 최대 아이템 행

# 판매코드 패턴: B000000, P00000000, A00000000
SALES_CODE_PAT = re.compile(r"^[BPA]\d{5,9}$")

logging.basicConfig(
    level=logging.INFO,
    format="[parse-quotes] %(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("parse-quotes")


def safe_decimal(val):
    """숫자 값을 Decimal로 변환. 실패 시 None."""
    if val is None:
        return None
    try:
        d = Decimal(str(val))
        if d == 0:
            return None
        return d
    except (InvalidOperation, ValueError):
        return None


def safe_str(val):
    """문자열 변환. None이면 None, strip 적용."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_date(val):
    """날짜 파싱. 다양한 형식 지원."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y년 %m월 %d일"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # 숫자만 있는 경우 (20260106)
    m = re.match(r"(\d{4})(\d{2})(\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        except ValueError:
            pass
    return None


def calc_margin(unit_price, cost_price):
    """이익률 계산: (판매단가-원가)/판매단가 × 100."""
    if unit_price and cost_price and unit_price > 0:
        return round(float((unit_price - cost_price) / unit_price * 100), 2)
    return None


def parse_xlsx(filepath):
    """xlsx 파일 파싱. (header, items) 반환."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # 전체 행을 미리 읽기
    rows = {}
    for row in ws.iter_rows(min_row=1, max_row=ITEM_MAX_ROW, max_col=16, values_only=True):
        # iter_rows는 순서대로 반환
        pass

    # 다시 읽기 (read_only 모드에서는 재탐색 불가하므로 한번에 수집)
    wb.close()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ITEM_MAX_ROW, max_col=16, values_only=True):
        all_rows.append(row)
    wb.close()

    if len(all_rows) < XLSX_DATA_START:
        return None, []

    # 헤더 정보 (0-indexed)
    # CS번호: Row 5의 B 또는 C열 (양식에 따라 다름)
    cs_number = None
    if len(all_rows) > 4:
        for col_idx in (1, 2):  # B, C
            val = safe_str(all_rows[4][col_idx])
            if val and re.match(r"^[A-Z]{3}\d{9,}", val):
                cs_number = val
                break
        if not cs_number:
            cs_number = safe_str(all_rows[4][1]) or safe_str(all_rows[4][2])

    header = {
        "cs_number": cs_number,
        "customer": safe_str(all_rows[6][3]) if len(all_rows) > 6 else None,    # Row 7, D
        "subject": safe_str(all_rows[8][3]) if len(all_rows) > 8 else None,     # Row 9, D
        "contact": safe_str(all_rows[9][3]) if len(all_rows) > 9 else None,     # Row 10, D
        "quote_date": parse_date(all_rows[12][3] if len(all_rows) > 12 else None),  # Row 13, D
        "validity": safe_str(all_rows[13][3]) if len(all_rows) > 13 else None,  # Row 14, D
        "total_amount": safe_decimal(all_rows[17][4]) if len(all_rows) > 17 else None,  # Row 18, E
    }

    # 아이템 파싱 (Row 22+ = index 21+)
    items = []
    for i in range(XLSX_DATA_START - 1, len(all_rows)):
        row = all_rows[i]
        if len(row) < 9:
            continue

        # B열(순번)이 숫자가 아니면 아이템 끝
        line_no = row[1]
        if not isinstance(line_no, (int, float)):
            continue

        product_name = safe_str(row[2])   # C
        if not product_name:
            continue

        d_val = safe_str(row[3])          # D
        item_code = None
        remark = None
        if d_val and SALES_CODE_PAT.match(d_val):
            item_code = d_val
        elif d_val:
            remark = d_val

        unit_price = safe_decimal(row[7])  # H
        cost_price = safe_decimal(row[12]) if len(row) > 12 else None  # M
        ref_price = safe_decimal(row[13]) if len(row) > 13 else None   # N
        profit = safe_decimal(row[14]) if len(row) > 14 else None      # O

        items.append({
            "line_no": int(line_no),
            "product_name": product_name,
            "item_code": item_code,
            "unit": safe_str(row[5]),      # F
            "quantity": safe_decimal(row[6]),  # G
            "unit_price": unit_price,
            "amount": safe_decimal(row[8]),    # I
            "cost_price": cost_price,
            "ref_price": ref_price,
            "profit": profit,
            "margin_pct": calc_margin(unit_price, cost_price),
            "remark": remark,
        })

    return header, items


def parse_xls(filepath):
    """xls 파일 파싱. xlrd 사용. xlsx와 동일 구조 + 1행 오프셋."""
    import xlrd

    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    def cell_val(r, c):
        if r >= ws.nrows or c >= ws.ncols:
            return None
        cell = ws.cell(r, c)
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return None
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                return dt
            except Exception:
                return cell.value
        return cell.value

    # xls는 +1 오프셋: Row 6→cs_number, Row 8→customer, etc.
    # CS번호: Row 6의 B 또는 C열
    cs_number = None
    for col_idx in (1, 2):  # B, C
        val = safe_str(cell_val(5, col_idx))
        if val and re.match(r"^[A-Z]{3}\d{9,}", val):
            cs_number = val
            break
    if not cs_number:
        cs_number = safe_str(cell_val(5, 1)) or safe_str(cell_val(5, 2))

    header = {
        "cs_number": cs_number,
        "customer": safe_str(cell_val(7, 3)),       # Row 8, D
        "subject": safe_str(cell_val(9, 3)),        # Row 10, D
        "contact": safe_str(cell_val(10, 3)),       # Row 11, D
        "quote_date": parse_date(cell_val(13, 3)),  # Row 14, D
        "validity": safe_str(cell_val(14, 3)),      # Row 15, D
        "total_amount": safe_decimal(cell_val(18, 4)),  # Row 19, E
    }

    # 아이템 (Row 23+ = index 22+)
    items = []
    for r in range(XLS_DATA_START - 1, min(ws.nrows, ITEM_MAX_ROW)):
        line_no = cell_val(r, 1)  # B
        if not isinstance(line_no, (int, float)):
            continue

        product_name = safe_str(cell_val(r, 2))  # C
        if not product_name:
            continue

        d_val = safe_str(cell_val(r, 3))  # D
        item_code = None
        remark = None
        if d_val and SALES_CODE_PAT.match(d_val):
            item_code = d_val
        elif d_val:
            remark = d_val

        unit_price = safe_decimal(cell_val(r, 7))   # H
        cost_price = safe_decimal(cell_val(r, 12))   # M
        ref_price = safe_decimal(cell_val(r, 13))    # N
        profit = safe_decimal(cell_val(r, 14))       # O

        items.append({
            "line_no": int(line_no),
            "product_name": product_name,
            "item_code": item_code,
            "unit": safe_str(cell_val(r, 5)),    # F
            "quantity": safe_decimal(cell_val(r, 6)),  # G
            "unit_price": unit_price,
            "amount": safe_decimal(cell_val(r, 8)),    # I
            "cost_price": cost_price,
            "ref_price": ref_price,
            "profit": profit,
            "margin_pct": calc_margin(unit_price, cost_price),
            "remark": remark,
        })

    return header, items


def insert_quote(conn, header, items, file_name, file_year, company):
    """견적서 + 아이템 + 단가이력 DB 적재."""
    cur = conn.cursor()

    # 1. csagent.quotes INSERT (중복 시 SKIP)
    cur.execute("""
        INSERT INTO csagent.quotes (cs_number, customer, subject, quote_date, validity,
                               total_amount, contact, file_name, file_year, company)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (cs_number, file_name) DO NOTHING
        RETURNING id
    """, (
        header.get("cs_number") or "UNKNOWN",
        header.get("customer"),
        header.get("subject"),
        header.get("quote_date"),
        header.get("validity"),
        header.get("total_amount"),
        header.get("contact"),
        file_name,
        file_year,
        company,
    ))

    row = cur.fetchone()
    if row is None:
        # 이미 존재하는 견적 → skip
        cur.close()
        return 0

    quote_id = row[0]

    # 2. csagent.quote_items 배치 INSERT
    if items:
        item_rows = []
        for it in items:
            item_rows.append((
                quote_id, it["line_no"], it["product_name"], it["item_code"],
                it["unit"], it["quantity"], it["unit_price"], it["amount"],
                it["cost_price"], it["ref_price"], it["profit"], it["margin_pct"],
                it["remark"],
            ))

        execute_values(cur, """
            INSERT INTO csagent.quote_items
                (quote_id, line_no, product_name, item_code, unit, quantity,
                 unit_price, amount, cost_price, ref_price, profit, margin_pct, remark)
            VALUES %s
        """, item_rows, template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")

    # 3. csagent.item_price_history — 단가가 있는 아이템만
    price_rows = []
    for it in items:
        if it["unit_price"] and header.get("quote_date"):
            price_rows.append((
                it["item_code"],
                it["product_name"],
                header.get("customer"),
                it["unit_price"],
                it["cost_price"],
                it["margin_pct"],
                header["quote_date"],
                header.get("cs_number") or "UNKNOWN",
            ))

    if price_rows:
        execute_values(cur, """
            INSERT INTO csagent.item_price_history
                (item_code, product_name, customer, unit_price, cost_price,
                 margin_pct, quote_date, cs_number)
            VALUES %s
        """, price_rows, template="(%s,%s,%s,%s,%s,%s,%s,%s)")

    conn.commit()
    cur.close()
    return len(items)


def collect_files(base_dir, year_filter=None):
    """견적 파일 목록 수집. (filepath, file_year, company) 튜플 리스트."""
    files = []
    for year_dir in sorted(os.listdir(base_dir)):
        year_path = os.path.join(base_dir, year_dir)
        if not os.path.isdir(year_path):
            continue

        # 연도 추출
        m = re.search(r"(\d{4})", year_dir)
        if not m:
            continue
        file_year = int(m.group(1))

        if year_filter and file_year != year_filter:
            continue

        for company in sorted(os.listdir(year_path)):
            comp_path = os.path.join(year_path, company)
            if not os.path.isdir(comp_path):
                continue

            for f in sorted(os.listdir(comp_path)):
                if f.startswith("~$"):
                    continue  # 임시파일 제외
                ext = f.lower().rsplit(".", 1)[-1] if "." in f else ""
                if ext in ("xlsx", "xls"):
                    files.append((os.path.join(comp_path, f), file_year, company, ext))

    return files


def main():
    parser = argparse.ArgumentParser(description="견적 엑셀 → DB 파싱")
    parser.add_argument("--dry-run", action="store_true", help="DB 저장 없이 파싱 결과만 출력")
    parser.add_argument("--year", type=int, help="특정 연도만 파싱")
    args = parser.parse_args()

    files = collect_files(QUOTES_DIR, year_filter=args.year)
    log.info("파싱 대상: %d개 파일", len(files))

    if not files:
        log.info("대상 파일 없음")
        return

    conn = None
    if not args.dry_run:
        conn = psycopg2.connect(**DB_CONFIG)

    stats = {"ok": 0, "skip": 0, "err": 0, "items": 0}

    for filepath, file_year, company, ext in files:
        file_name = os.path.basename(filepath)
        try:
            if ext == "xlsx":
                header, items = parse_xlsx(filepath)
            else:
                header, items = parse_xls(filepath)

            if header is None:
                stats["skip"] += 1
                log.warning("파싱 실패(행 부족): %s", file_name)
                continue

            if args.dry_run:
                cs = header.get("cs_number", "?")
                log.info("[DRY] %s | CS=%s | items=%d", file_name[:50], cs, len(items))
                stats["ok"] += 1
                stats["items"] += len(items)
                continue

            count = insert_quote(conn, header, items, file_name, file_year, company)
            if count > 0:
                stats["ok"] += 1
                stats["items"] += count
            else:
                stats["skip"] += 1

        except Exception as e:
            stats["err"] += 1
            log.error("오류: %s — %s", file_name[:50], e)
            if conn:
                conn.rollback()

    if conn:
        conn.close()

    log.info("완료: 성공=%d, 건너뜀=%d, 오류=%d, 총아이템=%d",
             stats["ok"], stats["skip"], stats["err"], stats["items"])


if __name__ == "__main__":
    main()
