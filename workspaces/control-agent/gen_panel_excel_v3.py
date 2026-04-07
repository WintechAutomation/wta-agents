"""
대구텍 검사기 F2 #1 — 판넬별 전류용량 계산 및 차단기 선정표
Axis_Information 시트 기반으로 판넬 1~4 자동 생성
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 드라이버 전류 스펙 =====
# 파나소닉 A6, 단상 AC200V, 역률 0.85 기준 정격입력전류(A)
DRIVER_CURRENT = {
    'MADLN05BE': (0.9,  'AC200V 100W'),
    'MADLN10BE': (1.2,  'AC200V 100W'),
    'MADLN15BE': (1.5,  'AC200V 200W'),
    'MADLN23BE': (2.3,  'AC200V 400W'),
    'MBDLN25BE': (2.5,  'AC200V 400W'),
    'MCDLN35BE': (5.0,  'AC200V 750/800W'),
    'MCDLN35BL': (5.0,  'AC200V Linear'),
    'MDDLN55BL': (8.5,  'AC200V Linear'),
}

# 파스텍/EziServo (DC24V 기반) — DC24V 전류(A)
EZISERVO_CURRENT = {
    'Ezi-EC-42XL-A':       (2.0, 'DC24V StepServo'),
    'Ezi-EC-42M-A':        (1.5, 'DC24V StepServo'),
    'EZI-SPEED60H30CR100P': (3.0, 'DC24V ConveyorServo'),
    'MSMF042L1S2':         None,   # 드라이버 아님, 모터
    'MBDLN25BE':           None,
}

# ===== 원본 파일에서 축 정보 추출 =====
src_path = 'C:/MES/wta-agents/data/uploads/b96e8269-ad21-432a-b91d-0840dc9c34da/대구텍 검사기 F2 #1-버전2xlsx의 복사본_Copied[00].xlsx'
wb_src = openpyxl.load_workbook(src_path, data_only=True)
ws_src = wb_src['Axis_Information']

# 헤더 행 찾기
header_row = None
for i, row in enumerate(ws_src.iter_rows(values_only=True), 1):
    if row and 'Axis Name' in str(row):
        header_row = i
        break

panels = {}
current_panel = None

for i, row in enumerate(ws_src.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
    if not any(cell is not None for cell in row[:10]):
        continue

    panel_cell = row[1] if len(row) > 1 else None
    idx        = row[2] if len(row) > 2 else None
    atype      = row[3] if len(row) > 3 else None
    axis_name  = row[4] if len(row) > 4 else None
    actuator   = row[6] if len(row) > 6 else None
    drive      = row[7] if len(row) > 7 else None

    # 새 판넬 시작 판단
    if panel_cell and isinstance(panel_cell, str) and '판넬' in panel_cell:
        current_panel = panel_cell.strip()
        panels[current_panel] = []

    if current_panel is None:
        continue

    if axis_name is None or idx is None or atype is None:
        continue

    # 숫자 인덱스 행만
    if not isinstance(idx, (int, float)):
        continue

    axis_type = str(atype) if atype else ''

    # EziServo / StepServo (DC24V)
    if 'STEP' in axis_type or 'Ezi' in str(actuator):
        actuator_str = str(actuator) if actuator else ''
        dc_curr, dc_note = 0.0, 'DC24V'
        for k, v in EZISERVO_CURRENT.items():
            if k in actuator_str and v is not None:
                dc_curr, dc_note = v
                break
        panels[current_panel].append({
            'index': int(idx),
            'axis_name': str(axis_name),
            'type': 'STEP/EziServo',
            'actuator': actuator_str,
            'driver': '',
            'power_w': 0,
            'ac_curr': 0.0,
            'dc24v_curr': dc_curr,
            'note': dc_note,
        })
    else:
        # 파나소닉 서보 (AC200V)
        drive_str = str(drive) if drive else ''
        ac_curr = 0.0
        note = ''
        for k, (curr, n) in DRIVER_CURRENT.items():
            if k in drive_str:
                ac_curr = curr
                note = n
                break
        # 출력 파악
        power_w = 0
        if 'Absolute[' in axis_type:
            try:
                power_w = int(axis_type.split('[')[1].split('W')[0])
            except:
                pass

        panels[current_panel].append({
            'index': int(idx),
            'axis_name': str(axis_name),
            'type': 'Servo',
            'actuator': str(actuator) if actuator else '',
            'driver': drive_str,
            'power_w': power_w,
            'ac_curr': ac_curr,
            'dc24v_curr': 0.0,
            'note': note,
        })

# ===== 스타일 =====
hdr_font   = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
hdr_fill   = PatternFill('solid', fgColor='2F5496')
sub_fill   = PatternFill('solid', fgColor='4472C4')
sub_font   = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
yellow     = PatternFill('solid', fgColor='FFFF99')
green      = PatternFill('solid', fgColor='E2EFDA')
blue       = PatternFill('solid', fgColor='D6E4F0')
orange     = PatternFill('solid', fgColor='FCE4D6')
thin       = Side(style='thin', color='AAAAAA')
med        = Side(style='medium', color='555555')
border     = Border(left=thin, right=thin, top=thin, bottom=thin)
center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def cell_set(ws, addr, val, font=None, fill=None, align=None, bdr=None):
    c = ws[addr]
    c.value = val
    if font:  c.font = font
    if fill:  c.fill = fill
    if align: c.alignment = align
    if bdr:   c.border = bdr

# ===== 출력 워크북 =====
wb = openpyxl.Workbook()

# ===== Sheet1: 종합 요약 =====
ws1 = wb.active
ws1.title = '판넬별 차단기 선정'

# 제목
ws1.merge_cells('A1:P1')
ws1['A1'] = '대구텍 검사기 F2 #1 — 판넬별 전류용량 계산 및 차단기 선정표'
ws1['A1'].font = Font(name='맑은 고딕', bold=True, size=13, color='1F3864')
ws1['A1'].alignment = center
ws1.row_dimensions[1].height = 28

# 공통 정보
info = [('A2','전원:'),('B2','AC 단상 200V'),('C2','역률(PF):'),('D2',0.85),
        ('E2','안전계수:'),('F2',1.25),('G2','작성일:'),('H2','2026-04-03')]
for addr, val in info:
    ws1[addr] = val
    ws1[addr].font = Font(name='맑은 고딕', size=9, bold=(addr in ['A2','C2','E2','G2']))
for addr in ['B2','D2','F2','H2']:
    ws1[addr].fill = yellow
ws1.row_dimensions[2].height = 16
ws1.row_dimensions[3].height = 5

# 헤더 (4~5행)
# A~H: 축 정보 (병합)
axis_cols   = ['A','B','C','D','E','F','G','H']
axis_labels = ['판넬','No.','Axis Name','드라이버','정격출력\n(W)','AC전류\n(A)','DC24V전류\n(A)','비고']
for col, lbl in zip(axis_cols, axis_labels):
    ws1.merge_cells(f'{col}4:{col}5')
    ws1[f'{col}4'] = lbl
    ws1[f'{col}4'].font = hdr_font
    ws1[f'{col}4'].fill = hdr_fill
    ws1[f'{col}4'].alignment = center
    ws1[f'{col}4'].border = border

# I4~P4: 판넬 차단기 선정
ws1.merge_cells('I4:P4')
ws1['I4'] = '판넬 차단기 선정'
ws1['I4'].font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
ws1['I4'].fill = PatternFill('solid', fgColor='1F3864')
ws1['I4'].alignment = center
ws1['I4'].border = border

sub_lbls = ['AC 합계\n(A)','AC×1.25\n(A)','AC 권장\n차단기(A)','AC 선정\n차단기(A)',
            'DC24V합계\n(A)','DC24V×1.25','DC24V\n권장(A)','비고']
for j, (col, lbl) in enumerate(zip(['I','J','K','L','M','N','O','P'], sub_lbls)):
    ws1[f'{col}5'] = lbl
    ws1[f'{col}5'].font = sub_font
    ws1[f'{col}5'].fill = sub_fill
    ws1[f'{col}5'].alignment = center
    ws1[f'{col}5'].border = border

ws1.row_dimensions[4].height = 18
ws1.row_dimensions[5].height = 35

# 컬럼 너비
col_widths = {
    'A':10,'B':6,'C':22,'D':14,'E':9,'F':9,'G':9,'H':14,
    'I':10,'J':10,'K':11,'L':11,'M':11,'N':11,'O':11,'P':14
}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

# 데이터 채우기
ROW = 6
panel_summary = {}   # panel_name -> {ac_start, ac_end, dc_start, dc_end}

for panel_name, axes in panels.items():
    p_start = ROW

    for ax in axes:
        ws1.row_dimensions[ROW].height = 17
        vals = [
            panel_name if ax == axes[0] else '',
            ax['index'],
            ax['axis_name'],
            ax['driver'] or ax['actuator'],
            ax['power_w'] if ax['power_w'] else '',
            ax['ac_curr'] if ax['ac_curr'] else '',
            ax['dc24v_curr'] if ax['dc24v_curr'] else '',
            ax['note'],
        ]
        for j, (col, val) in enumerate(zip(axis_cols, vals)):
            c = ws1[f'{col}{ROW}']
            c.value = val
            c.font = Font(name='맑은 고딕', size=9)
            c.alignment = center if j in [0,1,4,5,6] else left_al
            c.border = border
            if col == 'A':
                c.fill = blue
            if col in ['F','G'] and val:
                c.fill = PatternFill('solid', fgColor='FFF2CC')
        ROW += 1

    # 합계행
    ws1.row_dimensions[ROW].height = 18
    ws1.merge_cells(f'A{ROW}:H{ROW}')
    ws1[f'A{ROW}'] = f'▶ {panel_name} 합계'
    ws1[f'A{ROW}'].font = Font(name='맑은 고딕', bold=True, size=9)
    ws1[f'A{ROW}'].alignment = center
    ws1[f'A{ROW}'].fill = green
    ws1[f'A{ROW}'].border = border

    ac_sum_formula  = f'=SUM(F{p_start}:F{ROW-1})'
    dc_sum_formula  = f'=SUM(G{p_start}:G{ROW-1})'

    ws1[f'I{ROW}'] = ac_sum_formula
    ws1[f'J{ROW}'] = f'=I{ROW}*1.25'
    ws1[f'K{ROW}'] = '(자동계산)'
    ws1[f'L{ROW}'] = ''
    ws1[f'M{ROW}'] = dc_sum_formula
    ws1[f'N{ROW}'] = f'=M{ROW}*1.25'
    ws1[f'O{ROW}'] = '(자동계산)'
    ws1[f'P{ROW}'] = ''

    for col in ['I','J','K','L','M','N','O','P']:
        ws1[f'{col}{ROW}'].font = Font(name='맑은 고딕', bold=True, size=9)
        ws1[f'{col}{ROW}'].alignment = center
        ws1[f'{col}{ROW}'].fill = green
        ws1[f'{col}{ROW}'].border = border

    panel_summary[panel_name] = {'sum_row': ROW}
    ROW += 1
    ws1.row_dimensions[ROW].height = 5
    ROW += 1

# 수식 설명 박스
ROW += 1
ws1.merge_cells(f'A{ROW}:P{ROW}')
ws1[f'A{ROW}'] = '【 수식 설명 】'
ws1[f'A{ROW}'].font = Font(name='맑은 고딕', bold=True, size=10, color='1F3864')
ws1[f'A{ROW}'].alignment = left_al
ROW += 1
formulas = [
    '① AC 전류 (A) = 드라이버 정격입력전류 (파나소닉 A6 스펙시트 기준, 단상 AC200V)',
    '② DC24V 전류 (A) = EziServo/EziSpeed 정격전류 (파스텍 스펙시트 기준)',
    '③ AC 합계 = Σ(각 축 AC 운전전류)',
    '④ AC×1.25 = 합계전류 × 안전계수 1.25 → 차단기 정격 산정 기준',
    '⑤ DC24V 합계 = Σ(EziServo/EziSpeed DC 전류)',
    '⑥ Linear 드라이버(MCDLN35BL, MDDLN55BL)는 실측 전류 입력 권장',
]
for f in formulas:
    ws1.merge_cells(f'A{ROW}:P{ROW}')
    ws1[f'A{ROW}'] = f
    ws1[f'A{ROW}'].font = Font(name='맑은 고딕', size=9)
    ws1[f'A{ROW}'].alignment = left_al
    ws1[f'A{ROW}'].fill = PatternFill('solid', fgColor='F2F2F2')
    ws1.row_dimensions[ROW].height = 15
    ROW += 1


# ===== Sheet2: 파나소닉 A6 드라이버 규격표 =====
ws2 = wb.create_sheet('파나소닉A6 드라이버 규격')
ws2.merge_cells('A1:G1')
ws2['A1'] = '파나소닉 A6 서보드라이버 규격표 (단상 AC200V 기준)'
ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color='1F3864')
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 22
ws2.row_dimensions[2].height = 5

hdrs = ['모델','정격출력(W)','정격입력전류(A)','입력전압','통신','비고','비고2']
for j, h in enumerate(hdrs):
    col = get_column_letter(j+1)
    ws2[f'{col}3'] = h
    ws2[f'{col}3'].font = hdr_font
    ws2[f'{col}3'].fill = hdr_fill
    ws2[f'{col}3'].alignment = center
    ws2[f'{col}3'].border = border
ws2.row_dimensions[3].height = 22

a6_data = [
    ['MADLN05BE',  100, 0.9,  'AC100/200V', 'EtherCAT', '소형 서보, I/O'],
    ['MADLN10BE',  100, 1.2,  'AC100/200V', 'EtherCAT', ''],
    ['MADLN15BE',  200, 1.5,  'AC100/200V', 'EtherCAT', ''],
    ['MADLN23BE',  400, 2.3,  'AC100/200V', 'EtherCAT', ''],
    ['MBDLN25BE',  400, 2.5,  'AC100/200V', 'EtherCAT', ''],
    ['MBDLN55BE',  750, 5.0,  'AC200V',     'EtherCAT', ''],
    ['MCDLN35BE',  750, 5.0,  'AC200V',     'EtherCAT', '800W 포함'],
    ['MCDLN35BL',  '-', 5.0,  'AC200V',     'EtherCAT', 'Linear Motor Driver'],
    ['MDDLN55BL',  '-', 8.5,  'AC200V',     'EtherCAT', 'Linear Motor Driver (고전류)'],
]
for i, row in enumerate(a6_data):
    r = i + 4
    ws2.row_dimensions[r].height = 16
    for j, val in enumerate(row):
        col = get_column_letter(j+1)
        ws2[f'{col}{r}'] = val
        ws2[f'{col}{r}'].font = Font(name='맑은 고딕', size=9)
        ws2[f'{col}{r}'].alignment = center
        ws2[f'{col}{r}'].border = border
        if j == 0:
            ws2[f'{col}{r}'].fill = blue
        if j == 2:
            ws2[f'{col}{r}'].fill = yellow

ws2_widths = {'A':16,'B':12,'C':14,'D':14,'E':12,'F':20,'G':10}
for col, w in ws2_widths.items():
    ws2.column_dimensions[col].width = w


# ===== Sheet3: 파스텍 드라이버 규격표 =====
ws3 = wb.create_sheet('파스텍 EziServo 규격')
ws3.merge_cells('A1:G1')
ws3['A1'] = '파스텍 EziServo / EziSpeed 규격표'
ws3['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color='1F3864')
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 22
ws3.row_dimensions[2].height = 5

hdrs3 = ['모델','전원','정격전류(A)','통신','최대토크','비고','']
for j, h in enumerate(hdrs3):
    col = get_column_letter(j+1)
    ws3[f'{col}3'] = h
    ws3[f'{col}3'].font = hdr_font
    ws3[f'{col}3'].fill = hdr_fill
    ws3[f'{col}3'].alignment = center
    ws3[f'{col}3'].border = border
ws3.row_dimensions[3].height = 22

ezi_data = [
    ['Ezi-EC-42XL-A',       'DC 24V',    2.0, 'EtherCAT', '1.2 N·m', 'EziServo EC 소형'],
    ['Ezi-EC-42M-A',        'DC 24V',    1.5, 'EtherCAT', '0.6 N·m', 'EziServo EC 미니'],
    ['EZI-SPEED60H30CR100P','DC 24V',    3.0, 'EtherCAT', 'N/A',     '컨베이어 전용 서보'],
    ['EM542S',              'DC 20~50V', 4.2, 'Pulse/Dir','N/A',     '스텝드라이버 (단독)'],
    ['EM806',               'DC 24~80V', 8.0, 'Pulse/Dir','N/A',     '스텝드라이버 (고전류)'],
]
for i, row in enumerate(ezi_data):
    r = i + 4
    ws3.row_dimensions[r].height = 16
    for j, val in enumerate(row):
        col = get_column_letter(j+1)
        ws3[f'{col}{r}'] = val
        ws3[f'{col}{r}'].font = Font(name='맑은 고딕', size=9)
        ws3[f'{col}{r}'].alignment = center
        ws3[f'{col}{r}'].border = border
        if j == 0: ws3[f'{col}{r}'].fill = orange
        if j == 2: ws3[f'{col}{r}'].fill = yellow

ws3_widths = {'A':26,'B':14,'C':12,'D':12,'E':12,'F':22,'G':10}
for col, w in ws3_widths.items():
    ws3.column_dimensions[col].width = w


# ===== Sheet4: 차단기 규격 참조표 =====
ws4 = wb.create_sheet('차단기 규격 참조표')
ws4.merge_cells('A1:F1')
ws4['A1'] = '차단기 규격 참조표 (MCCB / MCB 3P, AC 200V)'
ws4['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color='1F3864')
ws4['A1'].alignment = center
ws4.row_dimensions[1].height = 22
ws4.row_dimensions[2].height = 5

hdrs4 = ['정격전류(A)', '적용 부하전류 범위', '제조사', '모델 (3P)', '비고', '']
for j, h in enumerate(hdrs4):
    col = get_column_letter(j+1)
    ws4[f'{col}3'] = h
    ws4[f'{col}3'].font = hdr_font
    ws4[f'{col}3'].fill = hdr_fill
    ws4[f'{col}3'].alignment = center
    ws4[f'{col}3'].border = border
ws4.row_dimensions[3].height = 22

breaker_data = [
    [5,  '~4A',   'LS/Chint', 'ABN5c / NXB-63',  '소형 I/O, EziServo 2~3개'],
    [10, '4~8A',  'LS/Chint', 'ABN10c / NXB-63', '소형 서보 판넬 (100W×5축 이하)'],
    [15, '8~12A', 'LS/Chint', 'ABN15c / NXB-63', ''],
    [20,'12~16A', 'LS/Chint', 'ABN20c / NXB-63', '중형 서보 판넬'],
    [25,'16~20A', 'LS/Chint', 'ABN25c / NXB-63', ''],
    [30,'20~24A', 'LS/Chint', 'ABN30c / NXB-63', '다축 서보 판넬'],
    [40,'24~32A', 'LS/Chint', 'ABN40c',           ''],
    [50,'32~40A', 'LS/Chint', 'ABN50c',           '대형 판넬 (800W 3축+)'],
    [63,'40~50A', 'LS/Chint', 'ABN63c',           ''],
    [75,'50~60A', 'LS/Chint', 'ABN75c',           '메인 차단기 (복합 판넬)'],
    [100,'60~80A','LS/Chint', 'ABN100c',           ''],
]
for i, row in enumerate(breaker_data):
    r = i + 4
    ws4.row_dimensions[r].height = 16
    for j, val in enumerate(row):
        col = get_column_letter(j+1)
        ws4[f'{col}{r}'] = val
        ws4[f'{col}{r}'].font = Font(name='맑은 고딕', size=9)
        ws4[f'{col}{r}'].alignment = center
        ws4[f'{col}{r}'].border = border
        if j == 0: ws4[f'{col}{r}'].fill = green

ws4_widths = {'A':12,'B':18,'C':12,'D':18,'E':22,'F':8}
for col, w in ws4_widths.items():
    ws4.column_dimensions[col].width = w


# ===== 저장 =====
out = 'C:/MES/wta-agents/workspaces/control-agent/대구텍_F2_1_판넬별_차단기선정표.xlsx'
wb.save(out)
print('저장 완료:', out)
