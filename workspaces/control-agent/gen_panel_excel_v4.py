"""
대구텍 검사기 F2 #1 — 판넬별 전류용량 계산 및 차단기 선정표 v4
- v2 형식 기반 (제목, 상단 요약, R/S/T 편중분석, 본문 구조 유지)
- 실제 Axis_Information 데이터 반영 (판넬1~4)
- 드라이버 모델 입력 → VLOOKUP 전류 자동계산
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 공통 스타일 헬퍼 =====
def mfont(bold=False, size=9, color='000000'):
    return Font(name='맑은 고딕', bold=bold, size=size, color=color)
def mfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)
def mborder():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)
def mcenter():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)
def mleft():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

C_HEADER  = '2F5496'
C_SUBHDR  = '4472C4'
C_TITLE   = '1F3864'
C_PAN     = '00B0F0'   # 파나소닉 (하늘)
C_EZI     = 'ED7D31'   # EziServo (주황)
C_DC      = '70AD47'   # DC24V (초록)
C_MISC    = '7030A0'   # 기타 (보라)
C_SUM     = 'E2EFDA'
C_FORMULA = 'FFF2CC'
C_INFO    = 'D6E4F0'
thin_b    = mborder()
center    = mcenter()
left_a    = mleft()

def sc(ws, row, col, val, bold=False, size=9, color='000000',
       fill=None, align=None, bdr=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = mfont(bold=bold, size=size, color=color)
    if fill:  c.fill = fill
    if align: c.alignment = align
    if bdr:   c.border = bdr
    if fmt:   c.number_format = fmt
    return c

# ============================================================
# SHEET 2: 파나소닉 A6 드라이버 규격 (VLOOKUP 참조용)
# ============================================================
ws2 = wb.create_sheet('파나소닉A6규격')
ws2.title = '파나소닉A6규격'

ws2.merge_cells('A1:D1')
ws2['A1'] = '파나소닉 A6 드라이버 규격 (VLOOKUP 참조)'
ws2['A1'].font = mfont(bold=True, size=10, color=C_TITLE)
ws2['A1'].alignment = center

hdrs2 = ['드라이버 모델', '정격출력(W)', '정격입력전류(A)', '비고']
for j, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=3, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws2.row_dimensions[3].height = 20

# A6 드라이버 데이터 (단상 AC200V 기준)
a6_specs = [
    ('MADLN05BE',  100, 0.9,  '50W~100W 서보'),
    ('MADLN10BE',  100, 1.2,  '100W 서보'),
    ('MADLN15BE',  200, 1.5,  '200W 서보'),
    ('MADLN23BE',  400, 2.3,  '400W 서보'),
    ('MBDLN25BE',  400, 2.5,  '400W 서보'),
    ('MBDLN55BE',  750, 5.0,  '750W 서보'),
    ('MCDLN35BE',  800, 5.0,  '750W~800W 서보'),
    ('MCDLN35BL',    0, 5.0,  'Linear Motor Driver'),
    ('MDDLN55BL',    0, 8.5,  'Linear Motor Driver (대형)'),
]
for i, (model, watt, curr, note) in enumerate(a6_specs):
    r = i + 4
    ws2.row_dimensions[r].height = 16
    data = [model, watt, curr, note]
    for j, val in enumerate(data, 1):
        c = ws2.cell(row=r, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j != 4 else left_a
        c.border = thin_b
        if j == 1: c.fill = mfill(C_INFO)
        if j == 3: c.fill = mfill('FFFF99')

ws2_widths = {'A': 16, 'B': 12, 'C': 16, 'D': 22}
for col, w in ws2_widths.items():
    ws2.column_dimensions[col].width = w


# ============================================================
# SHEET 3: 파스텍 EziServo 규격 (VLOOKUP 참조용)
# ============================================================
ws3 = wb.create_sheet('파스텍EziServo규격')
ws3.title = '파스텍EziServo규격'

ws3.merge_cells('A1:D1')
ws3['A1'] = '파스텍 EziServo / EziSpeed 규격 (VLOOKUP 참조)'
ws3['A1'].font = mfont(bold=True, size=10, color=C_TITLE)
ws3['A1'].alignment = center

for j, h in enumerate(['드라이버/액추에이터 모델', '전원', '정격전류(A)', '비고'], 1):
    c = ws3.cell(row=3, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws3.row_dimensions[3].height = 20

ezi_specs = [
    ('Ezi-EC-42XL-A',        'DC 24V', 2.0, 'EziServo EC 소형'),
    ('Ezi-EC-42M-A',         'DC 24V', 1.5, 'EziServo EC 미니'),
    ('EZI-SPEED60H30CR100P', 'DC 24V', 3.0, 'EziSpeed 컨베이어'),
    ('EM542S',               'DC 24~50V', 4.2, '파스텍 스텝드라이버'),
    ('EM806',                'DC 24~80V', 8.0, '파스텍 스텝드라이버'),
]
for i, (model, pwr, curr, note) in enumerate(ezi_specs):
    r = i + 4
    ws3.row_dimensions[r].height = 16
    for j, val in enumerate([model, pwr, curr, note], 1):
        c = ws3.cell(row=r, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j != 4 else left_a
        c.border = thin_b
        if j == 1: c.fill = mfill('FCE4D6')
        if j == 3: c.fill = mfill('FFFF99')

ws3_widths = {'A': 26, 'B': 14, 'C': 14, 'D': 22}
for col, w in ws3_widths.items():
    ws3.column_dimensions[col].width = w


# ============================================================
# SHEET 4: 차단기 규격 참조
# ============================================================
ws4 = wb.create_sheet('차단기규격')
ws4.merge_cells('A1:E1')
ws4['A1'] = '차단기 규격 참조표 (MCCB/MCB 3P, AC200V)'
ws4['A1'].font = mfont(bold=True, size=10, color=C_TITLE)
ws4['A1'].alignment = center

for j, h in enumerate(['정격(A)', '적용 전류범위', '제조사', '모델명', '비고'], 1):
    c = ws4.cell(row=3, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws4.row_dimensions[3].height = 20

bkr = [
    (5,  '~4A',    'LS/Chint','ABN5c / NXB-63',  'EziServo 2~3개'),
    (10, '4~8A',   'LS/Chint','ABN10c / NXB-63', '100W 서보 5축 이하'),
    (15, '8~12A',  'LS/Chint','ABN15c / NXB-63', ''),
    (20, '12~16A', 'LS/Chint','ABN20c / NXB-63', '중형 판넬'),
    (25, '16~20A', 'LS/Chint','ABN25c / NXB-63', ''),
    (30, '20~24A', 'LS/Chint','ABN30c / NXB-63', '다축 서보'),
    (40, '24~32A', 'LS/Chint','ABN40c',           ''),
    (50, '32~40A', 'LS/Chint','ABN50c',           '대형 판넬'),
    (63, '40~50A', 'LS/Chint','ABN63c',           ''),
    (75, '50~60A', 'LS/Chint','ABN75c',           '메인 차단기'),
    (100,'60~80A', 'LS/Chint','ABN100c',          ''),
]
for i, row_data in enumerate(bkr):
    r = i + 4
    ws4.row_dimensions[r].height = 15
    for j, val in enumerate(row_data, 1):
        c = ws4.cell(row=r, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center
        c.border = thin_b
        if j == 1: c.fill = mfill(C_SUM)

for col, w in {'A':10,'B':16,'C':12,'D':18,'E':22}.items():
    ws4.column_dimensions[col].width = w


# ============================================================
# SHEET 1: 전류·용량 계산표 (v2 형식 기반)
# ============================================================
ws = wb.active
ws.title = '전류·용량 계산표'

# 컬럼 너비 (A~Q)
col_widths = {
    1:8, 2:16, 3:20, 4:8, 5:10, 6:5, 7:6, 8:6, 9:6, 10:10,
    11:7, 12:7, 13:7, 14:7, 15:10, 16:16, 17:30
}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── 1행: 시트 제목 ──
ws.merge_cells('A1:Q1')
ws['A1'] = '대구텍 검사기 F2 #1 — 판넬별 전류용량 계산 및 차단기 선정표'
ws['A1'].font = mfont(bold=True, size=14, color=C_TITLE)
ws['A1'].alignment = center
ws['A1'].fill = mfill('D9E2F3')
ws.row_dimensions[1].height = 28

# ── 2~4행: 전체 요약 박스 ──
ws.merge_cells('A2:Q2')
ws['A2'] = '▶ 전체 부하 요약 (상단 자동계산 — 하단 합계행 수식 연동)'
ws['A2'].font = mfont(bold=True, size=10, color='FFFFFF')
ws['A2'].fill = mfill(C_TITLE)
ws['A2'].alignment = center
ws.row_dimensions[2].height = 18

sum_hdrs = [(1,'구분'),(2,'총 정격전류(A)'),(3,'안전계수×1.25'),(4,'권장 차단기(A)'),(5,'비고')]
for col, label in sum_hdrs:
    c = ws.cell(row=3, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[3].height = 18

# 요약 행 4~7 (수식은 나중에 채움 — 행 번호 확정 후)
sum_rows_def = [
    (4, 'AC 서보드라이버 (파나소닉 A6)',   mfill('EBF3FB')),
    (5, 'DC24V EziServo/EziSpeed',          mfill('FEF4E8')),
    (6, 'DC 24V 공통 부하 (I/O·센서 등)',   mfill('EDF7E3')),
    (7, '기타 (HMI·조명·PC 등)',            mfill('F3EBF9')),
]
for row, label, fill in sum_rows_def:
    c = ws.cell(row=row, column=1, value=label)
    c.font = mfont(bold=True, size=9)
    c.fill = fill
    c.alignment = left_a
    c.border = thin_b
    for col in range(2, 6):
        cell = ws.cell(row=row, column=col, value='(자동)')
        cell.fill = fill
        cell.font = mfont(size=9, color='888888')
        cell.alignment = center
        cell.border = thin_b
    ws.row_dimensions[row].height = 16

ws.row_dimensions[8].height = 6

# ── 9행: R/S/T 편중 분석 헤더 ──
ws.merge_cells('A9:Q9')
ws['A9'] = '▶ R/S/T 상별 전류 편중 분석 (AC 서보드라이버 기준)'
ws['A9'].font = mfont(bold=True, size=10, color='FFFFFF')
ws['A9'].fill = mfill('833C00')
ws['A9'].alignment = center
ws.row_dimensions[9].height = 18

phase_hdrs = ['','R상(A)','S상(A)','T상(A)','최대-최소(A)','불균형률(%)','판정','기준: 불균형률 10% 이하 권장']
for i, h in enumerate(phase_hdrs, 1):
    c = ws.cell(row=10, column=i, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill('C55A11')
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[10].height = 18

for i, label in enumerate(['AC 서보드라이버 합계','판넬1','판넬2','판넬3','판넬4'], 11):
    ws.cell(row=i, column=1, value=label).font = mfont(bold=True, size=9)
    ws.cell(row=i, column=1).border = thin_b
    ws.cell(row=i, column=1).alignment = left_a
    for col in range(2, 9):
        c = ws.cell(row=i, column=col, value='')
        c.border = thin_b
        c.alignment = center
    ws.row_dimensions[i].height = 15

ws.row_dimensions[16].height = 6

# ── 17행: 본문 메인 헤더 ──
HEADER_ROW = 17

# 단일 컬럼 (세로 병합)
for col, label in [(1,'판넬'),(2,'Axis Name'),(3,'드라이버 모델\n(직접 입력)'),
                    (4,'출력(W)\n[자동]'),(5,'정격전류(A)\n[자동]'),
                    (6,'상수'),(7,'R상'),(8,'S상'),(9,'T상'),
                    (10,'운전전류(A)')]:
    ws.merge_cells(f'{get_column_letter(col)}{HEADER_ROW}:{get_column_letter(col)}{HEADER_ROW+1}')
    c = ws.cell(row=HEADER_ROW, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b

# 전원 구분 (가로 병합)
ws.merge_cells(f'K{HEADER_ROW}:N{HEADER_ROW}')
ws.cell(row=HEADER_ROW, column=11, value='전원 구분').font = mfont(bold=True, size=9, color='FFFFFF')
ws.cell(row=HEADER_ROW, column=11).fill = mfill(C_SUBHDR)
ws.cell(row=HEADER_ROW, column=11).alignment = center
ws.cell(row=HEADER_ROW, column=11).border = thin_b

for col, label in [(11,'AC 서보\n(파나소닉)'),(12,'DC 24V\n(EziServo)'),(13,'DC 24V\n(공통)'),(14,'기타')]:
    c = ws.cell(row=HEADER_ROW+1, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b

for col, label in [(15,'차단기\n선정(A)'),(16,'제조사/모델'),(17,'비고/수식설명')]:
    ws.merge_cells(f'{get_column_letter(col)}{HEADER_ROW}:{get_column_letter(col)}{HEADER_ROW+1}')
    c = ws.cell(row=HEADER_ROW, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b

ws.row_dimensions[HEADER_ROW].height = 20
ws.row_dimensions[HEADER_ROW+1].height = 26

# ===== 실제 판넬 데이터 =====
# (판넬명, axis_name, 드라이버모델, 상수, R, S, T, 전원구분, 비고)
# 전원구분: 'AC'=파나소닉서보, 'EZI'=EziServo/EziSpeed DC, 'DC'=DC24V공통, 'ETC'=기타
PANEL_DATA = [
    # === 판넬 1 ===
    ('판넬1','LO_Y1',       'MDDLN55BL',             3,1,1,1,'AC', 'Linear Motor Driver (AC200V)'),
    ('',    'LO_X1',        'MCDLN35BL',             3,0,1,1,'AC', 'Linear Motor Driver (AC200V)'),
    ('',    'LO_Z1',        'MADLN05BE',             1,1,0,0,'AC', 'I=VA/(220×PF=0.85)'),
    ('',    'ULO_Y1',       'MCDLN35BL',             3,1,1,0,'AC', 'Linear Motor Driver'),
    ('',    'ULO_X1',       'MCDLN35BL',             3,0,1,1,'AC', 'Linear Motor Driver'),
    ('',    'ULO_Z1',       'MADLN05BE',             1,0,0,1,'AC', ''),
    ('',    'NG_EV',        'MCDLN35BE',             1,1,0,0,'AC', '800W 서보 (단상)'),
    ('',    'NG_FD',        'MADLN05BE',             1,0,1,0,'AC', ''),
    ('',    'NG_FD_Z',      'MBDLN25BE',             1,1,0,0,'AC', ''),
    ('',    'NG_FD_ALIGN',  'MADLN05BE',             1,0,0,1,'AC', ''),
    ('',    'NG_FD_INDEX',  'MBDLN25BE',             1,1,0,0,'AC', ''),
    ('',    'JIG_ULO_Y',    'MADLN05BE',             1,0,1,0,'AC', ''),
    ('',    'JIG_ULO_Z',    'MADLN05BE',             1,0,0,1,'AC', ''),
    ('',    'LO_R1 (#1)',   'Ezi-EC-42XL-A',         1,0,0,0,'EZI','EziServo EC (DC24V)'),
    ('',    'LO_R1 (#2)',   'Ezi-EC-42XL-A',         1,0,0,0,'EZI','EziServo EC (DC24V)'),
    # === 판넬 2 ===
    ('판넬2','GOOD_EV',     'MCDLN35BE',             1,1,0,0,'AC', '800W 서보'),
    ('',    'GOOD_FD',      'MADLN05BE',             1,0,1,0,'AC', ''),
    ('',    'GOOD_FD_Z',    'MBDLN25BE',             1,1,0,0,'AC', ''),
    ('',    'GOOD_FD_ALIGN','MADLN05BE',             1,0,0,1,'AC', ''),
    ('',    'GOOD_FD_INDEX','MBDLN25BE',             1,0,1,0,'AC', ''),
    ('',    'TURN_Y',       'MBDLN25BE',             1,1,0,0,'AC', ''),
    ('',    'TURN_GRIPPER_Y1','MADLN15BE',           1,0,1,0,'AC', ''),
    ('',    'TURN_GRIPPER_X1','MADLN15BE',           1,0,0,1,'AC', ''),
    ('',    'TURN_GRIPPER_Z1','MADLN05BE',           1,1,0,0,'AC', '50W 모터'),
    ('',    'TURN_GRIPPER_Y2','MADLN15BE',           1,0,1,0,'AC', ''),
    ('',    'TURN_GRIPPER_X2','MADLN15BE',           1,0,0,1,'AC', ''),
    ('',    'TURN_GRIPPER_Z2','MADLN05BE',           1,1,0,0,'AC', '50W 모터'),
    ('',    'BOT_MAC_Y',    'MADLN15BE',             1,0,1,0,'AC', ''),
    # === 판넬 3 ===
    ('판넬3','BOT_MAC_X',   'MADLN15BE',             1,0,0,1,'AC', ''),
    ('',    'BOT_MAC_Z',    'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'BOT_MIC_Y',    'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'BOT_MIC_X',    'MADLN15BE',             1,0,0,1,'AC', ''),
    ('',    'BOT_MIC_Z',    'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'LO_ATC_X',     'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'LO_ATC_Z',     'MADLN15BE',             1,0,0,1,'AC', ''),
    ('',    'SIDE_X1',      'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'SIDE_Y1',      'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'SIDE_T1',      'MADLN05BE',             1,0,0,1,'AC', '50W'),
    ('',    'SIDE_X2',      'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'SIDE_Y2',      'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'SIDE_T2',      'MADLN05BE',             1,0,0,1,'AC', '50W'),
    ('',    'SIDE_X3',      'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'SIDE_Y3',      'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'SIDE_T3',      'MADLN05BE',             1,0,0,1,'AC', '50W'),
    ('',    'LO_EV',        'MCDLN35BE',             1,1,0,0,'AC', '800W 서보'),
    ('',    'LO_FD',        'MADLN05BE',             1,0,1,0,'AC', ''),
    ('',    'LO_FD_Z',      'MBDLN25BE',             1,0,0,1,'AC', ''),
    ('',    'LO_FD_ALIGN',  'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'LO_FD_INDEX',  'MBDLN25BE',             1,0,1,0,'AC', ''),
    ('',    'LO_EV_CONV',   'EZI-SPEED60H30CR100P',  1,0,0,0,'EZI','컨베이어 EziSpeed (DC24V)'),
    ('',    'ULO_EV_CONV',  'EZI-SPEED60H30CR100P',  1,0,0,0,'EZI',''),
    ('',    'NG_EV_LO_CONV','EZI-SPEED60H30CR100P',  1,0,0,0,'EZI',''),
    ('',    'NG_EV_ULO_CONV','EZI-SPEED60H30CR100P', 1,0,0,0,'EZI',''),
    ('',    'GOOD_EV_LO_CONV','EZI-SPEED60H30CR100P',1,0,0,0,'EZI',''),
    ('',    'GOOD_EV_ULO_CONV','EZI-SPEED60H30CR100P',1,0,0,0,'EZI',''),
    # === 판넬 4 ===
    ('판넬4','TOP_MAC_Y',   'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'TOP_MAC_X',    'MADLN15BE',             1,0,0,1,'AC', ''),
    ('',    'TOP_MAC_Z',    'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'TOP_MIC_Y',    'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'TOP_MIC_X',    'MADLN15BE',             1,0,0,1,'AC', ''),
    ('',    'TOP_MIC_Z',    'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'INSPEC_CONV1', 'MADLN15BE',             1,0,1,0,'AC', ''),
    ('',    'INSPEC_CONV2', 'MADLN15BE',             1,0,0,1,'AC', ''),
    ('',    'JIG_PLATE_CONV1','MADLN15BE',           1,1,0,0,'AC', ''),
    ('',    'JIG_PLATE_CONV2','MADLN15BE',           1,0,1,0,'AC', ''),
    ('',    'LO_JIG_TRANSFER_Y','MADLN05BE',         1,0,0,1,'AC', ''),
    ('',    'LO_JIG_TRANSFER_Z','MADLN05BE',         1,1,0,0,'AC', ''),
    ('',    'TURN_PLATE',   'MADLN05BE',             1,0,1,0,'AC', ''),
    ('',    'INSEPC_Y',     'MDDLN55BL',             3,1,1,1,'AC', 'Linear Motor Driver'),
    ('',    'INSPEC_X',     'MCDLN35BL',             3,0,1,1,'AC', 'Linear Motor Driver'),
    ('',    'INSPEC_Z1',    'MADLN05BE',             1,1,0,0,'AC', ''),
    ('',    'INSPEC_Z2',    'MADLN05BE',             1,0,1,0,'AC', ''),
    ('',    'INSPEC_R1',    'Ezi-EC-42XL-A',         1,0,0,0,'EZI','EziServo EC (DC24V)'),
    ('',    'INSEPC_R2',    'Ezi-EC-42XL-A',         1,0,0,0,'EZI',''),
    ('',    'TURN_GRIPPER_R1','Ezi-EC-42XL-A',       1,0,0,0,'EZI',''),
    ('',    'TURN_GRIPPER_R2','Ezi-EC-42XL-A',       1,0,0,0,'EZI',''),
    ('',    'TURN_R1',      'Ezi-EC-42M-A',          1,0,0,0,'EZI','EziServo EC 미니'),
    ('',    'TURN_R2',      'Ezi-EC-42M-A',          1,0,0,0,'EZI',''),
]

# DC24V 공통 부하 — 센서/I/O는 지건승님이 정리 후 추가 예정
DC_DATA = [
    ('공통', 'EziServo/센서/I/O DC24V', '(추후 입력)', 0, 0, 1, 0,0,0,'DC','센서·솔레노이드·I/O 정리 후 추가'),
]

MISC_DATA = [
    ('기타','HMI/PC/조명 등', '(추후 입력)', 0, 0, 1, 0,0,0,'ETC','추후 직접 입력'),
]

# ── 섹션 헤더 헬퍼 ──
def section_hdr(ws, r, label, color):
    ws.merge_cells(f'A{r}:Q{r}')
    c = ws[f'A{r}']
    c.value = label
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(color)
    c.alignment = left_a
    c.border = thin_b
    ws.row_dimensions[r].height = 16

def data_row_write(ws, r, panel, axis, model, phase, R, S, T,
                   pwr_type, note, row_fill):
    # C열: 드라이버 모델 (직접 입력)
    # D열: 출력W (VLOOKUP — 파나소닉 또는 파스텍)
    # E열: 정격전류A (VLOOKUP)
    # J열: 운전전류 = E × 0.85 (운전비율)

    c_col = 3   # 드라이버 모델
    d_col = 4   # 출력W
    e_col = 5   # 정격전류
    j_col = 10  # 운전전류

    # 전원구분 체크
    ac_flag  = '●' if pwr_type == 'AC'  else ''
    ezi_flag = '●' if pwr_type == 'EZI' else ''
    dc_flag  = '●' if pwr_type == 'DC'  else ''
    etc_flag = '●' if pwr_type == 'ETC' else ''

    vals = [panel, axis, model, '', '', phase, R, S, T, '',
            ac_flag, ezi_flag, dc_flag, etc_flag, '', '', note]

    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = mfont(size=9)
        c.alignment = center if i not in (1,2,3,17) else (left_a if i in (2,3,17) else center)
        c.border = thin_b
        if row_fill: c.fill = mfill(row_fill)
        if i == 1 and panel:
            c.font = mfont(bold=True, size=9, color=C_TITLE)

    # D열: 출력W VLOOKUP (파나소닉 → 파스텍 순)
    ws.cell(row=r, column=4).value = (
        f'=IFERROR(VLOOKUP(C{r},파나소닉A6규격!$A:$C,2,FALSE),'
        f'IFERROR(VLOOKUP(C{r},파스텍EziServo규격!$A:$C,1,FALSE),""))'
    )
    ws.cell(row=r, column=4).font = mfont(size=9, color='555555')
    ws.cell(row=r, column=4).alignment = center
    ws.cell(row=r, column=4).border = thin_b
    if row_fill: ws.cell(row=r, column=4).fill = mfill(row_fill)

    # E열: 정격전류A VLOOKUP
    ws.cell(row=r, column=5).value = (
        f'=IFERROR(VLOOKUP(C{r},파나소닉A6규격!$A:$C,3,FALSE),'
        f'IFERROR(VLOOKUP(C{r},파스텍EziServo규격!$A:$C,3,FALSE),""))'
    )
    ws.cell(row=r, column=5).font = mfont(size=9, color='1F497D')
    ws.cell(row=r, column=5).alignment = center
    ws.cell(row=r, column=5).border = thin_b
    ws.cell(row=r, column=5).fill = mfill('FFFF99')  # 자동계산 셀 강조

    # J열: 운전전류 = 정격전류 × 0.9
    ws.cell(row=r, column=10).value = f'=IFERROR(E{r}*0.9,"")'
    ws.cell(row=r, column=10).font = mfont(size=9)
    ws.cell(row=r, column=10).alignment = center
    ws.cell(row=r, column=10).border = thin_b
    if row_fill: ws.cell(row=r, column=10).fill = mfill(row_fill)

    ws.row_dimensions[r].height = 17

# 소계 행 헬퍼
def subtotal_row(ws, r, label, start, end, color):
    ws.merge_cells(f'A{r}:C{r}')
    ws.cell(row=r, column=1, value=f'  {label} 소계')
    ws.cell(row=r, column=1).font = mfont(bold=True, size=9)
    ws.cell(row=r, column=1).fill = mfill(color)
    ws.cell(row=r, column=1).alignment = left_a
    ws.cell(row=r, column=1).border = thin_b
    # E열: 정격전류 합계
    ws.cell(row=r, column=5).value = f'=SUM(E{start}:E{end})'
    ws.cell(row=r, column=5).number_format = '0.0'
    ws.cell(row=r, column=5).font = mfont(bold=True, size=9)
    ws.cell(row=r, column=5).fill = mfill(color)
    ws.cell(row=r, column=5).alignment = center
    ws.cell(row=r, column=5).border = thin_b
    # J열: 운전전류 합계
    ws.cell(row=r, column=10).value = f'=SUM(J{start}:J{end})'
    ws.cell(row=r, column=10).number_format = '0.0'
    ws.cell(row=r, column=10).font = mfont(bold=True, size=9)
    ws.cell(row=r, column=10).fill = mfill(color)
    ws.cell(row=r, column=10).alignment = center
    ws.cell(row=r, column=10).border = thin_b
    # O열: 권장 차단기 = E × 1.25
    ws.cell(row=r, column=15).value = f'=CEILING(E{r}*1.25,5)'
    ws.cell(row=r, column=15).font = mfont(bold=True, size=9, color='C00000')
    ws.cell(row=r, column=15).fill = mfill(color)
    ws.cell(row=r, column=15).alignment = center
    ws.cell(row=r, column=15).border = thin_b
    for col in [4,6,7,8,9,11,12,13,14,16]:
        ws.cell(row=r, column=col).fill = mfill(color)
        ws.cell(row=r, column=col).border = thin_b
    ws.cell(row=r, column=17).fill = mfill(color)
    ws.cell(row=r, column=17).border = thin_b
    ws.cell(row=r, column=17).value = 'E×1.25→CEILING 5A 단위'
    ws.cell(row=r, column=17).font = mfont(size=8, color='555555')
    ws.row_dimensions[r].height = 17
    return r

# ─── 본문 데이터 작성 ───
ROW = HEADER_ROW + 2

# 판넬별 섹션 분리
panels_order = ['판넬1', '판넬2', '판넬3', '판넬4']
panel_colors = {
    'AC':  {'판넬1':'EBF3FB','판넬2':'EBF3FB','판넬3':'EBF3FB','판넬4':'EBF3FB'},
    'EZI': {'판넬1':'FEF4E8','판넬2':'FEF4E8','판넬3':'FEF4E8','판넬4':'FEF4E8'},
}
section_colors = {'AC':'00B0F0','EZI':'ED7D31','DC':'70AD47','ETC':'7030A0'}
fill_colors    = {'AC':'EBF3FB','EZI':'FEF4E8','DC':'EDF7E3','ETC':'F3EBF9'}

panel_sum_rows = {}  # 판넬명 → 소계행 번호 (AC, EZI 각각)

for pidx, pname in enumerate(panels_order):
    rows_in_panel = [(i, d) for i, d in enumerate(PANEL_DATA) if d[0] == pname or (d[0]=='' and i>0)]

# 전체를 판넬별로 묶어 처리
current_panel = None
ac_start = None
ezi_start = None
ac_rows  = []
ezi_rows = []

for pname in panels_order:
    # 해당 판넬 데이터 추출
    in_panel = False
    p_ac_rows = []
    p_ezi_rows = []
    for d in PANEL_DATA:
        if d[0] == pname:
            in_panel = True
        if in_panel and d[0] != pname and d[0] != '':
            in_panel = False
        if in_panel:
            if d[7] == 'AC':
                p_ac_rows.append(d)
            elif d[7] == 'EZI':
                p_ezi_rows.append(d)

    # AC 서보 섹션
    if p_ac_rows:
        section_hdr(ws, ROW, f'■ {pname} — 파나소닉 A6 서보드라이버 (AC200V)', C_PAN)
        ROW += 1
        ac_s = ROW
        for d in p_ac_rows:
            panel_, axis_, model_, ph_, R_, S_, T_, ptype_, note_ = d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8]
            data_row_write(ws, ROW, panel_ if ROW == ac_s else '', axis_, model_,
                           ph_, R_, S_, T_, ptype_, note_, fill_colors['AC'])
            ROW += 1
        subtotal_row(ws, ROW, f'{pname} AC 소계', ac_s, ROW-1, C_SUM)
        panel_sum_rows.setdefault(pname, {})['ac_sum'] = ROW
        ROW += 2

    # EziServo/DC 섹션
    if p_ezi_rows:
        section_hdr(ws, ROW, f'■ {pname} — EziServo/EziSpeed (DC 24V)', C_EZI)
        ROW += 1
        ezi_s = ROW
        for d in p_ezi_rows:
            panel_, axis_, model_, ph_, R_, S_, T_, ptype_, note_ = d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8]
            data_row_write(ws, ROW, '', axis_, model_,
                           ph_, R_, S_, T_, ptype_, note_, fill_colors['EZI'])
            ROW += 1
        subtotal_row(ws, ROW, f'{pname} EziServo 소계', ezi_s, ROW-1, C_SUM)
        panel_sum_rows.setdefault(pname, {})['ezi_sum'] = ROW
        ROW += 2

# DC 24V 공통
section_hdr(ws, ROW, '■ DC 24V 공통 부하 (센서/I/O/솔레노이드) — 수량 직접 입력', C_DC)
ROW += 1
dc_s = ROW
for d in DC_DATA:
    panel_, axis_, model_, watt_, curr_, ph_, R_, S_, T_, ptype_, note_ = d
    data_row_write(ws, ROW, panel_, axis_, model_, ph_, R_, S_, T_, ptype_, note_, fill_colors['DC'])
    ws.cell(row=ROW, column=5).value = curr_ if curr_ else ''
    ws.cell(row=ROW, column=5).fill = mfill('FFFF99')
    ROW += 1
subtotal_row(ws, ROW, 'DC24V 공통 소계', dc_s, ROW-1, C_SUM)
dc_sum_row = ROW
ROW += 2

# 기타 부하
section_hdr(ws, ROW, '■ 기타 부하 (HMI / PC / 조명 등) — 전류 직접 입력', C_MISC)
ROW += 1
etc_s = ROW
for d in MISC_DATA:
    panel_, axis_, model_, watt_, curr_, ph_, R_, S_, T_, ptype_, note_ = d
    data_row_write(ws, ROW, panel_, axis_, model_, ph_, R_, S_, T_, ptype_, note_, fill_colors['ETC'])
    ws.cell(row=ROW, column=5).value = curr_
    ws.cell(row=ROW, column=5).fill = mfill('FFFF99')
    ROW += 1
subtotal_row(ws, ROW, '기타 소계', etc_s, ROW-1, C_SUM)
etc_sum_row = ROW
ROW += 2

# ── 수식 설명 ──
section_hdr(ws, ROW, '【 수식/VLOOKUP 설명 】', '595959')
ROW += 1
formula_notes = [
    '① C열 드라이버 모델 입력 → D열(출력W), E열(정격전류A) VLOOKUP 자동 계산 (파나소닉A6규격, 파스텍EziServo규격 시트 참조)',
    '② 운전전류(J열) = 정격전류(E) × 0.9  (90% 운전 가정)',
    '③ 소계행 O열 = 정격전류합계 × 1.25 → CEILING(5A 단위) = 차단기 권장 정격',
    '④ R/S/T 상 할당(G/H/I): 1=해당상 연결, 0=미연결. R/S/T 합계로 편중 분석',
    '⑤ Linear 드라이버(MCDLN35BL, MDDLN55BL): 스펙 전류 기입, 실측값으로 수정 권장',
    '⑥ DC24V 공통 부하 E열: 직접 입력 (VLOOKUP 제외)',
]
for note in formula_notes:
    ws.merge_cells(f'A{ROW}:Q{ROW}')
    ws.cell(row=ROW, column=1, value=note)
    ws.cell(row=ROW, column=1).font = mfont(size=9)
    ws.cell(row=ROW, column=1).fill = mfill('FFF2CC')
    ws.cell(row=ROW, column=1).alignment = left_a
    ws.row_dimensions[ROW].height = 14
    ROW += 1

# ── 저장 ──
out = 'C:/MES/wta-agents/workspaces/control-agent/대구텍_F2_1_판넬별_차단기선정_v2.xlsx'
wb.save(out)
print('저장 완료:', out)
