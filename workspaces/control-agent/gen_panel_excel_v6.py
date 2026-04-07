"""
대구텍 검사기 F2 #1 판넬별 전류·용량 계산 및 차단기 선정표 v6
- 형식: 판넬별 섹션 (판넬1→2→3→4→DC→기타→합계)
- 각 판넬 끝에 소계 + 차단기 선정행
- 카탈로그 정격전류 기준, 설계전류 = 정격×0.7
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 공통 스타일 =====
def mfont(bold=False, size=9, color='000000', name='맑은 고딕'):
    return Font(name=name, bold=bold, size=size, color=color)
def mfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)
def mborder(style='thin', color='AAAAAA'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def mcenter(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def mleft(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

C_HEADER   = '2F5496'
C_SUBHDR   = '4472C4'
C_TITLE    = '1F3864'
C_PANASONIC= '00B0F0'
C_FASTECH  = 'ED7D31'
C_DC       = '70AD47'
C_MISC     = '7030A0'
C_SUM      = 'E2EFDA'
C_BREAKER  = 'FFFF99'

thin_b = mborder('thin')
center = mcenter()
left_a = mleft()

# ===== SHEET 1: 판넬별 차단기 선정표 =====
ws = wb.active
ws.title = '판넬별 차단기 선정표'

col_widths = {1:5,2:20,3:24,4:8,5:10,6:6,7:5,8:9,9:9,10:9,11:9,12:9,13:9,14:9,15:9,16:9,17:32}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── 1행: 제목 ──
ws.merge_cells('A1:Q1')
c = ws['A1']
c.value = '대구텍 검사기 F2 #1 — 판넬별 전류·용량 계산 및 차단기 선정표'
c.font = Font(name='맑은 고딕', bold=True, size=14, color=C_TITLE)
c.alignment = center
c.fill = mfill('D9E2F3')
ws.row_dimensions[1].height = 28

# ── 2행: 기준 정보 ──
info = [('A2','전원:'),('B2','AC 220V / 3상'),('C2','역률(PF):'),('D2',0.85),
        ('E2','안전계수:'),('F2',1.25),('G2','수용률:'),('H2',0.7),('I2','작성일:'),('J2','2026-04-03')]
for addr, val in info:
    ws[addr] = val
    ws[addr].font = Font(name='맑은 고딕', size=9, bold=(addr[0] in 'ACEGI' and addr[1]=='2'))
for addr in ['B2','D2','F2','H2','J2']:
    ws[addr].fill = mfill('FFFF99')
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 5

# ── 4~5행: 컬럼 헤더 ──
for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
    ws.merge_cells(f'{col_letter}4:{col_letter}5')

for col, label in [(1,'판넬'),(2,'축/유닛명'),(3,'드라이버/장치 모델'),(4,'출력\n(W)'),
                   (5,'정격전류\n(A)'),(6,'상수'),(7,'R상\n할당'),(8,'S상\n할당'),
                   (9,'T상\n할당'),(10,'설계전류\n(A)')]:
    c = ws.cell(row=4, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b

ws.merge_cells('K4:N4')
ws['K4'] = '전원 구분'
ws['K4'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['K4'].fill = mfill(C_SUBHDR)
ws['K4'].alignment = center
ws['K4'].border = thin_b

for col, label in [(11,'AC 3상\n220V'),(12,'AC 단상\n220V'),(13,'DC 24V'),(14,'기타')]:
    c = ws.cell(row=5, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b

ws.merge_cells('O4:O5')
ws['O4'] = '권장\n차단기(A)'
ws['O4'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['O4'].fill = mfill(C_HEADER)
ws['O4'].alignment = center
ws['O4'].border = thin_b

ws.merge_cells('P4:P5')
ws['P4'] = '제조사/모델'
ws['P4'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['P4'].fill = mfill(C_HEADER)
ws['P4'].alignment = center
ws['P4'].border = thin_b

ws.merge_cells('Q4:Q5')
ws['Q4'] = '비고'
ws['Q4'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['Q4'].fill = mfill(C_HEADER)
ws['Q4'].alignment = left_a
ws['Q4'].border = thin_b

ws.row_dimensions[4].height = 20
ws.row_dimensions[5].height = 24

ROW = 6

# ===== 헬퍼 함수 =====
def vlookup_e(r):
    return (f'=IFERROR(VLOOKUP(C{r},파나소닉A6규격!$A:$C,3,FALSE),'
            f'IFERROR(VLOOKUP(C{r},파스텍EziServo규격!$A:$C,3,FALSE),""))')

def data_row(ws, r, panel, axis, model, watt, r_flag, s_flag, t_flag, util,
             ac3=False, ac1=False, dc=False, misc=False, note='', fill_color=None):
    fill = mfill(fill_color) if fill_color else None
    vals = [panel, axis, model, watt, '',  # E='' placeholder, set by set_e_j
            1, r_flag, s_flag, t_flag, '',  # J='' placeholder
            '●' if ac3 else '', '●' if ac1 else '',
            '●' if dc else '', '●' if misc else '',
            '', '', note]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = mfont(size=9)
        c.alignment = left_a if i in (2,3,17) else center
        c.border = thin_b
        if fill: c.fill = fill
        if i == 1: c.font = mfont(bold=True, size=9, color=C_TITLE)
    ws.row_dimensions[r].height = 17
    # VLOOKUP 수식 삽입
    ws[f'E{r}'] = vlookup_e(r)
    ws[f'E{r}'].font = mfont(size=9, color='1F3864')
    ws[f'E{r}'].alignment = center
    ws[f'E{r}'].border = thin_b
    if fill: ws[f'E{r}'].fill = fill
    ws[f'J{r}'] = f'=IF(E{r}="","",ROUND(E{r}*{util},3))'
    ws[f'J{r}'].font = mfont(size=9, color='375623')
    ws[f'J{r}'].alignment = center
    ws[f'J{r}'].border = thin_b
    if fill: ws[f'J{r}'].fill = fill

def data_row_fixed(ws, r, panel, axis, model, watt, rated_i, r_flag, s_flag, t_flag, op_i,
                   ac3=False, ac1=False, dc=False, misc=False, note='', fill_color=None):
    """VLOOKUP 없이 직접 값 입력 (DC 부하용)"""
    fill = mfill(fill_color) if fill_color else None
    vals = [panel, axis, model, watt, rated_i, 1, r_flag, s_flag, t_flag, op_i,
            '●' if ac3 else '', '●' if ac1 else '',
            '●' if dc else '', '●' if misc else '',
            '', '', note]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = mfont(size=9)
        c.alignment = left_a if i in (2,3,17) else center
        c.border = thin_b
        if fill: c.fill = fill
        if i == 1: c.font = mfont(bold=True, size=9, color=C_TITLE)
    ws.row_dimensions[r].height = 17

def section_header(ws, r, label, color):
    ws.merge_cells(f'A{r}:Q{r}')
    c = ws[f'A{r}']
    c.value = label
    c.font = mfont(bold=True, size=10, color='FFFFFF')
    c.fill = mfill(color)
    c.alignment = left_a
    c.border = thin_b
    ws.row_dimensions[r].height = 18

def panel_sum_row(ws, r, panel_name, start, end, breaker_a, mfr, model_str, note=''):
    """판넬 소계 + 차단기 선정 행"""
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'] = f'  ▶ {panel_name} 합계'
    ws[f'A{r}'].font = mfont(bold=True, size=9, color='FFFFFF')
    ws[f'A{r}'].fill = mfill(C_TITLE)
    ws[f'A{r}'].alignment = left_a
    ws[f'A{r}'].border = thin_b
    # E: 정격전류 합계
    ws[f'E{r}'] = f'=SUM(E{start}:E{end})'
    ws[f'E{r}'].font = mfont(bold=True, size=9)
    ws[f'E{r}'].number_format = '0.00'
    ws[f'E{r}'].alignment = center
    ws[f'E{r}'].border = thin_b
    ws[f'E{r}'].fill = mfill(C_SUM)
    # F blank
    for col in [6,7,8,9]:
        ws.cell(row=r, column=col).border = thin_b
        ws.cell(row=r, column=col).fill = mfill(C_SUM)
    # J: 설계전류 합계
    ws[f'J{r}'] = f'=SUM(J{start}:J{end})'
    ws[f'J{r}'].font = mfont(bold=True, size=9)
    ws[f'J{r}'].number_format = '0.00'
    ws[f'J{r}'].alignment = center
    ws[f'J{r}'].border = thin_b
    ws[f'J{r}'].fill = mfill(C_SUM)
    # K~N
    for col in range(11, 15):
        ws.cell(row=r, column=col).border = thin_b
        ws.cell(row=r, column=col).fill = mfill(C_SUM)
    # O: 권장차단기
    ws[f'O{r}'] = breaker_a
    ws[f'O{r}'].font = mfont(bold=True, size=10, color='1F3864')
    ws[f'O{r}'].alignment = center
    ws[f'O{r}'].border = thin_b
    ws[f'O{r}'].fill = mfill(C_BREAKER)
    # P: 제조사/모델
    ws[f'P{r}'] = f'{mfr} {model_str}'
    ws[f'P{r}'].font = mfont(bold=True, size=9)
    ws[f'P{r}'].alignment = center
    ws[f'P{r}'].border = thin_b
    ws[f'P{r}'].fill = mfill(C_BREAKER)
    # Q: 비고
    ws[f'Q{r}'] = note if note else f'=J{r}&"A×1.25="&TEXT(J{r}*1.25,"0.00")&"A → {breaker_a}A 선정"'
    ws[f'Q{r}'].font = mfont(size=9, color='555555')
    ws[f'Q{r}'].alignment = left_a
    ws[f'Q{r}'].border = thin_b
    ws[f'Q{r}'].fill = mfill(C_BREAKER)
    ws.row_dimensions[r].height = 18

# ============================================================
# 판넬 1  (AC 3상 파나소닉 13축)
# ============================================================
section_header(ws, ROW, '■ 판넬 1 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
p1_start = ROW
P1_COLOR = 'EBF3FB'

panel1_data = [
    # (축명, 모델, 출력W, R, S, T, 수용률)
    ('LO_Y1',       'MDDLN55BL', 1000, 1,1,1, 0.7),
    ('LO_X1',       'MCDLN35BL',  750, 1,1,1, 0.7),
    ('LO_Z1',       'MADLN05BE',   50, 1,1,1, 0.7),
    ('ULO_Y1',      'MCDLN35BL',  750, 1,1,1, 0.7),
    ('ULO_X1',      'MCDLN35BL',  750, 1,1,1, 0.7),
    ('ULO_Z1',      'MADLN05BE',   50, 1,1,1, 0.7),
    ('NG_EV',       'MCDLN35BE',  750, 1,1,1, 0.7),
    ('NG_FD',       'MADLN05BE',   50, 1,1,1, 0.7),
    ('NG_FD_Z',     'MBDLN25BE',  400, 1,1,1, 0.7),
    ('NG_FD_ALIGN', 'MADLN05BE',   50, 1,1,1, 0.7),
    ('NG_FD_INDEX', 'MBDLN25BE',  400, 1,1,1, 0.7),
    ('JIG_ULO_Y',   'MADLN05BE',   50, 1,1,1, 0.7),
    ('JIG_ULO_Z',   'MADLN05BE',   50, 1,1,1, 0.7),
]
for axis, model, watt, r, s, t, util in panel1_data:
    data_row(ws, ROW, '판넬1', axis, model, watt, r, s, t, util, ac3=True, fill_color=P1_COLOR)
    ROW += 1
p1_end = ROW - 1
panel_sum_row(ws, ROW, '판넬 1', p1_start, p1_end, 20, 'LS', 'ABN20c')
ROW += 2

# ============================================================
# 판넬 2  (AC 3상 파나소닉 13축)
# ============================================================
section_header(ws, ROW, '■ 판넬 2 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
p2_start = ROW
P2_COLOR = 'DEEAF1'

panel2_data = [
    ('GOOD_EV',         'MCDLN35BE',  750, 1,1,1, 0.7),
    ('GOOD_FD',         'MADLN05BE',   50, 1,1,1, 0.7),
    ('GOOD_FD_Z',       'MBDLN25BE',  400, 1,1,1, 0.7),
    ('GOOD_FD_ALIGN',   'MADLN05BE',   50, 1,1,1, 0.7),
    ('GOOD_FD_INDEX',   'MBDLN25BE',  400, 1,1,1, 0.7),
    ('TURN_Y',          'MBDLN25BE',  400, 1,1,1, 0.7),
    ('TURN_GRIPPER_Y1', 'MADLN15BE',  200, 1,1,1, 0.7),
    ('TURN_GRIPPER_X1', 'MADLN15BE',  200, 1,1,1, 0.7),
    ('TURN_GRIPPER_Z1', 'MADLN05BE',   50, 1,1,1, 0.7),
    ('TURN_GRIPPER_Y2', 'MADLN15BE',  200, 1,1,1, 0.7),
    ('TURN_GRIPPER_X2', 'MADLN15BE',  200, 1,1,1, 0.7),
    ('TURN_GRIPPER_Z2', 'MADLN05BE',   50, 1,1,1, 0.7),
    ('BOT_MAC_Y',       'MADLN15BE',  200, 1,1,1, 0.7),
]
for axis, model, watt, r, s, t, util in panel2_data:
    data_row(ws, ROW, '판넬2', axis, model, watt, r, s, t, util, ac3=True, fill_color=P2_COLOR)
    ROW += 1
p2_end = ROW - 1
panel_sum_row(ws, ROW, '판넬 2', p2_start, p2_end, 15, 'LS', 'ABN15c')
ROW += 2

# ============================================================
# 판넬 3  (AC 3상 파나소닉 21축)
# ============================================================
section_header(ws, ROW, '■ 판넬 3 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
p3_start = ROW
P3_COLOR = 'EBF3FB'

panel3_data = [
    ('BOT_MAC_X',   'MADLN15BE',  200, 1,1,1, 0.7),
    ('BOT_MAC_Z',   'MADLN05BE',   50, 1,1,1, 0.7),
    ('BOT_MIC_Y',   'MADLN15BE',  200, 1,1,1, 0.7),
    ('BOT_MIC_X',   'MADLN15BE',  200, 1,1,1, 0.7),
    ('BOT_MIC_Z',   'MADLN05BE',   50, 1,1,1, 0.7),
    ('LO_ATC_X',    'MADLN15BE',  200, 1,1,1, 0.7),
    ('LO_ATC_Z',    'MADLN15BE',  200, 1,1,1, 0.7),
    ('SIDE_X1',     'MADLN05BE',   50, 1,1,1, 0.7),
    ('SIDE_Y1',     'MADLN15BE',  200, 1,1,1, 0.7),
    ('SIDE_T1',     'MADLN05BE',   50, 1,1,1, 0.7),
    ('SIDE_X2',     'MADLN05BE',   50, 1,1,1, 0.7),
    ('SIDE_Y2',     'MADLN15BE',  200, 1,1,1, 0.7),
    ('SIDE_T2',     'MADLN05BE',   50, 1,1,1, 0.7),
    ('SIDE_X3',     'MADLN05BE',   50, 1,1,1, 0.7),
    ('SIDE_Y3',     'MADLN15BE',  200, 1,1,1, 0.7),
    ('SIDE_T3',     'MADLN05BE',   50, 1,1,1, 0.7),
    ('LO_EV',       'MCDLN35BE',  750, 1,1,1, 0.7),
    ('LO_FD',       'MADLN05BE',   50, 1,1,1, 0.7),
    ('LO_FD_Z',     'MBDLN25BE',  400, 1,1,1, 0.7),
    ('LO_FD_ALIGN', 'MADLN05BE',   50, 1,1,1, 0.7),
    ('LO_FD_INDEX', 'MBDLN25BE',  400, 1,1,1, 0.7),
]
for axis, model, watt, r, s, t, util in panel3_data:
    data_row(ws, ROW, '판넬3', axis, model, watt, r, s, t, util, ac3=True, fill_color=P3_COLOR)
    ROW += 1
p3_end = ROW - 1
panel_sum_row(ws, ROW, '판넬 3', p3_start, p3_end, 20, 'LS', 'ABN20c')
ROW += 2

# ============================================================
# 판넬 4  (AC 3상 파나소닉 17축)
# ============================================================
section_header(ws, ROW, '■ 판넬 4 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
p4_start = ROW
P4_COLOR = 'DEEAF1'

panel4_data = [
    ('TOP_MAC_Y',       'MADLN15BE',  200, 1,1,1, 0.7),
    ('TOP_MAC_X',       'MADLN15BE',  200, 1,1,1, 0.7),
    ('TOP_MAC_Z',       'MADLN05BE',   50, 1,1,1, 0.7),
    ('TOP_MIC_Y',       'MADLN15BE',  200, 1,1,1, 0.7),
    ('TOP_MIC_X',       'MADLN15BE',  200, 1,1,1, 0.7),
    ('TOP_MIC_Z',       'MADLN05BE',   50, 1,1,1, 0.7),
    ('INSPEC_CONV1',    'MADLN15BE',  200, 1,1,1, 0.7),
    ('INSPEC_CONV2',    'MADLN15BE',  200, 1,1,1, 0.7),
    ('JIG_PLATE_CONV1', 'MADLN15BE',  200, 1,1,1, 0.7),
    ('JIG_PLATE_CONV2', 'MADLN15BE',  200, 1,1,1, 0.7),
    ('LO_JIG_TRANS_Y',  'MADLN05BE',   50, 1,1,1, 0.7),
    ('LO_JIG_TRANS_Z',  'MADLN05BE',   50, 1,1,1, 0.7),
    ('TURN_PLATE',      'MADLN05BE',   50, 1,1,1, 0.7),
    ('INSEPC_Y',        'MDDLN55BL', 1000, 1,1,1, 0.7),
    ('INSPEC_X',        'MCDLN35BL',  750, 1,1,1, 0.7),
    ('INSPEC_Z1',       'MADLN05BE',   50, 1,1,1, 0.7),
    ('INSPEC_Z2',       'MADLN05BE',   50, 1,1,1, 0.7),
]
for axis, model, watt, r, s, t, util in panel4_data:
    data_row(ws, ROW, '판넬4', axis, model, watt, r, s, t, util, ac3=True, fill_color=P4_COLOR)
    ROW += 1
p4_end = ROW - 1
panel_sum_row(ws, ROW, '판넬 4', p4_start, p4_end, 20, 'LS', 'ABN20c')
ROW += 2

# ============================================================
# DC 24V 부하 — EZS2 스텝서보 + I/O·Brake·센서·솔레노이드·EZI-SPEED
# ============================================================
section_header(ws, ROW, '■ DC 24V 부하 — EZS2 스텝서보 · 서보I/O · Brake · 센서 · 솔레노이드 · EZI-SPEED', C_DC)
ROW += 1
DC_COLOR = 'EDF7E3'

dc_data = [
    # EZS2 스텝서보 (판넬별)
    ('판넬1', 'LO_R1',          'EZS2-EC-42M',           0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬1 LO 회전축'),
    ('판넬1', 'ULO_R2',         'EZS2-EC-42M',           0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬1 ULO 회전축'),
    ('판넬2', 'TURN_R1',        'EZS2-EC-42M',           0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬2 TURN 회전1'),
    ('판넬2', 'TURN_R2',        'EZS2-EC-42M',           0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬2 TURN 회전2'),
    ('판넬2', 'TURN_GRIPPER_R1','EZS2-EC-42XL',          0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬2 그리퍼R1'),
    ('판넬2', 'TURN_GRIPPER_R2','EZS2-EC-42XL',          0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬2 그리퍼R2'),
    ('판넬4', 'INSPEC_R1',      'EZS2-EC-42M',           0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬4 INSPEC 회전1'),
    ('판넬4', 'INSPEC_R2',      'EZS2-EC-42M',           0, 1.2, 0,0,0, 1.2, False,False,True,  '판넬4 INSPEC 회전2'),
    # 서보 컨트롤러 전원 (1P24/1N24)
    ('공통-1P', '서보I/O (A6B)', 'A6B I/O',              0, 0.64,0,0,0, 0.64,False,False,True,  '64축×10mA=640mA'),
    ('공통-1P', '서보Brake',     'Brake',                 0, 9.0, 0,0,0, 9.0, False,False,True,  '18개×500mA=9000mA'),
    ('공통-1P', '리밋센서',      '리밋센서',               0, 0.64,0,0,0, 0.64,False,False,True,  '64개×10mA=640mA'),
    ('공통-1P', '호밍센서',      '호밍센서',               0, 0.06,0,0,0, 0.06,False,False,True,  '6개×10mA=60mA'),
    # I/O 전원 (2P24/2N24)
    ('공통-2P', '도어',          '도어스위치',             0, 3.0, 0,0,0, 3.0, False,False,True,  '10개×300mA=3000mA'),
    ('공통-2P', 'I/O 모듈',      '파스텍 IO단자대',        0, 2.34,0,0,0, 2.34,False,False,True,  '18개×130mA=2340mA'),
    ('공통-2P', '입력부하(센서)','포토/근접센서',           0, 2.5, 0,0,0, 2.5, False,False,True,  '250개×10mA=2500mA'),
    ('공통-2P', '출력부하',      '릴레이·솔레노이드',       0, 5.4, 0,0,0, 5.4, False,False,True,  '180개×30mA=5400mA'),
    # EZI-SPEED (3P24/3N24)
    ('공통-3P', 'LO_EV_TOP',    'EZI-SPEED60H30CR100P',  0, 1.0, 0,0,0, 1.0, False,False,True,  'LO이젝터 상부'),
    ('공통-3P', 'LO_EV_BOT',    'EZI-SPEED60H30CR100P',  0, 1.0, 0,0,0, 1.0, False,False,True,  'LO이젝터 하부'),
    ('공통-3P', 'GOOD_EV_TOP',  'EZI-SPEED60H30CR100P',  0, 1.0, 0,0,0, 1.0, False,False,True,  'GOOD이젝터 상부'),
    ('공통-3P', 'GOOD_EV_BOT',  'EZI-SPEED60H30CR100P',  0, 1.0, 0,0,0, 1.0, False,False,True,  'GOOD이젝터 하부'),
    ('공통-3P', 'NG_EV_TOP',    'EZI-SPEED60H30CR100P',  0, 1.0, 0,0,0, 1.0, False,False,True,  'NG이젝터 상부'),
    ('공통-3P', 'NG_EV_BOT',    'EZI-SPEED60H30CR100P',  0, 1.0, 0,0,0, 1.0, False,False,True,  'NG이젝터 하부'),
]
dc_start = ROW
for d in dc_data:
    panel, axis, model, watt, ri, r, s, t, oi, ac3, ac1, dc, note = d
    data_row_fixed(ws, ROW, panel, axis, model, watt, ri, r, s, t, oi,
                   ac3=ac3, ac1=ac1, dc=dc, note=note, fill_color=DC_COLOR)
    ROW += 1
dc_end = ROW - 1
# DC 소계
ws.merge_cells(f'A{ROW}:D{ROW}')
ws[f'A{ROW}'] = '  ▶ DC 24V 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill(C_DC)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{dc_start}:E{dc_end})'
ws[f'E{ROW}'].font = mfont(bold=True, size=9)
ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center
ws[f'E{ROW}'].border = thin_b
ws[f'E{ROW}'].fill = mfill(C_SUM)
ws[f'J{ROW}'] = f'=SUM(J{dc_start}:J{dc_end})'
ws[f'J{ROW}'].font = mfont(bold=True, size=9)
ws[f'J{ROW}'].number_format = '0.00'
ws[f'J{ROW}'].alignment = center
ws[f'J{ROW}'].border = thin_b
ws[f'J{ROW}'].fill = mfill(C_SUM)
for col in [6,7,8,9,11,12,13,14,15,16]:
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
ws[f'Q{ROW}'] = 'DC 전원장치: EZS2(9.6A)+1P(10.34A)+2P(13.24A)+3P(6.0A) → SMPS 각각 선정'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].alignment = left_a
ws[f'Q{ROW}'].border = thin_b
ws[f'Q{ROW}'].fill = mfill(C_SUM)
ws.row_dimensions[ROW].height = 17
dc_sum_row = ROW
ROW += 2

# ============================================================
# 기타 부하 — UPS · 조명CTRL · SMPS (AC 단상 220V)
# ============================================================
section_header(ws, ROW, '■ 기타 부하 — UPS · 조명CTRL · SMPS (AC 단상 220V)', C_MISC)
ROW += 1
MISC_COLOR = 'F3EBF9'

misc_data = [
    ('공통', 'UPS 950VA',    'BX950MI-GR',           0, 1.5, 0,0,0, 1.5, False,True, False, 'AC단상 220V — 1대'),
    ('공통', 'UPS 1600VA',   'BX1600MI-GR',          0, 4.5, 0,0,0, 4.5, False,True, False, 'AC단상 220V — 3대 (각각 별도 차단기)'),
    ('공통', '조명CTRL',     'POD-22024-4-PEI-LKCB', 0, 5.0, 0,1,0, 5.0, False,True, False, '10개×500mA=5.0A → 10A MCB'),
    ('공통', 'SMPS 24V 20A', 'PRO ECO 480W 24V 20A', 0, 2.4, 0,0,1, 2.4, False,True, False, '2대×2.4A'),
    ('공통', 'SMPS 24V 10A', 'PRO ECO 480W 24V 10A', 0, 1.2, 1,0,0, 1.2, False,True, False, '1대×1.2A'),
    ('공통', 'HMI/PC',       '(직접 입력)',           0, 0.0, 0,0,0, 0.0, False,True, False, '← 실측값 입력'),
    ('공통', '기타',          '(직접 입력)',           0, 0.0, 0,0,0, 0.0, False,False,False, '← 실측값 입력'),
]
misc_start = ROW
for d in misc_data:
    panel, axis, model, watt, ri, r, s, t, oi, ac3, ac1, dc, note = d
    data_row_fixed(ws, ROW, panel, axis, model, watt, ri, r, s, t, oi,
                   ac3=ac3, ac1=ac1, dc=dc, note=note, fill_color=MISC_COLOR)
    ROW += 1
misc_end = ROW - 1
ws.merge_cells(f'A{ROW}:D{ROW}')
ws[f'A{ROW}'] = '  ▶ 기타 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill(C_MISC)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{misc_start}:E{misc_end})'
ws[f'E{ROW}'].font = mfont(bold=True, size=9)
ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center
ws[f'E{ROW}'].border = thin_b
ws[f'E{ROW}'].fill = mfill(C_SUM)
ws[f'J{ROW}'] = f'=SUM(J{misc_start}:J{misc_end})'
ws[f'J{ROW}'].font = mfont(bold=True, size=9)
ws[f'J{ROW}'].number_format = '0.00'
ws[f'J{ROW}'].alignment = center
ws[f'J{ROW}'].border = thin_b
ws[f'J{ROW}'].fill = mfill(C_SUM)
for col in [6,7,8,9,11,12,13,14,15,16]:
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
ws[f'Q{ROW}'] = '단상 부하 — UPS는 대당 별도 차단기, 조명CTRL 10A MCB 권장'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].alignment = left_a
ws[f'Q{ROW}'].border = thin_b
ws[f'Q{ROW}'].fill = mfill(C_SUM)
ws.row_dimensions[ROW].height = 17
misc_sum_row = ROW
ROW += 2

# ============================================================
# 전체 합계 (AC 3상)
# ============================================================
section_header(ws, ROW, '■ 전체 합계 — AC 3상 220V (판넬1+2+3+4 합산)', C_TITLE)
ROW += 1

all_ac_rows = [p1_end+1, p2_end+1, p3_end+1, p4_end+1]  # sum rows indices
sum_e = '+'.join([f'E{p1_end+1}', f'E{p2_end+1}', f'E{p3_end+1}', f'E{p4_end+1}'])
sum_j = '+'.join([f'J{p1_end+1}', f'J{p2_end+1}', f'J{p3_end+1}', f'J{p4_end+1}'])

ws.merge_cells(f'A{ROW}:D{ROW}')
ws[f'A{ROW}'] = '  ▶ AC 3상 전체 합계 (설계전류 기준)'
ws[f'A{ROW}'].font = mfont(bold=True, size=10, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill(C_TITLE)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'={sum_e}'
ws[f'E{ROW}'].font = mfont(bold=True, size=10)
ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center
ws[f'E{ROW}'].border = thin_b
ws[f'E{ROW}'].fill = mfill(C_SUM)
ws[f'J{ROW}'] = f'={sum_j}'
ws[f'J{ROW}'].font = mfont(bold=True, size=10)
ws[f'J{ROW}'].number_format = '0.00'
ws[f'J{ROW}'].alignment = center
ws[f'J{ROW}'].border = thin_b
ws[f'J{ROW}'].fill = mfill(C_SUM)
for col in [6,7,8,9,11,12,13,14]:
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
ws[f'O{ROW}'] = '→ 메인 차단기'
ws[f'O{ROW}'].font = mfont(bold=True, size=10, color='1F3864')
ws[f'O{ROW}'].alignment = center
ws[f'O{ROW}'].border = thin_b
ws[f'O{ROW}'].fill = mfill(C_BREAKER)
ws[f'P{ROW}'] = 'LS ABN75c'
ws[f'P{ROW}'].font = mfont(bold=True, size=9)
ws[f'P{ROW}'].alignment = center
ws[f'P{ROW}'].border = thin_b
ws[f'P{ROW}'].fill = mfill(C_BREAKER)
ws[f'Q{ROW}'] = f'=J{ROW}&"A × 1.25 = "&TEXT(J{ROW}*1.25,"0.00")&"A → 메인 MCCB 선정"'
ws[f'Q{ROW}'].font = mfont(size=9, color='555555')
ws[f'Q{ROW}'].alignment = left_a
ws[f'Q{ROW}'].border = thin_b
ws[f'Q{ROW}'].fill = mfill(C_BREAKER)
ws.row_dimensions[ROW].height = 20

# ============================================================
# SHEET 2: 파나소닉A6규격 (VLOOKUP 소스)
# ============================================================
ws2 = wb.create_sheet('파나소닉A6규격')
ws2.merge_cells('A1:D1')
ws2['A1'] = '파나소닉 A6 서보드라이버 정격입력전류 (카탈로그 기준, 3상 AC200V)'
ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws2['A1'].alignment = center
ws2['A1'].fill = mfill('D6E4F0')
ws2.row_dimensions[1].height = 22

for j, h in enumerate(['드라이버 모델','정격출력 (W)','정격입력전류 (A)','비고'], 1):
    c = ws2.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws2.row_dimensions[2].height = 20

pan_lookup = [
    ('MADLN05BE',   50,  0.3, '50W 서보드라이버'),
    ('MADLN10BE',  100,  0.5, '100W 서보드라이버'),
    ('MADLN15BE',  200,  1.0, '200W 서보드라이버'),
    ('MBDLN25BE',  400,  1.5, '400W 서보드라이버'),
    ('MCDLN35BL',  750,  2.4, '750W 서보드라이버 (BL)'),
    ('MCDLN35BE',  750,  2.4, '750W 서보드라이버 (BE)'),
    ('MDDLN55BL', 1000,  3.4, '1kW 서보드라이버 (BL)'),
    ('MDDLN55BE', 1000,  3.4, '1kW 서보드라이버 (BE)'),
    ('MDDLN85BL', 1500,  5.5, '1.5kW 서보드라이버 (BL)'),
]
for i, (model, watt, curr, note) in enumerate(pan_lookup, 3):
    fill = mfill('EBF3FB') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate([model, watt, curr, note], 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j <= 3 else left_a
        c.border = thin_b
        c.fill = fill
    ws2.row_dimensions[i].height = 16

note_r = len(pan_lookup) + 4
ws2.merge_cells(f'A{note_r}:D{note_r}')
ws2[f'A{note_r}'] = '※ A열 모델명과 메인시트 C열 코드 정확히 일치 필요  |  카탈로그 정격입력전류 기준  |  모델 추가 시 이 시트에 행 추가'
ws2[f'A{note_r}'].font = mfont(size=8, color='555555')
ws2[f'A{note_r}'].fill = mfill('FFF2CC')
ws2[f'A{note_r}'].alignment = left_a

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 26

# ============================================================
# SHEET 3: 파스텍EziServo규격 (VLOOKUP 소스)
# ============================================================
ws3 = wb.create_sheet('파스텍EziServo규격')
ws3.merge_cells('A1:C1')
ws3['A1'] = '파스텍 EZS2/EziServo 정격전류 (VLOOKUP 소스)'
ws3['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws3['A1'].alignment = center
ws3['A1'].fill = mfill('FCE4D6')
ws3.row_dimensions[1].height = 22

for j, h in enumerate(['드라이버 모델','정격전류 (A)','비고'], 1):
    c = ws3.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws3.row_dimensions[2].height = 18

fas_lookup = [
    ('EZS2-EC-42M',          1.2, 'DC 24V 입력 스텝서보'),
    ('EZS2-EC-42XL',         1.2, 'DC 24V 입력 스텝서보 (XL)'),
    ('Ezi-EC-42M-A',         1.5, 'AC 입력 스텝서보'),
    ('Ezi-EC-42XL-A',        2.0, 'AC 입력 스텝서보 (XL)'),
    ('EZI-SPEED60H30CR100P', 1.0, 'DC 24V 입력 스텝모터드라이버'),
]
for i, (model, curr, note) in enumerate(fas_lookup, 3):
    fill = mfill('FEF4E8') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate([model, curr, note], 1):
        c = ws3.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j <= 2 else left_a
        c.border = thin_b
        c.fill = fill
    ws3.row_dimensions[i].height = 16

ws3.column_dimensions['A'].width = 26
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 28

# ============================================================
# SHEET 4: 차단기 규격 참조표
# ============================================================
ws4 = wb.create_sheet('차단기 규격 참조표')
ws4.merge_cells('A1:E1')
ws4['A1'] = '차단기 규격 참조표 (MCCB / MCB 3P, AC 220V)'
ws4['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws4['A1'].alignment = center
ws4.row_dimensions[1].height = 22
ws4.row_dimensions[2].height = 5

for j, h in enumerate(['정격전류 (A)','적용 부하전류 범위 (A)','제조사','모델 (3P)','비고'], 1):
    c = ws4.cell(row=3, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws4.row_dimensions[3].height = 28

breaker_data = [
    [3,  '~2.4A',  'LS/Chint', 'ABN3c / NXB-63',  '조명, 소형 제어회로'],
    [5,  '2.4~4A', 'LS/Chint', 'ABN5c / NXB-63',  '소형 I/O, 팬'],
    [10, '4~8A',   'LS/Chint', 'ABN10c / NXB-63', '소형 서보 1~2축'],
    [15, '8~12A',  'LS/Chint', 'ABN15c / NXB-63', '판넬2 권장'],
    [20, '12~16A', 'LS/Chint', 'ABN20c / NXB-63', '판넬1/3/4 권장'],
    [25, '16~20A', 'LS/Chint', 'ABN25c / NXB-63', ''],
    [30, '20~24A', 'LS/Chint', 'ABN30c / NXB-63', '다축 서보 판넬'],
    [40, '24~32A', 'LS/Chint', 'ABN40c',           ''],
    [50, '32~40A', 'LS/Chint', 'ABN50c',           ''],
    [63, '40~50A', 'LS/Chint', 'ABN63c',           '대형 판넬'],
    [75, '50~60A', 'LS/Chint', 'ABN75c',           '메인 차단기 (판넬 합산)'],
]
for i, row in enumerate(breaker_data, 4):
    ws4.row_dimensions[i].height = 16
    for j, val in enumerate(row, 1):
        c = ws4.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j <= 4 else left_a
        c.border = thin_b
        if j == 1: c.fill = mfill(C_SUM)

ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 20
ws4.column_dimensions['E'].width = 22

out_path = 'C:/MES/wta-agents/workspaces/control-agent/대구텍_F2_1_판넬별_차단기선정표_v3.xlsx'
wb.save(out_path)
print('저장 완료:', out_path)
print(f'판넬1({len(panel1_data)}축) / 판넬2({len(panel2_data)}축) / 판넬3({len(panel3_data)}축) / 판넬4({len(panel4_data)}축)')
print(f'DC부하({len(dc_data)}항목) / 기타({len(misc_data)}항목)')
