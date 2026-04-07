import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# 스타일 정의
header_font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
sub_header_fill = PatternFill('solid', fgColor='4472C4')
sub_header_font = Font(name='맑은 고딕', bold=True, size=9, color='FFFFFF')
yellow_fill = PatternFill('solid', fgColor='FFFF99')
green_fill = PatternFill('solid', fgColor='E2EFDA')
blue_fill = PatternFill('solid', fgColor='D6E4F0')
thin = Side(style='thin', color='AAAAAA')
medium = Side(style='medium', color='555555')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== Sheet1: 판넬별 차단기 선정 =====
ws = wb.active
ws.title = '판넬별 차단기 선정'

# 제목
ws.merge_cells('A1:O1')
ws['A1'] = '판넬별 전류 용량 계산 및 차단기 선정표'
ws['A1'].font = Font(name='맑은 고딕', bold=True, size=13, color='1F3864')
ws['A1'].alignment = center
ws.row_dimensions[1].height = 25

# 공통 정보 행
info = [('A2','전원:'),('B2','AC 220V / 3상'),('C2','역률(PF):'),('D2',0.85),
        ('E2','안전계수:'),('F2',1.25),('G2','작성일:'),('H2','2026-04-03')]
for addr, val in info:
    ws[addr] = val
    ws[addr].font = Font(name='맑은 고딕', size=9, bold=(addr in ['A2','C2','E2','G2']))
for addr in ['B2','D2','F2','H2']:
    ws[addr].fill = yellow_fill

ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 5

# 헤더 행 (4~5행)
# A~G: 개별 축 정보 (세로 병합)
main_cols = ['A','B','C','D','E','F','G']
main_labels = ['판넬명','축 번호','모터명 / 드라이버','정격출력\n(W)','정격전류\n(A)','운전전류\n(A)','비고']
for col, label in zip(main_cols, main_labels):
    ws.merge_cells(f'{col}4:{col}5')
    ws[f'{col}4'] = label
    ws[f'{col}4'].font = header_font
    ws[f'{col}4'].fill = header_fill
    ws[f'{col}4'].alignment = center
    ws[f'{col}4'].border = thin_border

# H4~O4: 판넬 차단기 선정 (가로 병합)
ws.merge_cells('H4:O4')
ws['H4'] = '판넬 차단기 선정'
ws['H4'].font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
ws['H4'].fill = PatternFill('solid', fgColor='1F3864')
ws['H4'].alignment = center
ws['H4'].border = thin_border

sub_cols = ['H','I','J','K','L','M','N','O']
sub_labels = ['총 운전전류\n(A)','안전계수\n(×1.25)','권장\n차단기(A)','선정\n차단기(A)','제조사','모델명','비고','']
for col, label in zip(sub_cols, sub_labels):
    ws[f'{col}5'] = label
    ws[f'{col}5'].font = sub_header_font
    ws[f'{col}5'].fill = sub_header_fill
    ws[f'{col}5'].alignment = center
    ws[f'{col}5'].border = thin_border

ws.row_dimensions[4].height = 18
ws.row_dimensions[5].height = 30

# ===== 판넬 1 샘플 데이터 =====
sample = [
    ['판넬 1\n(예시)', '축1', '서보모터 (Panasonic MHMF042L1S)', 400, 2.1, 1.8, ''],
    ['', '축2', '서보모터 (Panasonic MHMF082L1S)', 750, 4.1, 3.5, ''],
    ['', '축3', '스텝모터 (Leadshine 57HS)', 150, 2.8, 2.2, 'DC 24V 별도'],
    ['', '기타', 'I/O 모듈, HMI, 팬 등', '', '', 2.0, '공통부하'],
]
start = 6

for i, row in enumerate(sample):
    r = start + i
    ws.row_dimensions[r].height = 18
    for j, val in enumerate(row):
        col = get_column_letter(j+1)
        ws[f'{col}{r}'] = val
        ws[f'{col}{r}'].font = Font(name='맑은 고딕', size=9)
        ws[f'{col}{r}'].alignment = center if j in [0,1,3,4,5] else left_align
        ws[f'{col}{r}'].border = thin_border
        if col == 'A':
            ws[f'{col}{r}'].fill = blue_fill

# 판넬 1 합계 행
sr1 = start + len(sample)
ws.row_dimensions[sr1].height = 18
ws.merge_cells(f'A{sr1}:G{sr1}')
ws[f'A{sr1}'] = '▶ 판넬 1 합계'
ws[f'A{sr1}'].font = Font(name='맑은 고딕', bold=True, size=9)
ws[f'A{sr1}'].alignment = center
ws[f'A{sr1}'].fill = green_fill
ws[f'A{sr1}'].border = thin_border

# H: 합계, I: ×1.25, J: 권장, K: 선정, L: 제조사, M: 모델
ws[f'H{sr1}'] = f'=SUM(F{start}:F{sr1-1})'
ws[f'I{sr1}'] = f'=H{sr1}*1.25'
ws[f'J{sr1}'] = '30A (권장)'
ws[f'K{sr1}'] = 30
ws[f'L{sr1}'] = 'LS/Chint'
ws[f'M{sr1}'] = 'ABN30c'
ws[f'N{sr1}'] = ''
for col in ['H','I','J','K','L','M','N','O']:
    ws[f'{col}{sr1}'].font = Font(name='맑은 고딕', bold=True, size=9)
    ws[f'{col}{sr1}'].alignment = center
    ws[f'{col}{sr1}'].fill = green_fill
    ws[f'{col}{sr1}'].border = thin_border

ws.row_dimensions[sr1+1].height = 5

# ===== 판넬 2 빈 템플릿 =====
p2_start = sr1 + 2
for i in range(6):
    r = p2_start + i
    ws.row_dimensions[r].height = 18
    ws[f'A{r}'] = '판넬 2' if i == 0 else ''
    ws[f'B{r}'] = f'축{i+1}' if i < 5 else '기타'
    for col in main_cols:
        ws[f'{col}{r}'].border = thin_border
        ws[f'{col}{r}'].font = Font(name='맑은 고딕', size=9)
        ws[f'{col}{r}'].alignment = center
        if col == 'A':
            ws[f'{col}{r}'].fill = blue_fill

sr2 = p2_start + 6
ws.row_dimensions[sr2].height = 18
ws.merge_cells(f'A{sr2}:G{sr2}')
ws[f'A{sr2}'] = '▶ 판넬 2 합계'
ws[f'A{sr2}'].font = Font(name='맑은 고딕', bold=True, size=9)
ws[f'A{sr2}'].alignment = center
ws[f'A{sr2}'].fill = green_fill
ws[f'A{sr2}'].border = thin_border
ws[f'H{sr2}'] = f'=SUM(F{p2_start}:F{sr2-1})'
ws[f'I{sr2}'] = f'=H{sr2}*1.25'
for col in ['H','I','J','K','L','M','N','O']:
    ws[f'{col}{sr2}'].font = Font(name='맑은 고딕', bold=True, size=9)
    ws[f'{col}{sr2}'].alignment = center
    ws[f'{col}{sr2}'].fill = green_fill
    ws[f'{col}{sr2}'].border = thin_border

# ===== 컬럼 너비 =====
col_widths = {
    'A':12,'B':8,'C':30,'D':10,'E':10,'F':10,'G':14,
    'H':12,'I':12,'J':12,'K':12,'L':10,'M':12,'N':10,'O':10
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ===== Sheet2: 차단기 규격 참조표 =====
ws2 = wb.create_sheet('차단기 규격 참조표')
ws2.merge_cells('A1:E1')
ws2['A1'] = '차단기 규격 참조표 (MCCB / MCB 3P, AC 220V)'
ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color='1F3864')
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 22
ws2.row_dimensions[2].height = 5

headers2 = ['정격전류\n(A)', '적용 부하전류\n범위(A)', '제조사', '모델 (3P)', '비고']
for j, h in enumerate(headers2):
    col = get_column_letter(j+1)
    ws2[f'{col}3'] = h
    ws2[f'{col}3'].font = header_font
    ws2[f'{col}3'].fill = header_fill
    ws2[f'{col}3'].alignment = center
    ws2[f'{col}3'].border = thin_border
ws2.row_dimensions[3].height = 28

breaker_data = [
    [3,  '~2.4A',  'LS/Chint', 'ABN3c / NXB-63',   '조명, 소형 제어회로'],
    [5,  '2.4~4A', 'LS/Chint', 'ABN5c / NXB-63',   '소형 I/O, 팬'],
    [10, '4~8A',   'LS/Chint', 'ABN10c / NXB-63',  '소형 서보 1~2축'],
    [15, '8~12A',  'LS/Chint', 'ABN15c / NXB-63',  ''],
    [20, '12~16A', 'LS/Chint', 'ABN20c / NXB-63',  '중형 서보'],
    [25, '16~20A', 'LS/Chint', 'ABN25c / NXB-63',  ''],
    [30, '20~24A', 'LS/Chint', 'ABN30c / NXB-63',  '다축 서보 판넬'],
    [40, '24~32A', 'LS/Chint', 'ABN40c',            ''],
    [50, '32~40A', 'LS/Chint', 'ABN50c',            ''],
    [63, '40~50A', 'LS/Chint', 'ABN63c',            '대형 판넬'],
    [75, '50~60A', 'LS/Chint', 'ABN75c',            '메인 차단기'],
    [100,'60~80A', 'LS/Chint', 'ABN100c',           ''],
]
for i, row in enumerate(breaker_data):
    r = i + 4
    ws2.row_dimensions[r].height = 16
    for j, val in enumerate(row):
        col = get_column_letter(j+1)
        ws2[f'{col}{r}'] = val
        ws2[f'{col}{r}'].font = Font(name='맑은 고딕', size=9)
        ws2[f'{col}{r}'].alignment = center
        ws2[f'{col}{r}'].border = thin_border
        if j == 0:
            ws2[f'{col}{r}'].fill = green_fill

ws2_widths = {'A':14,'B':18,'C':12,'D':18,'E':22}
for col, w in ws2_widths.items():
    ws2.column_dimensions[col].width = w

out_path = 'C:/MES/wta-agents/workspaces/control-agent/판넬별_차단기선정.xlsx'
wb.save(out_path)
print('저장 완료:', out_path)
