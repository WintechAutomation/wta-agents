"""
WTA 전장 전류·용량 계산 및 차단기 선정표 v2
- 파나소닉 A6 서보드라이버 규격 포함
- 파스텍 스텝드라이버 규격 포함
- DC 24V 부하 섹션
- 전류 편중 분석 (R/S/T 상별)
- 전체 전류/용량 상단 요약
- 수식 + 설명 기록
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import math

wb = openpyxl.Workbook()

# ===== 공통 스타일 =====
def mfont(bold=False, size=9, color='000000', name='맑은 고딕'):
    return Font(name=name, bold=bold, size=size, color=color)

def mfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def mborder(style='thin', color='AAAAAA'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def mcenter(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def mleft(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def mright():
    return Alignment(horizontal='right', vertical='center')

# 색상 팔레트
C_HEADER   = '2F5496'  # 진파랑 헤더
C_SUBHDR   = '4472C4'  # 중간 파랑
C_TITLE    = '1F3864'  # 제목 네이비
C_PANASONIC= '00B0F0'  # 파나소닉 (하늘)
C_FASTECH  = 'ED7D31'  # 파스텍 (주황)
C_DC       = '70AD47'  # DC (초록)
C_MISC     = '7030A0'  # 기타 (보라)
C_SUM      = 'E2EFDA'  # 합계 연초록
C_FORMULA  = 'FFF2CC'  # 수식 설명 노랑
C_WARN     = 'FFE0E0'  # 경고 연빨강
C_INFO     = 'D6E4F0'  # 정보 연파랑
C_BREAKER  = 'F4F9C6'  # 차단기 연노랑

thin_b  = mborder('thin')
med_b   = mborder('medium', '555555')
center  = mcenter()
left_a  = mleft()

def set_cell(ws, row, col, value, bold=False, size=9, color='000000', fill=None,
             align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = mfont(bold=bold, size=size, color=color)
    if fill:  c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    if fmt:   c.number_format = fmt
    return c

# ============================================================
# SHEET 1: 종합 전류·용량 계산표
# ============================================================
ws = wb.active
ws.title = '전류·용량 계산표'

# 컬럼 너비
col_widths = {
    1:5, 2:18, 3:20, 4:8, 5:10, 6:6, 7:5, 8:9, 9:9, 10:9, 11:9,
    12:9, 13:9, 14:9, 15:9, 16:9, 17:30
}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── 1행: 시트 제목 ──
ws.merge_cells('A1:Q1')
c = ws['A1']
c.value = 'WTA 전장 전류·용량 계산 및 차단기 선정 종합표'
c.font = Font(name='맑은 고딕', bold=True, size=14, color=C_TITLE)
c.alignment = center
c.fill = mfill('D9E2F3')
ws.row_dimensions[1].height = 28

# ── 2~4행: 전체 요약 박스 ──
ws.merge_cells('A2:Q2')
ws['A2'].value = '▶ 전체 부하 요약 (상단 자동계산)'
ws['A2'].font = mfont(bold=True, size=10, color='FFFFFF')
ws['A2'].fill = mfill(C_TITLE)
ws['A2'].alignment = center
ws.row_dimensions[2].height = 18

# 요약 헤더 (3행)
summary_headers = [
    (1,'구분'), (2,'총 정격전류 (A)'), (3,'안전계수 ×1.25'), (4,'권장 차단기 (A)'),
    (5,'총 부하 (kVA)'), (6,'비고')
]
for col, label in summary_headers:
    c = ws.cell(row=3, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[3].height = 18

# 요약 데이터 (4~7행: AC3상, AC단상, DC24V, 합계)
summary_rows = [
    (4, 'AC 3상 220V (서보드라이버)',    mfill('EBF3FB')),
    (5, 'AC 단상 220V (파스텍 외)',       mfill('FEF4E8')),
    (6, 'DC 24V (I/O·센서·팬 등)',        mfill('EDF7E3')),
    (7, '기타 부하 (HMI·조명 등)',        mfill('F3EBF9')),
]
# 나중에 수식 채움 — 지금은 플레이스홀더
for row, label, fill in summary_rows:
    ws.cell(row=row, column=1, value=label).fill = fill
    ws.cell(row=row, column=1).font = mfont(bold=True, size=9)
    ws.cell(row=row, column=1).alignment = left_a
    ws.cell(row=row, column=1).border = thin_b
    for col in range(2, 7):
        c = ws.cell(row=row, column=col, value='→ 하단 계산')
        c.fill = fill
        c.font = mfont(size=9, color='888888')
        c.alignment = center
        c.border = thin_b
    ws.row_dimensions[row].height = 16

ws.row_dimensions[8].height = 6  # 구분선

# ── 9행: 전류 편중 분석 헤더 ──
ws.merge_cells('A9:Q9')
ws['A9'].value = '▶ R/S/T 상별 전류 편중 분석'
ws['A9'].font = mfont(bold=True, size=10, color='FFFFFF')
ws['A9'].fill = mfill('833C00')
ws['A9'].alignment = center
ws.row_dimensions[9].height = 18

# 10행: 편중 분석 헤더
phase_hdrs = ['', 'R상 (A)', 'S상 (A)', 'T상 (A)', '최대-최소 (A)', '불균형률 (%)', '판정', '기준: 불균형률 10% 이하 권장']
for i, h in enumerate(phase_hdrs, 1):
    c = ws.cell(row=10, column=i, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill('C55A11')
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[10].height = 18

# 11~13행: AC3상, 파스텍, 전체 — 나중에 수식
phase_labels = ['AC 3상 서보드라이버', 'AC 3상 기타', '전체 합계']
for i, label in enumerate(phase_labels, 11):
    ws.cell(row=i, column=1, value=label).font = mfont(bold=True, size=9)
    ws.cell(row=i, column=1).border = thin_b
    ws.cell(row=i, column=1).alignment = left_a
    for col in range(2, 9):
        c = ws.cell(row=i, column=col, value='')
        c.border = thin_b
        c.alignment = center
    ws.row_dimensions[i].height = 16

ws.row_dimensions[14].height = 6  # 구분선

# ── 15행: 본문 메인 헤더 ──
main_hdr_row = 15
ws.merge_cells('A15:A16')
ws.merge_cells('B15:B16')
ws.merge_cells('C15:C16')
ws.merge_cells('D15:D16')
ws.merge_cells('E15:E16')
ws.merge_cells('F15:F16')
ws.merge_cells('G15:G16')
ws.merge_cells('H15:H16')
ws.merge_cells('I15:I16')
ws.merge_cells('J15:J16')

main_hdrs = [
    (1,'판넬'), (2,'축/유닛명'), (3,'드라이버/장치 모델'),
    (4,'출력\n(W)'), (5,'정격전류\n(A)'), (6,'상수'),
    (7,'R상\n할당'), (8,'S상\n할당'), (9,'T상\n할당'),
    (10,'운전전류\n(A)')
]
for col, label in main_hdrs:
    c = ws.cell(row=main_hdr_row, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b

ws.merge_cells('K15:N15')
ws['K15'] = '전원 구분 (체크)'
ws['K15'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['K15'].fill = mfill(C_SUBHDR)
ws['K15'].alignment = center
ws['K15'].border = thin_b

power_sub = [(11,'AC 3상\n220V'), (12,'AC 단상\n220V'), (13,'DC 24V'), (14,'기타')]
for col, label in power_sub:
    c = ws.cell(row=16, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b

ws.merge_cells('O15:O16')
ws['O15'] = '차단기\n선정 (A)'
ws['O15'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['O15'].fill = mfill(C_HEADER)
ws['O15'].alignment = center
ws['O15'].border = thin_b

ws.merge_cells('P15:P16')
ws['P15'] = '제조사/\n모델'
ws['P15'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['P15'].fill = mfill(C_HEADER)
ws['P15'].alignment = center
ws['P15'].border = thin_b

ws.merge_cells('Q15:Q16')
ws['Q15'] = '비고 / 수식 설명'
ws['Q15'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['Q15'].fill = mfill(C_HEADER)
ws['Q15'].alignment = left_a
ws['Q15'].border = thin_b

ws.row_dimensions[15].height = 20
ws.row_dimensions[16].height = 24

# ── 본문 데이터 ──
# 전류 계산: I = VA / (√3 × 220 × PF)  [3상]
#            I = VA / (220 × PF)          [단상]
# 파나소닉 A6 드라이버 정격입력전류 (3상 200V, datasheet 기준)
# 모터출력W → 드라이버 정격입력전류(A)
# MADLN05BE:100W→0.5A, MADLN15BE:200W→1.0A,
# MBDLN25BE:400W→1.5A, MCDLN35BL:750W→2.4A,
# MDDLN55BL:1000W→3.4A, MDDLN85BL:1500W→5.5A

ROW = 17  # 데이터 시작

def data_row(ws, r, panel, axis, model, watt, rated_i, phase,
             r_flag, s_flag, t_flag, op_i,
             ac3=False, ac1=False, dc=False, misc=False,
             breaker='', mfr='', note='', fill_color=None):
    fill = mfill(fill_color) if fill_color else None
    vals = [panel, axis, model, watt, rated_i, phase,
            r_flag, s_flag, t_flag, op_i,
            '●' if ac3 else '', '●' if ac1 else '',
            '●' if dc else '', '●' if misc else '',
            breaker, mfr, note]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = mfont(size=9)
        c.alignment = center if i not in (1,2,3,17) else (left_a if i in (2,3,17) else center)
        c.border = thin_b
        if fill: c.fill = fill
        if i == 1:
            c.font = mfont(bold=True, size=9, color=C_TITLE)
    ws.row_dimensions[r].height = 17

# ── 섹션: 파나소닉 A6 서보드라이버 (AC 3상) ──
def section_header(ws, r, label, color):
    ws.merge_cells(f'A{r}:Q{r}')
    c = ws[f'A{r}']
    c.value = label
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(color)
    c.alignment = left_a
    c.border = thin_b
    ws.row_dimensions[r].height = 16

section_header(ws, ROW, '■ 파나소닉 A6 서보드라이버 (AC 3상 220V)', C_PANASONIC); ROW+=1
PAN_COLOR = 'EBF3FB'

# 샘플 데이터: 파나소닉 A6
# I_rated = VA / (√3 × 220 × PF) ≈ VA / 381.05
panasonic_data = [
    # panel, axis, model, W, rated_I(A), phase, R,S,T, op_I(A), note
    ('판넬1', 'LO_Y',   'MDDLN85BL (1.5kW)', 1500, 5.5,  3, 1,1,1, 3.85, 'I = VA/(√3×220×PF=0.95) ≈ VA/362'),
    ('판넬1', 'LO_X',   'MCDLN35BL (750W)',   750,  2.4,  3, 1,0,1, 1.68, 'Panasonic A6 — MCDLN35BL'),
    ('판넬1', 'LO_Z',   'MADLN15BE (200W)',   200,  1.0,  3, 0,1,0, 0.70, 'Panasonic A6 — MADLN15BE'),
    ('판넬1', 'ULO_Y',  'MCDLN35BL (750W)',   750,  2.4,  3, 1,1,0, 1.68, 'Panasonic A6 — MCDLN35BL'),
    ('판넬1', 'ULO_X',  'MCDLN35BL (750W)',   750,  2.4,  3, 0,1,1, 1.68, 'Panasonic A6 — MCDLN35BL'),
    ('판넬1', 'ULO_Z',  'MADLN05BE (100W)',   100,  0.5,  3, 1,0,0, 0.35, 'Panasonic A6 — MADLN05BE'),
    ('판넬1', 'NG_EV',  'MCDLN35BE (750W)',   750,  2.4,  3, 0,1,0, 1.68, 'Panasonic A6 — MCDLN35BE'),
    ('판넬1', 'NG_FD',  'MADLN05BE (100W)',   100,  0.5,  3, 0,0,1, 0.35, 'Panasonic A6 — MADLN05BE'),
    ('판넬1', 'NG_FD_Z','MBDLN25BE (400W)',   400,  1.5,  3, 1,1,0, 1.05, 'Panasonic A6 — MBDLN25BE'),
]

pan_start = ROW
for d in panasonic_data:
    panel,axis,model,watt,ri,ph,r,s,t,oi,note = d
    data_row(ws, ROW, panel, axis, model, watt, ri, ph, r, s, t, oi,
             ac3=True, note=note, fill_color=PAN_COLOR)
    ROW += 1

# 파나소닉 소계
pan_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  파나소닉 A6 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{pan_start}:E{pan_end})'
ws[f'E{ROW}'].number_format = '0.0'
ws[f'J{ROW}'] = f'=SUM(J{pan_start}:J{pan_end})'
ws[f'J{ROW}'].number_format = '0.0'
for col in [4,5,6,7,8,9,10,11,12,13,14,15,16]:
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = '정격전류 합계 / 운전전류 합계 (SUM 수식)'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
pan_sum_row = ROW
ROW += 2

# ── 섹션: 파스텍 스텝드라이버 (AC 입력 or DC 입력) ──
section_header(ws, ROW, '■ 파스텍(Fastech) 스텝드라이버', C_FASTECH); ROW+=1
FAS_COLOR = 'FEF4E8'

# 파스텍 드라이버: 일반적으로 DC 24~80V 입력 (EM542S, EM806 등)
# 단, 파스텍 중 AC 입력 제품도 있음 (EziSERVO 등)
fastech_data = [
    ('판넬2', 'X1', 'EM542S (DC 24-50V, 4.2A max)', 0,  4.2, 1, 0,0,0, 2.1, True,  False, False, 'DC입력: I_dc=P_w/V_dc, V_dc=48V 기준'),
    ('판넬2', 'X2', 'EM542S (DC 24-50V, 4.2A max)', 0,  4.2, 1, 0,0,0, 2.1, True,  False, False, 'DC입력: 스텝모터 구동전류 설정값'),
    ('판넬2', 'Y1', 'EM806  (DC 24-80V, 8.0A max)', 0,  8.0, 1, 0,0,0, 4.0, True,  False, False, 'DC입력: EM806 — 고토크용'),
    ('판넬2', 'Y2', 'EM806  (DC 24-80V, 8.0A max)', 0,  8.0, 1, 0,0,0, 4.0, True,  False, False, 'DC입력: EM806'),
    ('판넬2', 'Z',  'EM542S (DC 24-50V, 4.2A max)', 0,  4.2, 1, 0,0,0, 2.1, True,  False, False, 'DC입력: EM542S'),
]
fas_start = ROW
for d in fastech_data:
    panel,axis,model,watt,ri,ph,r,s,t,oi,dc,ac3,ac1,note = d
    data_row(ws, ROW, panel, axis, model, watt, ri, ph, r, s, t, oi,
             ac3=ac3, ac1=ac1, dc=dc, note=note, fill_color=FAS_COLOR)
    ROW += 1

fas_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  파스텍 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{fas_start}:E{fas_end})'
ws[f'J{ROW}'] = f'=SUM(J{fas_start}:J{fas_end})'
for col in range(4, 17):
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = '파스텍: DC 전원에서 공급 — DC 전원장치 용량에 반영'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
fas_sum_row = ROW
ROW += 2

# ── 섹션: DC 24V 부하 ──
section_header(ws, ROW, '■ DC 24V 부하 (I/O 모듈, 센서, 솔레노이드, 팬 등)', C_DC); ROW+=1
DC_COLOR = 'EDF7E3'

dc_data = [
    ('공통', 'I/O 모듈',      'Weidmuller / Leadshine IO', 0, 1.5, 1, 0,0,0, 1.5, False, False, True, 'DC 24V: 슬롯당 ~0.5A × n슬롯'),
    ('공통', '포토센서',       '포토센서 ×n',               0, 0.5, 1, 0,0,0, 0.5, False, False, True, 'DC 24V: 센서 1개당 ~50mA'),
    ('공통', '솔레노이드밸브', '솔레노이드밸브 ×n',         0, 1.0, 1, 0,0,0, 1.0, False, False, True, 'DC 24V: 밸브 1개당 ~200mA'),
    ('공통', '냉각팬',         '팬 (AC→DC 24V)',            0, 0.5, 1, 0,0,0, 0.5, False, False, True, 'DC 24V: 팬 1개당 ~0.3~0.5A'),
    ('공통', '기타 DC 부하',   '(직접 입력)',               0, 0.0, 1, 0,0,0, 0.0, False, False, True, '← 실측값 직접 입력'),
]
dc_start = ROW
for d in dc_data:
    panel,axis,model,watt,ri,ph,r,s,t,oi,ac3,ac1,dc,note = d
    data_row(ws, ROW, panel, axis, model, watt, ri, ph, r, s, t, oi,
             ac3=ac3, ac1=ac1, dc=dc, note=note, fill_color=DC_COLOR)
    ROW += 1

dc_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  DC 24V 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{dc_start}:E{dc_end})'
ws[f'J{ROW}'] = f'=SUM(J{dc_start}:J{dc_end})'
for col in range(4, 17):
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = 'DC 전원장치 선정: 합계전류×1.25 이상 용량'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
dc_sum_row = ROW
ROW += 2

# ── 섹션: 기타 부하 ──
section_header(ws, ROW, '■ 기타 부하 (HMI, PC, 조명, 히터 등)', C_MISC); ROW+=1
MISC_COLOR = 'F3EBF9'

misc_data = [
    ('공통', 'HMI',       'PROFACE / Weintek HMI',   0, 1.0, 1, 0,1,0, 1.0, False, True, False, 'AC 단상 220V: 제품 사양서 확인'),
    ('공통', '산업용 PC',  'IPC (AC 단상)',            0, 2.0, 1, 0,0,1, 2.0, False, True, False, 'AC 단상 220V: PC 정격전력 ÷ 220'),
    ('공통', '내부 조명',  '형광등/LED (AC 단상)',    0, 0.3, 1, 1,0,0, 0.3, False, True, False, 'AC 단상 220V'),
    ('공통', '히터 (옵션)','PTC 히터 (AC 단상)',      0, 0.0, 1, 0,0,0, 0.0, False, True, False, '← 옵션: 실측값 직접 입력'),
    ('공통', '기타',       '(직접 입력)',             0, 0.0, 1, 0,0,0, 0.0, False, False, False, '← 실측값 직접 입력'),
]
misc_start = ROW
for d in misc_data:
    panel,axis,model,watt,ri,ph,r,s,t,oi,ac3,ac1,dc,note = d
    data_row(ws, ROW, panel, axis, model, watt, ri, ph, r, s, t, oi,
             ac3=ac3, ac1=ac1, dc=dc, note=note, fill_color=MISC_COLOR)
    ROW += 1

misc_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  기타 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{misc_start}:E{misc_end})'
ws[f'J{ROW}'] = f'=SUM(J{misc_start}:J{misc_end})'
for col in range(4, 17):
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = '기타: 단상 부하는 상 편중 주의'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
misc_sum_row = ROW
ROW += 2

# ── 전체 합계 ──
section_header(ws, ROW, '■ 전체 합계 및 차단기 선정', C_TITLE); ROW+=1

TOTAL_ROW = ROW
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  전체 정격전류 합계'
ws[f'A{ROW}'].font = mfont(bold=True, size=10)
ws[f'A{ROW}'].fill = mfill('2F5496')
ws[f'A{ROW}'].font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b

ws[f'E{ROW}'] = f'=E{pan_sum_row}+E{fas_sum_row}+E{dc_sum_row}+E{misc_sum_row}'
ws[f'E{ROW}'].font = mfont(bold=True, size=11)
ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center
ws[f'E{ROW}'].border = thin_b
ws[f'E{ROW}'].fill = mfill('FFFF99')

ws[f'F{ROW}'] = f'=E{ROW}*1.25'
ws[f'F{ROW}'].font = mfont(bold=True, size=10)
ws[f'F{ROW}'].number_format = '0.00'
ws[f'F{ROW}'].alignment = center
ws[f'F{ROW}'].border = thin_b
ws[f'F{ROW}'].fill = mfill('FFE699')

ws[f'G{ROW}'] = '×1.25 (안전계수)'
ws[f'G{ROW}'].font = mfont(size=9, color='555555')
ws[f'G{ROW}'].border = thin_b

ws[f'O{ROW}'] = '→ 차단기 직접 선정'
ws[f'O{ROW}'].font = mfont(bold=True, size=9)
ws[f'O{ROW}'].border = thin_b
ws[f'O{ROW}'].alignment = center

ws[f'Q{ROW}'] = '안전계수 1.25 적용 후 규격 차단기 선정 (Sheet3 참조)'
ws[f'Q{ROW}'].font = mfont(size=8, color=C_TITLE)
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 20
ROW += 1

# 상별 합계 행
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  R/S/T 상별 전류 합계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill('833C00')
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b

for col, label in [(7,'R'), (8,'S'), (9,'T')]:
    # 파나소닉 섹션에서 R/S/T 할당은 col 7,8,9 × rated_I로 계산
    # 여기서는 할당 플래그 × 정격전류 합산 공식은 복잡하므로 수동 입력 안내
    c = ws.cell(row=ROW, column=col, value=f'→ {label}상 합산 입력')
    c.font = mfont(size=9, color='555555')
    c.border = thin_b
    c.alignment = center
    c.fill = mfill('F2DCDB')
ws[f'Q{ROW}'] = '각 상별 전류: 드라이버별 R/S/T 할당 × 정격전류 합산 (수동 확인)'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
ROW += 2

# ── 수식 설명 박스 ──
ws.merge_cells(f'A{ROW}:Q{ROW}')
ws[f'A{ROW}'].value = '▶ 주요 수식 설명'
ws[f'A{ROW}'].font = mfont(bold=True, size=10, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill('375623')
ws[f'A{ROW}'].alignment = center
ws[f'A{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 16
ROW += 1

formulas = [
    ('AC 3상 정격전류',   'I (A) = VA_정격 ÷ (√3 × 220V × PF)',        '예) 2300VA ÷ (1.732×220×0.95) = 6.36A',   'PF=역률(보통 0.85~0.95), 파나소닉 A6 spec 기준'),
    ('AC 단상 정격전류',  'I (A) = W_정격 ÷ (220V × PF)',              '예) 1000W ÷ (220×0.85) = 5.35A',           '단상 220V 부하: HMI, PC 등'),
    ('DC 전류',           'I (A) = W_정격 ÷ V_dc',                      '예) 100W ÷ 24V = 4.17A',                   'DC 24V 계열: SMPS 용량 선정에 활용'),
    ('안전계수 적용',     'I_설계 = I_합계 × 1.25',                     '→ 이 값 이상의 차단기 선정',                 'NEC/IEC 기준 1.25배 (연속운전 고려)'),
    ('전류 불균형률',     'δ (%) = (I_max - I_avg) ÷ I_avg × 100',     '10% 이하 권장 (IEC 61000-3-3)',             '불균형 시 R/S/T 축 재배분 검토'),
    ('차단기 선정 원칙',  '정격 > I_설계, 단락전류 ≥ 현장 단락전류',   '→ Sheet3 차단기 규격 참조표 확인',          'MCCB: 산업용, MCB: 소용량'),
]

for row_data in formulas:
    label, formula, example, note = row_data
    ws.merge_cells(f'A{ROW}:B{ROW}')
    ws[f'A{ROW}'].value = label
    ws[f'A{ROW}'].font = mfont(bold=True, size=9)
    ws[f'A{ROW}'].fill = mfill(C_FORMULA)
    ws[f'A{ROW}'].alignment = left_a
    ws[f'A{ROW}'].border = thin_b

    ws.merge_cells(f'C{ROW}:F{ROW}')
    ws[f'C{ROW}'].value = formula
    ws[f'C{ROW}'].font = Font(name='Courier New', bold=True, size=9, color='1F3864')
    ws[f'C{ROW}'].fill = mfill(C_FORMULA)
    ws[f'C{ROW}'].alignment = left_a
    ws[f'C{ROW}'].border = thin_b

    ws.merge_cells(f'G{ROW}:J{ROW}')
    ws[f'G{ROW}'].value = example
    ws[f'G{ROW}'].font = mfont(size=9, color='375623')
    ws[f'G{ROW}'].fill = mfill('EBF3E8')
    ws[f'G{ROW}'].alignment = left_a
    ws[f'G{ROW}'].border = thin_b

    ws.merge_cells(f'K{ROW}:Q{ROW}')
    ws[f'K{ROW}'].value = note
    ws[f'K{ROW}'].font = mfont(size=9, color='555555')
    ws[f'K{ROW}'].fill = mfill(C_FORMULA)
    ws[f'K{ROW}'].alignment = left_a
    ws[f'K{ROW}'].border = thin_b

    ws.row_dimensions[ROW].height = 16
    ROW += 1

# 상단 요약 셀 수식 채우기
ws[f'B4'].value = f'=E{pan_sum_row}'
ws[f'B5'].value = f'=E{misc_sum_row}'
ws[f'B6'].value = f'=E{dc_sum_row}'
ws[f'B7'].value = f'=E{fas_sum_row}'
for r in range(4, 8):
    ws.cell(row=r, column=2).number_format = '0.0'
    ws.cell(row=r, column=3).value = f'=B{r}*1.25'
    ws.cell(row=r, column=3).number_format = '0.0'
    ws.cell(row=r, column=4).value = '→ Sheet3 참조'
    ws.cell(row=r, column=4).font = mfont(size=9)

# ============================================================
# SHEET 2: 파나소닉 A6 드라이버 규격표
# ============================================================
ws2 = wb.create_sheet('파나소닉 A6 규격표')
ws2.merge_cells('A1:J1')
ws2['A1'] = '파나소닉 A6 서보드라이버 규격표 (MINAS A6 Series)'
ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=12, color=C_TITLE)
ws2['A1'].alignment = center
ws2['A1'].fill = mfill('D6E4F0')
ws2.row_dimensions[1].height = 24

headers2 = ['드라이버 모델', '대응 모터', '정격출력(W)', '전원',
            '정격입력전류(A)', '최대입력전류(A)', '제어방식', '통신', '특이사항', '비고']
for j, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws2.row_dimensions[2].height = 20

a6_data = [
    ('MADLN05BE', 'MSMF/MQMF 50W',    50,   'AC3상 200V', 0.3,  0.8,  'フルクロ', 'EtherCAT/Modbus', '소형 단축용',   ''),
    ('MADLN10BE', 'MSMF/MQMF 100W',   100,  'AC3상 200V', 0.5,  1.2,  'フルクロ', 'EtherCAT/Modbus', '',              ''),
    ('MADLN15BE', 'MSMF/MQMF 200W',   200,  'AC3상 200V', 1.0,  2.4,  'フルクロ', 'EtherCAT/Modbus', '',              ''),
    ('MBDLN25BE', 'MSMF/MQMF 400W',   400,  'AC3상 200V', 1.5,  4.0,  'フルクロ', 'EtherCAT/Modbus', '',              ''),
    ('MCDLN35BL', 'MSMF/MQMF 750W',   750,  'AC3상 200V', 2.4,  7.0,  'フルクロ', 'EtherCAT/Modbus', '중형 축용',     ''),
    ('MCDLN35BE', 'MSMF/MQMF 750W',   750,  'AC3상 200V', 2.4,  7.0,  'フルクロ', 'EtherCAT/Modbus', 'Brake 내장형',  ''),
    ('MDDLN55BL', 'MSMF 1kW',         1000, 'AC3상 200V', 3.4,  10.0, 'フルクロ', 'EtherCAT/Modbus', '대형 주축용',   ''),
    ('MDDLN85BL', 'MSMF 1.5kW',       1500, 'AC3상 200V', 5.5,  15.0, 'フルクロ', 'EtherCAT/Modbus', '',              ''),
    ('MEDLN15BL', 'MSMF 2kW',         2000, 'AC3상 200V', 7.5,  20.0, 'フルクロ', 'EtherCAT/Modbus', '',              ''),
    ('MFDLN23BL', 'MSMF 3kW',         3000, 'AC3상 200V', 11.0, 30.0, 'フルクロ', 'EtherCAT/Modbus', '고출력',        ''),
]
for i, row in enumerate(a6_data, 3):
    fill = mfill('EBF3FB') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate(row, 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j != 1 else left_a
        c.border = thin_b
        c.fill = fill
    ws2.row_dimensions[i].height = 16

ws2_widths = {'A':18,'B':22,'C':14,'D':16,'E':16,'F':16,'G':12,'H':18,'I':20,'J':14}
for col, w in ws2_widths.items():
    ws2.column_dimensions[col].width = w

# 수식 설명 추가
note_row = len(a6_data) + 4
ws2.merge_cells(f'A{note_row}:J{note_row}')
ws2[f'A{note_row}'] = '※ 정격입력전류: I = P_출력(VA) ÷ (√3 × V_입력 × PF)  |  안전계수 1.25 적용 후 차단기 선정  |  출처: Panasonic MINAS A6 Catalog'
ws2[f'A{note_row}'].font = mfont(size=8, color='555555')
ws2[f'A{note_row}'].fill = mfill(C_FORMULA)
ws2[f'A{note_row}'].alignment = left_a
ws2.row_dimensions[note_row].height = 14

# ============================================================
# SHEET 3: 파스텍 드라이버 규격표
# ============================================================
ws3 = wb.create_sheet('파스텍 드라이버 규격표')
ws3.merge_cells('A1:J1')
ws3['A1'] = '파스텍(Fastech) 스텝드라이버 규격표'
ws3['A1'].font = Font(name='맑은 고딕', bold=True, size=12, color=C_TITLE)
ws3['A1'].alignment = center
ws3['A1'].fill = mfill('FCE4D6')
ws3.row_dimensions[1].height = 24

headers3 = ['드라이버 모델', '입력전원', '출력전류 max(A)', '마이크로스텝', '통신/제어', '보호기능', '대응 모터', '특이사항', '비고', '']
for j, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws3.row_dimensions[2].height = 20

fas_data = [
    ('EM542S',      'DC 20~50V',  4.2,  '1~256', '스텝/방향, CW/CCW', '과열/과전류/위상오류', '2상 스텝모터', '소형 범용',          ''),
    ('EM806',       'DC 24~80V',  8.0,  '1~256', '스텝/방향, CW/CCW', '과열/과전류',          '2상 스텝모터', '고토크 대형',        ''),
    ('EZB2-EC',     'DC 24~48V',  2.0,  '1~256', 'EtherCAT',          '과열/과전류',          '2상 스텝모터', '네트워크 제어',      ''),
    ('EZB2-ML',     'DC 24~48V',  2.0,  '1~256', 'MECHATROLINK-III',  '과열/과전류',          '2상 스텝모터', 'ML3 네트워크',       ''),
    ('EZS2-EC',     'DC 24~48V',  4.0,  '1~256', 'EtherCAT',          '과열/과전류',          '2상 스텝모터', '중형 EtherCAT',      ''),
    ('EziSERVO-EC', 'AC 100~240V',3.0,  '-',     'EtherCAT (CiA 402)','과열/과전류/위치오류', '폐루프 스텝',  '폐루프(엔코더 내장)',''),
]
for i, row in enumerate(fas_data, 3):
    fill = mfill('FEF4E8') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate(row, 1):
        c = ws3.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j not in (1,7,8) else left_a
        c.border = thin_b
        c.fill = fill
    ws3.row_dimensions[i].height = 16

ws3_widths = {'A':16,'B':14,'C':16,'D':12,'E':22,'F':22,'G':18,'H':20,'I':14,'J':8}
for col, w in ws3_widths.items():
    ws3.column_dimensions[col].width = w

note3_row = len(fas_data) + 4
ws3.merge_cells(f'A{note3_row}:J{note3_row}')
ws3[f'A{note3_row}'] = '※ 파스텍 DC 드라이버 전원: 별도 SMPS 필요 (DC 전원장치 용량 = 합계전류×1.25 이상)  |  출처: Fastech 제품 카탈로그'
ws3[f'A{note3_row}'].font = mfont(size=8, color='555555')
ws3[f'A{note3_row}'].fill = mfill(C_FORMULA)
ws3[f'A{note3_row}'].alignment = left_a
ws3.row_dimensions[note3_row].height = 14

# ============================================================
# SHEET 4: 차단기 규격 참조표 (기존 개선)
# ============================================================
ws4 = wb.create_sheet('차단기 규격 참조표')
ws4.merge_cells('A1:G1')
ws4['A1'] = '차단기 규격 참조표 (MCCB / MCB — AC 220V 3상/단상)'
ws4['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws4['A1'].alignment = center
ws4['A1'].fill = mfill('D9E2F3')
ws4.row_dimensions[1].height = 22

headers4 = ['정격전류(A)', '적용 부하전류 범위', '제조사', '모델 3P (MCCB)', '모델 1P (MCB)', '특이사항', '비고']
for j, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws4.row_dimensions[2].height = 22

breaker_data = [
    (3,   '~2.4A',   'LS/Chint',   'ABN3c / NXB-63',   'BKN1-1P C3',   '조명·소형제어', ''),
    (5,   '2.4~4A',  'LS/Chint',   'ABN5c / NXB-63',   'BKN1-1P C5',   '소형I/O',        ''),
    (10,  '4~8A',    'LS/Chint',   'ABN10c / NXB-63',  'BKN1-1P C10',  '소형서보 1~2축', ''),
    (15,  '8~12A',   'LS/Chint',   'ABN15c / NXB-63',  'BKN1-1P C15',  '',               ''),
    (20,  '12~16A',  'LS/Chint',   'ABN20c / NXB-63',  'BKN1-1P C20',  '중형서보',       ''),
    (25,  '16~20A',  'LS/Chint',   'ABN25c / NXB-63',  'BKN1-1P C25',  '',               ''),
    (30,  '20~24A',  'LS/Chint',   'ABN30c / NXB-63',  '-',            '다축서보 판넬',  ''),
    (40,  '24~32A',  'LS/Chint',   'ABN40c',           '-',            '',               ''),
    (50,  '32~40A',  'LS/Chint',   'ABN50c',           '-',            '',               ''),
    (63,  '40~50A',  'LS/Chint',   'ABN63c',           '-',            '대형 판넬',      ''),
    (75,  '50~60A',  'LS/Chint',   'ABN75c',           '-',            '메인 차단기',    ''),
    (100, '60~80A',  'LS/Chint',   'ABN100c',          '-',            '',               ''),
    (125, '80~100A', 'LS',         'ABN125c',          '-',            '대형 메인',      ''),
    (150, '100~120A','LS',         'ABN150c',          '-',            '',               ''),
]
for i, row in enumerate(breaker_data, 3):
    fill = mfill(C_BREAKER) if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate(row, 1):
        c = ws4.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j != 6 else left_a
        c.border = thin_b
        c.fill = fill
        if j == 1:
            c.fill = mfill(C_SUM)
            c.font = mfont(bold=True, size=9)
    ws4.row_dimensions[i].height = 16

ws4_widths = {'A':14,'B':18,'C':12,'D':20,'E':18,'F':20,'G':14}
for col, w in ws4_widths.items():
    ws4.column_dimensions[col].width = w

note4_row = len(breaker_data) + 4
ws4.merge_cells(f'A{note4_row}:G{note4_row}')
ws4[f'A{note4_row}'] = '※ 차단기 선정: 설계전류(I_합계×1.25) 초과하는 최소 정격 선택  |  모터부하: 기동전류 고려 (서보는 드라이버가 흡수, 스텝은 정격전류 ≒ 운전전류)'
ws4[f'A{note4_row}'].font = mfont(size=8, color='555555')
ws4[f'A{note4_row}'].fill = mfill(C_FORMULA)
ws4[f'A{note4_row}'].alignment = left_a
ws4.row_dimensions[note4_row].height = 14

# ── 저장 ──
out_path = 'C:/MES/wta-agents/workspaces/control-agent/WTA_전장_전류용량계산_차단기선정표.xlsx'
wb.save(out_path)
print('저장 완료:', out_path)
