"""
대구텍 검사기 F2 #1 판넬별 전류·용량 계산 및 차단기 선정표 v7
- 지건승님 파일(WTA_장비전장설계용량선정(수정중2).xlsx) 양식 기준
- 출력(W), 정격전류(A) → VLOOKUP 자동조회
- 역률/안전계수/수용률 → 셀 참조 수식으로 어디에 반영되는지 명시
- 파스텍 드라이버 판넬별 배치
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 공통 스타일 =====
def mfont(bold=False, size=9, color='000000', name='맑은 고딕'):
    return Font(name=name, bold=bold, size=size, color=color)
def mfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)
def mborder(style='thin', color='AAAAAA'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def mcenter(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def mleft(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

C_HEADER   = '2F5496'
C_SUBHDR   = '4472C4'
C_TITLE    = '1F3864'
C_PANASONIC= '00B0F0'
C_FASTECH  = 'ED7D31'
C_DC       = '70AD47'
C_MISC     = '7030A0'
C_SUM      = 'E2EFDA'
C_BREAKER  = 'FFFF99'
C_INPUT    = 'FFF2CC'   # 입력값 셀 (노란색)
C_FORMULA  = 'EBF3FB'  # 수식 셀 (연파란)

thin_b = mborder('thin')
center = mcenter()
left_a = mleft()

# ===== SHEET 1 =====
ws = wb.active
ws.title = '판넬별 차단기 선정표'

col_widths = {1:5, 2:20, 3:24, 4:9, 5:11, 6:6, 7:5, 8:9, 9:9, 10:11, 11:9, 12:9, 13:9, 14:9, 15:10, 16:14, 17:38}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Row 1: 제목 ──
ws.merge_cells('A1:Q1')
c = ws['A1']
c.value = '대구텍 검사기 F2 #1 — 판넬별 전류·용량 계산 및 차단기 선정표'
c.font = Font(name='맑은 고딕', bold=True, size=14, color=C_TITLE)
c.alignment = center
c.fill = mfill('D9E2F3')
ws.row_dimensions[1].height = 28

# ── Row 2: 기준값 입력셀 (역률/안전계수/수용률) ──
# 셀 참조 규칙: PF=D2, 안전계수=F2, 수용률=H2
pairs = [
    ('A2','전원:'), ('B2','AC 220V / 3상'),
    ('C2','역률(PF):'), ('D2', 0.95),
    ('E2','안전계수:'), ('F2', 1.25),
    ('G2','수용률:'), ('H2', 0.7),
    ('I2','작성일:'), ('J2','2026-04-03'),
]
for addr, val in pairs:
    ws[addr] = val
    col_letter = addr[0]
    ws[addr].font = Font(name='맑은 고딕', size=9,
                         bold=(col_letter in 'ACEGI'))
# 입력셀 강조 (노란색)
for addr in ['D2','F2','H2']:
    ws[addr].fill = mfill(C_INPUT)
    ws[addr].font = Font(name='맑은 고딕', size=10, bold=True, color='8B0000')
ws.row_dimensions[2].height = 18

# 주석: 각 값이 어디에 쓰이는지
ws.merge_cells('K2:Q2')
ws['K2'] = '▶ 역률(PF): kVA 환산에 반영  ▶ 안전계수×1.25: 각 판넬 소계의 권장차단기 산출  ▶ 수용률: J열 설계전류 = 정격전류(E열)×수용률'
ws['K2'].font = Font(name='맑은 고딕', size=8, color='444444', italic=True)
ws['K2'].alignment = left_a
ws['K2'].fill = mfill('F0F0F0')
ws.row_dimensions[3].height = 5

# ── Rows 4~8: 전체 요약 박스 ──
ws.merge_cells('A4:Q4')
ws['A4'].value = '▶ 전체 부하 요약 (판넬별 소계 완성 후 자동계산)'
ws['A4'].font = mfont(bold=True, size=10, color='FFFFFF')
ws['A4'].fill = mfill(C_TITLE)
ws['A4'].alignment = center
ws.row_dimensions[4].height = 16

for j, h in enumerate(['구분','총 정격전류 (A)','×안전계수(1.25)','권장 차단기 (A)','총 부하 (kVA)','비고'], 1):
    c = ws.cell(row=5, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[5].height = 16

# 요약 데이터 행 (6~9) — sum_row 변수는 아래서 채움
summary_labels = [
    ('AC 3상 220V\n파나소닉 A6 서보드라이버', 'EBF3FB'),
    ('AC 단상 220V\n파스텍 EziServo/EziSpeed',  'FEF4E8'),
    ('DC 24V\n서보I/O·센서·솔레노이드 등',       'EDF7E3'),
    ('기타 부하\nUPS·HMI·조명 등',               'F3EBF9'),
]
for i, (label, clr) in enumerate(summary_labels):
    r = 6 + i
    ws.cell(row=r, column=1, value=label).fill = mfill(clr)
    ws.cell(row=r, column=1).font = mfont(bold=True, size=9)
    ws.cell(row=r, column=1).alignment = left_a
    ws.cell(row=r, column=1).border = thin_b
    for col in range(2, 7):
        ws.cell(row=r, column=col, value='→ 하단 계산').fill = mfill(clr)
        ws.cell(row=r, column=col).font = mfont(size=8, color='888888')
        ws.cell(row=r, column=col).alignment = center
        ws.cell(row=r, column=col).border = thin_b
    ws.row_dimensions[r].height = 22

ws.row_dimensions[10].height = 5

# ── Rows 11~13: R/S/T 편중 분석 ──
ws.merge_cells('A11:Q11')
ws['A11'].value = '▶ R/S/T 상별 전류 편중 분석 (참조용 — 상 배분 최적화 시 활용)'
ws['A11'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['A11'].fill = mfill('833C00')
ws['A11'].alignment = center
ws.row_dimensions[11].height = 16

for j, h in enumerate(['구분','R상 (A)','S상 (A)','T상 (A)','최대-최소 (A)','불균형률 (%)','판정','기준: 10% 이하 권장'], 1):
    c = ws.cell(row=12, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill('C55A11')
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[12].height = 16

for i, label in enumerate(['AC 3상 서보드라이버','기타 AC 부하','전체 합계'], 13):
    ws.cell(row=i, column=1, value=label).font = mfont(bold=True, size=9)
    ws.cell(row=i, column=1).border = thin_b
    ws.cell(row=i, column=1).alignment = left_a
    for col in range(2, 9):
        ws.cell(row=i, column=col).border = thin_b
        ws.cell(row=i, column=col).alignment = center
    ws.row_dimensions[i].height = 16

ws.row_dimensions[16].height = 5

# ── Row 17: 기준 표시 (데이터 섹션용 참조 재표시) ──
for addr, val in [('A17','기준:'), ('B17','역률(PF)=$D$2'), ('C17','안전계수=$F$2'), ('D17','수용률=$H$2')]:
    ws[addr] = val
    ws[addr].font = Font(name='맑은 고딕', size=8, color='555555')
    ws[addr].fill = mfill('F0F0F0')
ws.row_dimensions[17].height = 14

# ── Rows 18~19: 컬럼 헤더 ──
def write_headers(ws, row):
    for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
        ws.merge_cells(f'{col_letter}{row}:{col_letter}{row+1}')
    for col, label in [(1,'판넬'),(2,'축/유닛명'),(3,'드라이버/장치 모델'),
                       (4,'출력(W)\n[자동]'),(5,'정격전류(A)\n[자동]'),
                       (6,'상수'),(7,'R상\n할당'),(8,'S상\n할당'),(9,'T상\n할당'),
                       (10,'설계전류(A)\n=E×수용률')]:
        c = ws.cell(row=row, column=col, value=label)
        c.font = mfont(bold=True, size=9, color='FFFFFF')
        c.fill = mfill(C_HEADER)
        c.alignment = center
        c.border = thin_b
    ws.merge_cells(f'K{row}:N{row}')
    ws[f'K{row}'] = '전원 구분'
    ws[f'K{row}'].font = mfont(bold=True, size=9, color='FFFFFF')
    ws[f'K{row}'].fill = mfill(C_SUBHDR)
    ws[f'K{row}'].alignment = center
    ws[f'K{row}'].border = thin_b
    for col, label in [(11,'AC 3상\n220V'),(12,'AC 단상\n220V'),(13,'DC 24V'),(14,'기타')]:
        c = ws.cell(row=row+1, column=col, value=label)
        c.font = mfont(bold=True, size=9, color='FFFFFF')
        c.fill = mfill(C_SUBHDR)
        c.alignment = center
        c.border = thin_b
    for col_letter, label in [('O','권장\n차단기(A)'),('P','제조사/모델'),('Q','비고 / 수식 설명')]:
        ws.merge_cells(f'{col_letter}{row}:{col_letter}{row+1}')
        ws[f'{col_letter}{row}'] = label
        ws[f'{col_letter}{row}'].font = mfont(bold=True, size=9, color='FFFFFF')
        ws[f'{col_letter}{row}'].fill = mfill(C_HEADER)
        ws[f'{col_letter}{row}'].alignment = center if col_letter != 'Q' else left_a
        ws[f'{col_letter}{row}'].border = thin_b
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row+1].height = 24

write_headers(ws, 18)
ROW = 20

# ===== 셀 참조 상수 (기준값 행) =====
PF_REF   = '$D$2'   # 역률
SF_REF   = '$F$2'   # 안전계수
UR_REF   = '$H$2'   # 수용률

def vlookup_d(r):
    """출력W 자동조회"""
    return (f'=IFERROR(VLOOKUP(C{r},파나소닉A6규격!$A:$B,2,FALSE),'
            f'"")')

def vlookup_e(r):
    """정격전류 자동조회"""
    return (f'=IFERROR(VLOOKUP(C{r},파나소닉A6규격!$A:$C,3,FALSE),'
            f'IFERROR(VLOOKUP(C{r},파스텍EziServo규격!$A:$B,2,FALSE),""))')

def design_j(r):
    """설계전류 = 정격전류 × 수용률(셀참조)"""
    return f'=IF(E{r}="","",ROUND(E{r}*{UR_REF},3))'

def section_header(ws, r, label, color):
    ws.merge_cells(f'A{r}:Q{r}')
    c = ws[f'A{r}']
    c.value = label
    c.font = mfont(bold=True, size=10, color='FFFFFF')
    c.fill = mfill(color)
    c.alignment = left_a
    c.border = thin_b
    ws.row_dimensions[r].height = 18

def data_row(ws, r, panel, axis, model, r_flag, s_flag, t_flag,
             ac3=False, ac1=False, dc=False, misc=False,
             note='', fill_color=None, fixed_d=None, fixed_e=None, fixed_j=None):
    """데이터 행 삽입. fixed_*는 직접 값 입력(DC 부하용), None이면 VLOOKUP"""
    fill = mfill(fill_color) if fill_color else None
    # A~C: 판넬/축/모델
    for col, val in [(1,panel),(2,axis),(3,model)]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = mfont(size=9) if col != 1 else mfont(bold=True, size=9, color=C_TITLE)
        c.alignment = left_a if col >= 2 else center
        c.border = thin_b
        if fill: c.fill = fill
    # D: 출력W (VLOOKUP 또는 직접)
    ws[f'D{r}'] = fixed_d if fixed_d is not None else vlookup_d(r)
    ws[f'D{r}'].font = mfont(size=9, color='1F3864' if fixed_d is None else '000000')
    ws[f'D{r}'].alignment = center
    ws[f'D{r}'].border = thin_b
    if fill: ws[f'D{r}'].fill = fill
    # E: 정격전류 (VLOOKUP 또는 직접)
    ws[f'E{r}'] = fixed_e if fixed_e is not None else vlookup_e(r)
    ws[f'E{r}'].font = mfont(size=9, color='1F3864' if fixed_e is None else '000000')
    ws[f'E{r}'].alignment = center
    ws[f'E{r}'].border = thin_b
    ws[f'E{r}'].number_format = '0.000'
    if fill: ws[f'E{r}'].fill = fill
    # F: 상수
    ws[f'F{r}'] = 1
    ws[f'F{r}'].font = mfont(size=9)
    ws[f'F{r}'].alignment = center
    ws[f'F{r}'].border = thin_b
    if fill: ws[f'F{r}'].fill = fill
    # G/H/I: R/S/T
    for col, val in [(7,r_flag),(8,s_flag),(9,t_flag)]:
        ws.cell(row=r, column=col, value=val).font = mfont(size=9)
        ws.cell(row=r, column=col).alignment = center
        ws.cell(row=r, column=col).border = thin_b
        if fill: ws.cell(row=r, column=col).fill = fill
    # J: 설계전류 (수식 또는 직접)
    ws[f'J{r}'] = fixed_j if fixed_j is not None else design_j(r)
    ws[f'J{r}'].font = mfont(size=9, color='375623')
    ws[f'J{r}'].alignment = center
    ws[f'J{r}'].border = thin_b
    ws[f'J{r}'].number_format = '0.000'
    if fill: ws[f'J{r}'].fill = fill
    # K~N: 전원 구분
    for col, flag in [(11,ac3),(12,ac1),(13,dc),(14,misc)]:
        ws.cell(row=r, column=col, value='●' if flag else '').font = mfont(size=9)
        ws.cell(row=r, column=col).alignment = center
        ws.cell(row=r, column=col).border = thin_b
        if fill: ws.cell(row=r, column=col).fill = fill
    # O/P: 빈칸
    for col in [15, 16]:
        ws.cell(row=r, column=col).border = thin_b
        if fill: ws.cell(row=r, column=col).fill = fill
    # Q: 비고
    ws[f'Q{r}'] = note
    ws[f'Q{r}'].font = mfont(size=8, color='555555')
    ws[f'Q{r}'].alignment = left_a
    ws[f'Q{r}'].border = thin_b
    if fill: ws[f'Q{r}'].fill = fill
    ws.row_dimensions[r].height = 17

def sum_row(ws, r, label, start, end, breaker_a, mfr_model, color_sum=C_SUM, include_fastech=False, extra_ranges=None):
    """판넬 소계 + 차단기 선정행"""
    ws.merge_cells(f'A{r}:C{r}')
    ws[f'A{r}'] = f'  ▶ {label}'
    ws[f'A{r}'].font = mfont(bold=True, size=9, color='FFFFFF')
    ws[f'A{r}'].fill = mfill(C_TITLE)
    ws[f'A{r}'].alignment = left_a
    ws[f'A{r}'].border = thin_b
    # E: 정격전류 합계
    e_ranges = f'E{start}:E{end}'
    if extra_ranges:
        extra_e = ','.join([f'E{s}:E{e}' for s, e in extra_ranges])
        e_formula = f'=SUM({e_ranges},{extra_e})'
    else:
        e_formula = f'=SUM({e_ranges})'
    ws[f'D{r}'].border = thin_b
    ws[f'D{r}'].fill = mfill(color_sum)
    ws[f'E{r}'] = e_formula
    ws[f'E{r}'].font = mfont(bold=True, size=9)
    ws[f'E{r}'].number_format = '0.00'
    ws[f'E{r}'].alignment = center
    ws[f'E{r}'].border = thin_b
    ws[f'E{r}'].fill = mfill(color_sum)
    # F~I
    for col in [6,7,8,9]:
        ws.cell(row=r, column=col).border = thin_b
        ws.cell(row=r, column=col).fill = mfill(color_sum)
    # J: 설계전류 합계
    j_ranges = f'J{start}:J{end}'
    if extra_ranges:
        extra_j = ','.join([f'J{s}:J{e}' for s, e in extra_ranges])
        j_formula = f'=SUM({j_ranges},{extra_j})'
    else:
        j_formula = f'=SUM({j_ranges})'
    ws[f'J{r}'] = j_formula
    ws[f'J{r}'].font = mfont(bold=True, size=9)
    ws[f'J{r}'].number_format = '0.00'
    ws[f'J{r}'].alignment = center
    ws[f'J{r}'].border = thin_b
    ws[f'J{r}'].fill = mfill(color_sum)
    # K~N
    for col in range(11, 15):
        ws.cell(row=r, column=col).border = thin_b
        ws.cell(row=r, column=col).fill = mfill(color_sum)
    # O: 권장차단기
    ws[f'O{r}'] = breaker_a
    ws[f'O{r}'].font = mfont(bold=True, size=10, color='1F3864')
    ws[f'O{r}'].alignment = center
    ws[f'O{r}'].border = thin_b
    ws[f'O{r}'].fill = mfill(C_BREAKER)
    # P: 제조사/모델
    ws[f'P{r}'] = mfr_model
    ws[f'P{r}'].font = mfont(bold=True, size=9)
    ws[f'P{r}'].alignment = center
    ws[f'P{r}'].border = thin_b
    ws[f'P{r}'].fill = mfill(C_BREAKER)
    # Q: 비고 — 안전계수 셀 참조 표시
    ws[f'Q{r}'] = (f'=TEXT(J{r},"0.00")&"A×안전계수("&{SF_REF}&")="'
                   f'&TEXT(J{r}*{SF_REF},"0.00")&"A → {breaker_a}A 선정"')
    ws[f'Q{r}'].font = mfont(size=8, color='8B0000')
    ws[f'Q{r}'].alignment = left_a
    ws[f'Q{r}'].border = thin_b
    ws[f'Q{r}'].fill = mfill(C_BREAKER)
    ws.row_dimensions[r].height = 18

def repeat_headers(ws, row):
    """컬럼 헤더 반복 (섹션 구분용, 단순 1행)"""
    labels = ['판넬','축/유닛명','드라이버/장치 모델','출력W[자동]','정격전류[자동]',
              '상수','R상','S상','T상','설계전류(×수용률)',
              'AC3상','AC단상','DC24V','기타','권장차단기','제조사/모델','비고']
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = mfont(bold=True, size=8, color='FFFFFF')
        c.fill = mfill(C_SUBHDR)
        c.alignment = center
        c.border = thin_b
    ws.row_dimensions[row].height = 14

# ============================================================
# 판넬 1 — 파나소닉 A6 (13축)
# ============================================================
section_header(ws, ROW, '■ 판넬 1 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
P_CLR = 'EBF3FB'
p1_start = ROW

panel1 = [
    ('판넬1','LO_Y1',       'MDDLN55BL', 1,1,1),
    ('판넬1','LO_X1',       'MCDLN35BL', 1,1,1),
    ('판넬1','LO_Z1',       'MADLN05BE', 1,1,1),
    ('판넬1','ULO_Y1',      'MCDLN35BL', 1,1,1),
    ('판넬1','ULO_X1',      'MCDLN35BL', 1,1,1),
    ('판넬1','ULO_Z1',      'MADLN05BE', 1,1,1),
    ('판넬1','NG_EV',       'MCDLN35BE', 1,1,1),
    ('판넬1','NG_FD',       'MADLN05BE', 1,1,1),
    ('판넬1','NG_FD_Z',     'MBDLN25BE', 1,1,1),
    ('판넬1','NG_FD_ALIGN', 'MADLN05BE', 1,1,1),
    ('판넬1','NG_FD_INDEX', 'MBDLN25BE', 1,1,1),
    ('판넬1','JIG_ULO_Y',   'MADLN05BE', 1,1,1),
    ('판넬1','JIG_ULO_Z',   'MADLN05BE', 1,1,1),
]
for panel, axis, model, r, s, t in panel1:
    data_row(ws, ROW, panel, axis, model, r, s, t, ac3=True, fill_color=P_CLR,
             note=f'=IF(D{ROW}="","","출력:"&D{ROW}&"W  정격:"&TEXT(E{ROW},"0.0")&"A  설계:"&TEXT(J{ROW},"0.000")&"A")')
    ROW += 1
p1_end = ROW - 1
sum_row(ws, ROW, '판넬 1 AC합계', p1_start, p1_end, 20, 'LS ABN20c')
p1_sum = ROW
ROW += 1

# 파스텍 (판넬1 소속)
repeat_headers(ws, ROW)
ROW += 1
section_header(ws, ROW, '   ■ 파스텍 EziServo / EziSpeed (AC 단상 입력) — 판넬 1', C_FASTECH)
ROW += 1
F_CLR1 = 'FEF4E8'
fas1_start = ROW
fas1_data = [
    ('판넬1','LO_R1_1', 'Ezi-EC-42XL-A', 0,0,0, 'LO 회전 R1 #1'),
    ('판넬1','LO_R1_2', 'Ezi-EC-42XL-A', 0,0,0, 'LO 회전 R1 #2'),
]
for panel, axis, model, r, s, t, note in fas1_data:
    data_row(ws, ROW, panel, axis, model, r, s, t, ac1=True, fill_color=F_CLR1, note=note)
    ROW += 1
fas1_end = ROW - 1

# 판넬2 헤더 전 repeat
repeat_headers(ws, ROW)
ROW += 1

# ============================================================
# 판넬 2 — 파나소닉 A6 (13축)
# ============================================================
section_header(ws, ROW, '■ 판넬 2 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
P2_CLR = 'DEEAF1'
p2_start = ROW

panel2 = [
    ('판넬2','GOOD_EV',         'MCDLN35BE', 1,1,1),
    ('판넬2','GOOD_FD',         'MADLN05BE', 1,1,1),
    ('판넬2','GOOD_FD_Z',       'MBDLN25BE', 1,1,1),
    ('판넬2','GOOD_FD_ALIGN',   'MADLN05BE', 1,1,1),
    ('판넬2','GOOD_FD_INDEX',   'MBDLN25BE', 1,1,1),
    ('판넬2','TURN_Y',          'MBDLN25BE', 1,1,1),
    ('판넬2','TURN_GRIPPER_Y1', 'MADLN15BE', 1,1,1),
    ('판넬2','TURN_GRIPPER_X1', 'MADLN15BE', 1,1,1),
    ('판넬2','TURN_GRIPPER_Z1', 'MADLN05BE', 1,1,1),
    ('판넬2','TURN_GRIPPER_Y2', 'MADLN15BE', 1,1,1),
    ('판넬2','TURN_GRIPPER_X2', 'MADLN15BE', 1,1,1),
    ('판넬2','TURN_GRIPPER_Z2', 'MADLN05BE', 1,1,1),
    ('판넬2','BOT_MAC_Y',       'MADLN15BE', 1,1,1),
]
for panel, axis, model, r, s, t in panel2:
    data_row(ws, ROW, panel, axis, model, r, s, t, ac3=True, fill_color=P2_CLR)
    ROW += 1
p2_end = ROW - 1
sum_row(ws, ROW, '판넬 2 합계', p2_start, p2_end, 15, 'LS ABN15c')
p2_sum = ROW
ROW += 1
repeat_headers(ws, ROW)
ROW += 1

# ============================================================
# 판넬 3 — 파나소닉 A6 (21축)
# ============================================================
section_header(ws, ROW, '■ 판넬 3 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
P3_CLR = 'EBF3FB'
p3_start = ROW

panel3 = [
    ('판넬3','BOT_MAC_X',   'MADLN15BE', 1,1,1),
    ('판넬3','BOT_MAC_Z',   'MADLN05BE', 1,1,1),
    ('판넬3','BOT_MIC_Y',   'MADLN15BE', 1,1,1),
    ('판넬3','BOT_MIC_X',   'MADLN15BE', 1,1,1),
    ('판넬3','BOT_MIC_Z',   'MADLN05BE', 1,1,1),
    ('판넬3','LO_ATC_X',    'MADLN15BE', 1,1,1),
    ('판넬3','LO_ATC_Z',    'MADLN15BE', 1,1,1),
    ('판넬3','SIDE_X1',     'MADLN05BE', 1,1,1),
    ('판넬3','SIDE_Y1',     'MADLN15BE', 1,1,1),
    ('판넬3','SIDE_T1',     'MADLN05BE', 1,1,1),
    ('판넬3','SIDE_X2',     'MADLN05BE', 1,1,1),
    ('판넬3','SIDE_Y2',     'MADLN15BE', 1,1,1),
    ('판넬3','SIDE_T2',     'MADLN05BE', 1,1,1),
    ('판넬3','SIDE_X3',     'MADLN05BE', 1,1,1),
    ('판넬3','SIDE_Y3',     'MADLN15BE', 1,1,1),
    ('판넬3','SIDE_T3',     'MADLN05BE', 1,1,1),
    ('판넬3','LO_EV',       'MCDLN35BE', 1,1,1),
    ('판넬3','LO_FD',       'MADLN05BE', 1,1,1),
    ('판넬3','LO_FD_Z',     'MBDLN25BE', 1,1,1),
    ('판넬3','LO_FD_ALIGN', 'MADLN05BE', 1,1,1),
    ('판넬3','LO_FD_INDEX', 'MBDLN25BE', 1,1,1),
]
for panel, axis, model, r, s, t in panel3:
    data_row(ws, ROW, panel, axis, model, r, s, t, ac3=True, fill_color=P3_CLR)
    ROW += 1
p3_end = ROW - 1
sum_row(ws, ROW, '판넬 3 합계', p3_start, p3_end, 20, 'LS ABN20c')
p3_sum = ROW
ROW += 1

# 파스텍 (판넬3 소속 — EZI-SPEED)
repeat_headers(ws, ROW)
ROW += 1
section_header(ws, ROW, '   ■ 파스텍 EziSpeed (AC 단상 입력) — 판넬 3 SIDE 검사', C_FASTECH)
ROW += 1
F_CLR3 = 'FEF4E8'
fas3_start = ROW
fas3_data = [
    ('판넬3','EZI_SPEED_1', 'EZI-SPEED60H30CR100P', 0,0,0, 'SIDE 검사 #1'),
    ('판넬3','EZI_SPEED_2', 'EZI-SPEED60H30CR100P', 0,0,0, 'SIDE 검사 #2'),
    ('판넬3','EZI_SPEED_3', 'EZI-SPEED60H30CR100P', 0,0,0, 'SIDE 검사 #3'),
    ('판넬3','EZI_SPEED_4', 'EZI-SPEED60H30CR100P', 0,0,0, 'SIDE 검사 #4'),
    ('판넬3','EZI_SPEED_5', 'EZI-SPEED60H30CR100P', 0,0,0, 'SIDE 검사 #5'),
    ('판넬3','EZI_SPEED_6', 'EZI-SPEED60H30CR100P', 0,0,0, 'SIDE 검사 #6'),
]
for panel, axis, model, r, s, t, note in fas3_data:
    data_row(ws, ROW, panel, axis, model, r, s, t, ac1=True, fill_color=F_CLR3, note=note)
    ROW += 1
fas3_end = ROW - 1
repeat_headers(ws, ROW)
ROW += 1

# ============================================================
# 판넬 4 — 파나소닉 A6 (17축)
# ============================================================
section_header(ws, ROW, '■ 판넬 4 — AC 3상 서보드라이버 (파나소닉 A6)', C_PANASONIC)
ROW += 1
P4_CLR = 'DEEAF1'
p4_start = ROW

panel4 = [
    ('판넬4','TOP_MAC_Y',       'MADLN15BE', 1,1,1),
    ('판넬4','TOP_MAC_X',       'MADLN15BE', 1,1,1),
    ('판넬4','TOP_MAC_Z',       'MADLN05BE', 1,1,1),
    ('판넬4','TOP_MIC_Y',       'MADLN15BE', 1,1,1),
    ('판넬4','TOP_MIC_X',       'MADLN15BE', 1,1,1),
    ('판넬4','TOP_MIC_Z',       'MADLN05BE', 1,1,1),
    ('판넬4','INSPEC_CONV1',    'MADLN15BE', 1,1,1),
    ('판넬4','INSPEC_CONV2',    'MADLN15BE', 1,1,1),
    ('판넬4','JIG_PLATE_CONV1', 'MADLN15BE', 1,1,1),
    ('판넬4','JIG_PLATE_CONV2', 'MADLN15BE', 1,1,1),
    ('판넬4','LO_JIG_TRANS_Y',  'MADLN05BE', 1,1,1),
    ('판넬4','LO_JIG_TRANS_Z',  'MADLN05BE', 1,1,1),
    ('판넬4','TURN_PLATE',      'MADLN05BE', 1,1,1),
    ('판넬4','INSEPC_Y',        'MDDLN55BL', 1,1,1),
    ('판넬4','INSPEC_X',        'MCDLN35BL', 1,1,1),
    ('판넬4','INSPEC_Z1',       'MADLN05BE', 1,1,1),
    ('판넬4','INSPEC_Z2',       'MADLN05BE', 1,1,1),
]
for panel, axis, model, r, s, t in panel4:
    data_row(ws, ROW, panel, axis, model, r, s, t, ac3=True, fill_color=P4_CLR)
    ROW += 1
p4_end = ROW - 1
sum_row(ws, ROW, '판넬 4 AC합계', p4_start, p4_end, 20, 'LS ABN20c')
p4_sum = ROW
ROW += 1

# 파스텍 (판넬4 소속)
section_header(ws, ROW, '   ■ 파스텍 EziServo (AC 단상 입력) — 판넬 4', C_FASTECH)
ROW += 1
F_CLR4 = 'FEF4E8'
fas4_start = ROW
fas4_data = [
    ('판넬4','EZI_EC_XL_1', 'Ezi-EC-42XL-A', 0,0,0, '판넬4 EziServo XL #1'),
    ('판넬4','EZI_EC_XL_2', 'Ezi-EC-42XL-A', 0,0,0, '판넬4 EziServo XL #2'),
    ('판넬4','EZI_EC_XL_3', 'Ezi-EC-42XL-A', 0,0,0, '판넬4 EziServo XL #3'),
    ('판넬4','EZI_EC_XL_4', 'Ezi-EC-42XL-A', 0,0,0, '판넬4 EziServo XL #4'),
    ('판넬4','EZI_EC_M_1',  'Ezi-EC-42M-A',  0,0,0, '판넬4 EziServo M #1'),
    ('판넬4','EZI_EC_M_2',  'Ezi-EC-42M-A',  0,0,0, '판넬4 EziServo M #2'),
]
for panel, axis, model, r, s, t, note in fas4_data:
    data_row(ws, ROW, panel, axis, model, r, s, t, ac1=True, fill_color=F_CLR4, note=note)
    ROW += 1
fas4_end = ROW - 1

# 파스텍 전체 소계 (판넬1+3+4)
sum_row(ws, ROW, '파스텍 전체 소계 (판넬1+3+4)',
        fas1_start, fas1_end, 20, 'LS ABN20c',
        color_sum='FEF4E8',
        extra_ranges=[(fas3_start, fas3_end), (fas4_start, fas4_end)])
fas_all_sum = ROW
ROW += 2

# ============================================================
# DC 24V 부하
# ============================================================
section_header(ws, ROW, '■ DC 24V 부하 — 서보I/O · Brake · 센서 · 솔레노이드 · EZI-SPEED', C_DC)
ROW += 1
DC_CLR = 'EDF7E3'
dc_start = ROW

dc_items = [
    # (panel, axis, model, fixed_e, fixed_j, r, s, t, dc, note)
    ('공통-1P','서보I/O (A6B)', 'A6B I/O',          0.64,0.64, 0,0,0, True,  '64축×10mA=640mA'),
    ('공통-1P','서보Brake',     'Brake',             9.0, 9.0,  0,0,0, True,  '18개×500mA=9000mA'),
    ('공통-1P','리밋센서',      '리밋센서',           0.64,0.64, 0,0,0, True,  '64개×10mA=640mA'),
    ('공통-1P','호밍센서',      '호밍센서',           0.06,0.06, 0,0,0, True,  '6개×10mA=60mA'),
    ('공통-2P','도어',          '도어스위치',         3.0, 3.0,  0,0,0, True,  '10개×300mA=3000mA'),
    ('공통-2P','I/O 모듈',      '파스텍 IO단자대',    2.34,2.34, 0,0,0, True,  '18개×130mA=2340mA'),
    ('공통-2P','입력부하(센서)','포토/근접센서',       2.5, 2.5,  0,0,0, True,  '250개×10mA=2500mA'),
    ('공통-2P','출력부하',      '릴레이·솔레노이드',  5.4, 5.4,  0,0,0, True,  '180개×30mA=5400mA'),
    ('공통-3P','LO_EV_TOP',    'EZI-SPEED60H30CR100P',1.0,1.0, 0,0,0, True,  'LO이젝터 상부'),
    ('공통-3P','LO_EV_BOT',    'EZI-SPEED60H30CR100P',1.0,1.0, 0,0,0, True,  'LO이젝터 하부'),
    ('공통-3P','GOOD_EV_TOP',  'EZI-SPEED60H30CR100P',1.0,1.0, 0,0,0, True,  'GOOD이젝터 상부'),
    ('공통-3P','GOOD_EV_BOT',  'EZI-SPEED60H30CR100P',1.0,1.0, 0,0,0, True,  'GOOD이젝터 하부'),
    ('공통-3P','NG_EV_TOP',    'EZI-SPEED60H30CR100P',1.0,1.0, 0,0,0, True,  'NG이젝터 상부'),
    ('공통-3P','NG_EV_BOT',    'EZI-SPEED60H30CR100P',1.0,1.0, 0,0,0, True,  'NG이젝터 하부'),
]
for panel, axis, model, fe, fj, r, s, t, dc, note in dc_items:
    data_row(ws, ROW, panel, axis, model, r, s, t,
             dc=dc, fill_color=DC_CLR,
             fixed_d=0, fixed_e=fe, fixed_j=fj, note=note)
    ROW += 1
dc_end = ROW - 1

# DC 소계
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  ▶ DC 24V 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill(C_DC)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{dc_start}:E{dc_end})'
ws[f'E{ROW}'].font = mfont(bold=True); ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center; ws[f'E{ROW}'].border = thin_b; ws[f'E{ROW}'].fill = mfill(C_SUM)
ws[f'J{ROW}'] = f'=SUM(J{dc_start}:J{dc_end})'
ws[f'J{ROW}'].font = mfont(bold=True); ws[f'J{ROW}'].number_format = '0.00'
ws[f'J{ROW}'].alignment = center; ws[f'J{ROW}'].border = thin_b; ws[f'J{ROW}'].fill = mfill(C_SUM)
for col in [4,6,7,8,9,11,12,13,14,15,16]:
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
ws[f'Q{ROW}'] = f'DC 전원장치: 1P=10.34A / 2P=13.24A / 3P(EZI-SPEED)=6.0A → 각 SMPS 별도 선정'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].alignment = left_a; ws[f'Q{ROW}'].border = thin_b; ws[f'Q{ROW}'].fill = mfill(C_SUM)
ws.row_dimensions[ROW].height = 17
dc_sum = ROW
ROW += 2

# ============================================================
# 기타 부하 (AC 단상)
# ============================================================
section_header(ws, ROW, '■ 기타 부하 — UPS · 조명CTRL · SMPS (AC 단상 220V)', C_MISC)
ROW += 1
M_CLR = 'F3EBF9'
misc_start = ROW

misc_items = [
    ('공통','UPS 950VA',    'BX950MI-GR',           1.5,1.5, 0,0,0, False,True,  False, 'AC단상 220V'),
    ('공통','UPS 1600VA',   'BX1600MI-GR',          4.5,4.5, 0,0,0, False,True,  False, '3대 각각 별도 차단기'),
    ('공통','조명CTRL',     'POD-22024-4-PEI-LKCB', 5.0,5.0, 0,1,0, False,True,  False, '10개×500mA → 10A MCB'),
    ('공통','SMPS 24V 20A', 'PRO ECO 480W 24V 20A', 2.4,2.4, 0,0,1, False,True,  False, '2대×2.4A'),
    ('공통','SMPS 24V 10A', 'PRO ECO 480W 24V 10A', 1.2,1.2, 1,0,0, False,True,  False, '1대×1.2A'),
    ('공통','HMI/PC',       '(직접 입력)',           0.0,0.0, 0,0,0, False,True,  False, '← 실측값 입력'),
    ('공통','기타',          '(직접 입력)',           0.0,0.0, 0,0,0, False,False, False, '← 실측값 입력'),
]
for panel, axis, model, fe, fj, r, s, t, ac3, ac1, dc, note in misc_items:
    data_row(ws, ROW, panel, axis, model, r, s, t,
             ac3=ac3, ac1=ac1, dc=dc, fill_color=M_CLR,
             fixed_d=0, fixed_e=fe, fixed_j=fj, note=note)
    ROW += 1
misc_end = ROW - 1

ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  ▶ 기타 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill(C_MISC)
ws[f'A{ROW}'].alignment = left_a; ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{misc_start}:E{misc_end})'
ws[f'E{ROW}'].font = mfont(bold=True); ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center; ws[f'E{ROW}'].border = thin_b; ws[f'E{ROW}'].fill = mfill(C_SUM)
ws[f'J{ROW}'] = f'=SUM(J{misc_start}:J{misc_end})'
ws[f'J{ROW}'].font = mfont(bold=True); ws[f'J{ROW}'].number_format = '0.00'
ws[f'J{ROW}'].alignment = center; ws[f'J{ROW}'].border = thin_b; ws[f'J{ROW}'].fill = mfill(C_SUM)
for col in [4,6,7,8,9,11,12,13,14,15,16]:
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
ws[f'Q{ROW}'] = 'UPS 대당 별도 차단기 / 조명CTRL 10A MCB 권장'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].alignment = left_a; ws[f'Q{ROW}'].border = thin_b; ws[f'Q{ROW}'].fill = mfill(C_SUM)
ws.row_dimensions[ROW].height = 17
misc_sum = ROW
ROW += 2

# ============================================================
# 전체 합계 (AC 3상 판넬1+2+3+4)
# ============================================================
section_header(ws, ROW, '■ 전체 합계 — AC 3상 220V (판넬 1+2+3+4 합산)', C_TITLE)
ROW += 1

ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  ▶ AC 3상 전체 합계'
ws[f'A{ROW}'].font = mfont(bold=True, size=10, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill(C_TITLE)
ws[f'A{ROW}'].alignment = left_a; ws[f'A{ROW}'].border = thin_b

ws[f'D{ROW}'].border = thin_b; ws[f'D{ROW}'].fill = mfill(C_SUM)
ws[f'E{ROW}'] = f'=E{p1_sum}+E{p2_sum}+E{p3_sum}+E{p4_sum}'
ws[f'E{ROW}'].font = mfont(bold=True, size=10)
ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center; ws[f'E{ROW}'].border = thin_b; ws[f'E{ROW}'].fill = mfill(C_SUM)

ws[f'J{ROW}'] = f'=J{p1_sum}+J{p2_sum}+J{p3_sum}+J{p4_sum}'
ws[f'J{ROW}'].font = mfont(bold=True, size=10)
ws[f'J{ROW}'].number_format = '0.00'
ws[f'J{ROW}'].alignment = center; ws[f'J{ROW}'].border = thin_b; ws[f'J{ROW}'].fill = mfill(C_SUM)

for col in [6,7,8,9,11,12,13,14]:
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)

ws[f'O{ROW}'] = '→ 메인 차단기'
ws[f'O{ROW}'].font = mfont(bold=True, size=10, color='1F3864')
ws[f'O{ROW}'].alignment = center; ws[f'O{ROW}'].border = thin_b; ws[f'O{ROW}'].fill = mfill(C_BREAKER)
ws[f'P{ROW}'] = 'LS ABN75c'
ws[f'P{ROW}'].font = mfont(bold=True, size=9)
ws[f'P{ROW}'].alignment = center; ws[f'P{ROW}'].border = thin_b; ws[f'P{ROW}'].fill = mfill(C_BREAKER)
ws[f'Q{ROW}'] = (f'=TEXT(J{ROW},"0.00")&"A × 안전계수("&{SF_REF}&") = "'
                 f'&TEXT(J{ROW}*{SF_REF},"0.00")&"A  |  kVA = "'
                 f'&TEXT(E{ROW}*220*SQRT(3)/1000,"0.00")&"kVA  (역률반영: kW="'
                 f'&TEXT(E{ROW}*220*SQRT(3)/1000*{PF_REF},"0.00")&"kW)"')
ws[f'Q{ROW}'].font = mfont(size=8, color='8B0000')
ws[f'Q{ROW}'].alignment = left_a; ws[f'Q{ROW}'].border = thin_b; ws[f'Q{ROW}'].fill = mfill(C_BREAKER)
ws.row_dimensions[ROW].height = 22
total_row = ROW

# ── 상단 요약박스 수식 연결 (사후 채움) ──
# AC 3상 요약 (row 6)
ws['B6'] = f'=E{p1_sum}+E{p2_sum}+E{p3_sum}+E{p4_sum}'
ws['C6'] = f'=B6*{SF_REF}'
ws['D6'] = '→ Sheet4 참조'
ws['E6'] = f'=B6*220*SQRT(3)/1000'
ws['F6'] = f'=TEXT(B6,"0.00")&"A × {SF_REF} = "&TEXT(C6,"0.00")&"A 이상 선정"'
for col in 'BCDEF':
    ws[f'{col}6'].font = mfont(size=9)
    ws[f'{col}6'].alignment = center
    ws[f'{col}6'].border = thin_b
    ws[f'{col}6'].fill = mfill('EBF3FB')
# AC 단상 요약 (row 7 — 파스텍)
ws['B7'] = f'=E{fas_all_sum}'
ws['C7'] = f'=B7*{SF_REF}'
ws['D7'] = '→ Sheet4 참조'
ws['E7'] = f'=B7*220/1000'
ws['F7'] = f'=TEXT(B7,"0.00")&"A × {SF_REF} = "&TEXT(C7,"0.00")&"A 이상 선정"'
for col in 'BCDEF':
    ws[f'{col}7'].font = mfont(size=9)
    ws[f'{col}7'].alignment = center
    ws[f'{col}7'].border = thin_b
    ws[f'{col}7'].fill = mfill('FEF4E8')
# DC 24V 요약 (row 8)
ws['B8'] = f'=E{dc_sum}'
ws['C8'] = f'=B8*{SF_REF}'
ws['D8'] = 'SMPS 별도'
ws['E8'] = f'=B8*24/1000'
ws['F8'] = 'DC 전원장치 용량: 합계전류 × 1.25 이상'
for col in 'BCDEF':
    ws[f'{col}8'].font = mfont(size=9)
    ws[f'{col}8'].alignment = center
    ws[f'{col}8'].border = thin_b
    ws[f'{col}8'].fill = mfill('EDF7E3')
# 기타 요약 (row 9)
ws['B9'] = f'=E{misc_sum}'
ws['C9'] = f'=B9*{SF_REF}'
ws['D9'] = '→ Sheet4 참조'
ws['E9'] = f'=B9*220/1000'
ws['F9'] = f'=TEXT(B9,"0.00")&"A × {SF_REF} = "&TEXT(C9,"0.00")&"A 이상 선정"'
for col in 'BCDEF':
    ws[f'{col}9'].font = mfont(size=9)
    ws[f'{col}9'].alignment = center
    ws[f'{col}9'].border = thin_b
    ws[f'{col}9'].fill = mfill('F3EBF9')

# ============================================================
# SHEET 2: 파나소닉A6규격 (VLOOKUP 소스) — 사용자 파일과 동일하게
# ============================================================
ws2 = wb.create_sheet('파나소닉A6규격')
ws2.merge_cells('A1:D1')
ws2['A1'] = '파나소닉 A6 서보드라이버 정격입력전류 (카탈로그 기준, 3상 AC200V)'
ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws2['A1'].alignment = center
ws2['A1'].fill = mfill('D6E4F0')
ws2.row_dimensions[1].height = 22

for j, h in enumerate(['드라이버 모델','정격출력 (W)','정격입력전류 (A)','비고'], 1):
    c = ws2.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center; c.border = thin_b
ws2.row_dimensions[2].height = 20

pan_lookup = [
    ('MADLN05BE',  50, 0.3, '50W 서보드라이버'),
    ('MADLN10BE', 100, 0.5, '100W 서보드라이버'),
    ('MADLN15BE', 200, 1.0, '200W 서보드라이버'),
    ('MBDLN25BE', 400, 1.5, '400W 서보드라이버'),
    ('MCDLN35BL', 750, 2.4, '750W 서보드라이버 (BL)'),
    ('MCDLN35BE', 750, 2.4, '750W 서보드라이버 (BE)'),
    ('MDDLN55BL',1000, 3.4, '1kW 서보드라이버 (BL)'),
    ('MDDLN55BE',1000, 3.4, '1kW 서보드라이버 (BE)'),
    ('MDDLN85BL',1500, 5.5, '1.5kW 서보드라이버 (BL)'),
]
for i, (model, watt, curr, note) in enumerate(pan_lookup, 3):
    fill = mfill('EBF3FB') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate([model, watt, curr, note], 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j <= 3 else left_a
        c.border = thin_b; c.fill = fill
    ws2.row_dimensions[i].height = 16

note_r = len(pan_lookup) + 4
ws2.merge_cells(f'A{note_r}:D{note_r}')
ws2[f'A{note_r}'] = '※ A열 모델명과 메인시트 C열 코드 정확히 일치 필요  |  카탈로그 정격입력전류 기준  |  모델 추가 시 이 시트에 행 추가'
ws2[f'A{note_r}'].font = mfont(size=8, color='555555')
ws2[f'A{note_r}'].fill = mfill('FFF2CC')
ws2[f'A{note_r}'].alignment = left_a

# 하단 상세 규격표 (사용자 파일 그대로)
detail_r = note_r + 2
ws2.merge_cells(f'A{detail_r}:J{detail_r}')
ws2[f'A{detail_r}'] = '파나소닉 A6 서보드라이버 규격표 (MINAS A6 Series)'
ws2[f'A{detail_r}'].font = Font(name='맑은 고딕', bold=True, size=10, color=C_TITLE)
ws2[f'A{detail_r}'].alignment = center
ws2[f'A{detail_r}'].fill = mfill('D6E4F0')
ws2.row_dimensions[detail_r].height = 18
detail_r += 1

detail_headers = ['드라이버 모델','대응 모터','정격출력(W)','전원','정격입력전류(A)',
                  '최대입력전류(A)','제어방식','통신','특이사항','비고']
for j, h in enumerate(detail_headers, 1):
    c = ws2.cell(row=detail_r, column=j, value=h)
    c.font = mfont(bold=True, size=8, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center; c.border = thin_b

detail_data = [
    ('MADLN05BE','MSMF/MQMF 50W',   50, 'AC3상 200V',0.3, 0.8, 'Full-Closed','EtherCAT/Modbus','소형 단축용',None),
    ('MADLN10BE','MSMF/MQMF 100W', 100, 'AC3상 200V',0.5, 1.2, 'Full-Closed','EtherCAT/Modbus',None,None),
    ('MADLN15BE','MSMF/MQMF 200W', 200, 'AC3상 200V',1.0, 2.4, 'Full-Closed','EtherCAT/Modbus',None,None),
    ('MBDLN25BE','MSMF/MQMF 400W', 400, 'AC3상 200V',1.5, 4.0, 'Full-Closed','EtherCAT/Modbus',None,None),
    ('MCDLN35BL','MSMF/MQMF 750W', 750, 'AC3상 200V',2.4, 7.0, 'Full-Closed','EtherCAT/Modbus','중형 축용',None),
    ('MCDLN35BE','MSMF/MQMF 750W', 750, 'AC3상 200V',2.4, 7.0, 'Full-Closed','EtherCAT/Modbus','Brake 내장형',None),
    ('MDDLN55BL','MSMF 1kW',      1000, 'AC3상 200V',3.4,10.0, 'Full-Closed','EtherCAT/Modbus','대형 주축용',None),
    ('MDDLN85BL','MSMF 1.5kW',    1500, 'AC3상 200V',5.5,15.0, 'Full-Closed','EtherCAT/Modbus',None,None),
    ('MEDLN15BL','MSMF 2kW',      2000, 'AC3상 200V',7.5,20.0, 'Full-Closed','EtherCAT/Modbus',None,None),
    ('MFDLN23BL','MSMF 3kW',      3000, 'AC3상 200V',11.0,30.0,'Full-Closed','EtherCAT/Modbus','고출력',None),
]
for i, row_data in enumerate(detail_data, detail_r+1):
    fill = mfill('EBF3FB') if i % 2 == 0 else mfill('FFFFFF')
    for j, val in enumerate(row_data, 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = mfont(size=8)
        c.alignment = center if j != 9 else left_a
        c.border = thin_b; c.fill = fill
    ws2.row_dimensions[i].height = 14

note2_r = detail_r + len(detail_data) + 1
ws2.merge_cells(f'A{note2_r}:J{note2_r}')
ws2[f'A{note2_r}'] = '※ 정격입력전류: I = P_출력(VA) ÷ (√3 × V_입력 × PF)  |  안전계수 1.25 적용 후 차단기 선정  |  출처: Panasonic MINAS A6 Catalog'
ws2[f'A{note2_r}'].font = mfont(size=8, color='555555')
ws2[f'A{note2_r}'].fill = mfill('FFF2CC')
ws2[f'A{note2_r}'].alignment = left_a

for col, w in [('A',18),('B',18),('C',12),('D',14),('E',14),('F',14),('G',12),('H',16),('I',14),('J',12)]:
    ws2.column_dimensions[col].width = w

# ============================================================
# SHEET 3: 파스텍EziServo규격
# ============================================================
ws3 = wb.create_sheet('파스텍EziServo규격')
ws3.merge_cells('A1:C1')
ws3['A1'] = '파스텍 EZS2/EziServo 정격전류 (VLOOKUP 소스)'
ws3['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws3['A1'].alignment = center; ws3['A1'].fill = mfill('FCE4D6')
ws3.row_dimensions[1].height = 22

for j, h in enumerate(['드라이버 모델','정격전류 (A)','비고'], 1):
    c = ws3.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER); c.alignment = center; c.border = thin_b
ws3.row_dimensions[2].height = 18

fas_lookup = [
    ('EZS2-EC-42M',          1.2, 'DC 24V 입력 스텝서보'),
    ('EZS2-EC-42XL',         1.2, 'DC 24V 입력 스텝서보 (XL)'),
    ('Ezi-EC-42M-A',         1.5, 'AC 입력 스텝서보'),
    ('Ezi-EC-42XL-A',        2.0, 'AC 입력 스텝서보 (XL)'),
    ('EZI-SPEED60H30CR100P', 1.0, 'DC 24V 입력 스텝모터드라이버'),
]
for i, (model, curr, note) in enumerate(fas_lookup, 3):
    fill = mfill('FEF4E8') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate([model, curr, note], 1):
        c = ws3.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j <= 2 else left_a
        c.border = thin_b; c.fill = fill
    ws3.row_dimensions[i].height = 16

for col, w in [('A',26),('B',16),('C',30)]:
    ws3.column_dimensions[col].width = w

# ============================================================
# SHEET 4: 차단기 규격 참조표
# ============================================================
ws4 = wb.create_sheet('차단기 규격 참조표')
ws4.merge_cells('A1:E1')
ws4['A1'] = '차단기 규격 참조표 (MCCB / MCB 3P, AC 220V)'
ws4['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws4['A1'].alignment = center
ws4.row_dimensions[1].height = 22; ws4.row_dimensions[2].height = 5

for j, h in enumerate(['정격전류 (A)','적용 부하전류 범위 (A)','제조사','모델 (3P)','비고'], 1):
    c = ws4.cell(row=3, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER); c.alignment = center; c.border = thin_b
ws4.row_dimensions[3].height = 28

breaker_data = [
    [3,  '~2.4A',   'LS/Chint','ABN3c / NXB-63',   '조명, 소형 제어회로'],
    [5,  '2.4~4A',  'LS/Chint','ABN5c / NXB-63',   '소형 I/O, 팬'],
    [10, '4~8A',    'LS/Chint','ABN10c / NXB-63',  '소형 서보 1~2축'],
    [15, '8~12A',   'LS/Chint','ABN15c / NXB-63',  '판넬2 권장'],
    [20, '12~16A',  'LS/Chint','ABN20c / NXB-63',  '판넬1/3/4 권장'],
    [25, '16~20A',  'LS/Chint','ABN25c / NXB-63',  ''],
    [30, '20~24A',  'LS/Chint','ABN30c / NXB-63',  '다축 서보 판넬'],
    [40, '24~32A',  'LS/Chint','ABN40c',             ''],
    [50, '32~40A',  'LS/Chint','ABN50c',             ''],
    [63, '40~50A',  'LS/Chint','ABN63c',             '대형 판넬'],
    [75, '50~60A',  'LS/Chint','ABN75c',             '메인 차단기 (판넬 합산)'],
    [100,'60~80A',  'LS/Chint','ABN100c',            ''],
]
for i, row in enumerate(breaker_data, 4):
    ws4.row_dimensions[i].height = 16
    for j, val in enumerate(row, 1):
        c = ws4.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j <= 4 else left_a
        c.border = thin_b
        if j == 1: c.fill = mfill(C_SUM)

for col, w in [('A',14),('B',20),('C',12),('D',20),('E',24)]:
    ws4.column_dimensions[col].width = w

out_path = 'C:/MES/wta-agents/workspaces/control-agent/대구텍_F2_1_판넬별_차단기선정표_v3.xlsx'
wb.save(out_path)
print('저장 완료:', out_path)
print(f'판넬1({len(panel1)}축) / 판넬2({len(panel2)}축) / 판넬3({len(panel3)}축) / 판넬4({len(panel4)}축)')
print(f'파스텍: 판넬1({len(fas1_data)}) + 판넬3({len(fas3_data)}) + 판넬4({len(fas4_data)}) = {len(fas1_data)+len(fas3_data)+len(fas4_data)}축')
print(f'DC({len(dc_items)}항목) / 기타({len(misc_items)}항목)')
print('역률=$D$2 / 안전계수=$F$2 / 수용률=$H$2 - 모든 수식에 셀 참조 반영')
