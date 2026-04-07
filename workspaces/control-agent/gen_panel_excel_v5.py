"""
대구텍 검사기 F2 #1 판넬별 전류·용량 계산 및 차단기 선정표 v5
- 형식: WTA v2 형식 완전 유지
- 데이터: 전장설계 용량계산,전류분배 파일 기준으로 전체 반영
- 전류 계산: I = VA / (√3 × 220V × PF=0.95)
- 설계전류(운전전류): 정격전류 × 수용률 0.5
- VLOOKUP: 드라이버 모델 → 정격전류 자동조회
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 공통 스타일 =====
def mfont(bold=False, size=9, color='000000', name='맑은 고딕'):
    return Font(name=name, bold=bold, size=size, color=color)
def mfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)
def mborder(style='thin', color='AAAAAA'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def mcenter(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def mleft(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

C_HEADER   = '2F5496'
C_SUBHDR   = '4472C4'
C_TITLE    = '1F3864'
C_PANASONIC= '00B0F0'
C_FASTECH  = 'ED7D31'
C_DC       = '70AD47'
C_MISC     = '7030A0'
C_SUM      = 'E2EFDA'
C_FORMULA  = 'FFF2CC'

thin_b = mborder('thin')
center = mcenter()
left_a = mleft()

# ===== SHEET 1: 종합 전류·용량 계산표 =====
ws = wb.active
ws.title = '전류·용량 계산표'

col_widths = {1:5,2:20,3:24,4:8,5:10,6:6,7:5,8:9,9:9,10:9,11:9,12:9,13:9,14:9,15:9,16:9,17:32}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# 1행 제목
ws.merge_cells('A1:Q1')
c = ws['A1']
c.value = '대구텍 검사기 F2 #1 — 전류·용량 계산 및 차단기 선정 종합표'
c.font = Font(name='맑은 고딕', bold=True, size=14, color=C_TITLE)
c.alignment = center
c.fill = mfill('D9E2F3')
ws.row_dimensions[1].height = 28

# 2행 요약 헤더
ws.merge_cells('A2:Q2')
ws['A2'].value = '▶ 전체 부하 요약 (상단 자동계산)'
ws['A2'].font = mfont(bold=True, size=10, color='FFFFFF')
ws['A2'].fill = mfill(C_TITLE)
ws['A2'].alignment = center
ws.row_dimensions[2].height = 18

# 3행 요약 컬럼 헤더
for col, label in [(1,'구분'),(2,'총 정격전류 (A)'),(3,'안전계수 ×1.25'),(4,'권장 차단기 (A)'),(5,'비고')]:
    c = ws.cell(row=3, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[3].height = 18

# 4~7행 요약 데이터
summary_rows = [
    (4, 'AC 3상 220V — 파나소닉 A6 서보드라이버', mfill('EBF3FB')),
    (5, 'AC 단상 220V — EZS2 스텝드라이버 (DC 전원 별도)',  mfill('FEF4E8')),
    (6, 'DC 24V — 서보I/O·Brake·센서·솔레노이드·EZI-SPEED', mfill('EDF7E3')),
    (7, '기타 부하 — UPS·조명CTRL·SMPS', mfill('F3EBF9')),
]
for row, label, fill in summary_rows:
    ws.cell(row=row, column=1, value=label).fill = fill
    ws.cell(row=row, column=1).font = mfont(bold=True, size=9)
    ws.cell(row=row, column=1).alignment = left_a
    ws.cell(row=row, column=1).border = thin_b
    for col in range(2, 6):
        c = ws.cell(row=row, column=col, value='→ 하단 계산')
        c.fill = fill
        c.font = mfont(size=9, color='888888')
        c.alignment = center
        c.border = thin_b
    ws.row_dimensions[row].height = 16

ws.row_dimensions[8].height = 6

# 9행 R/S/T 분석
ws.merge_cells('A9:Q9')
ws['A9'].value = '▶ R/S/T 상별 전류 편중 분석'
ws['A9'].font = mfont(bold=True, size=10, color='FFFFFF')
ws['A9'].fill = mfill('833C00')
ws['A9'].alignment = center
ws.row_dimensions[9].height = 18

for i, h in enumerate(['','R상 (A)','S상 (A)','T상 (A)','최대-최소 (A)','불균형률 (%)','판정','기준: 불균형률 10% 이하 권장'], 1):
    c = ws.cell(row=10, column=i, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill('C55A11')
    c.alignment = center
    c.border = thin_b
ws.row_dimensions[10].height = 18

for i, label in enumerate(['AC 3상 서보드라이버','AC 3상 기타','전체 합계'], 11):
    ws.cell(row=i, column=1, value=label).font = mfont(bold=True, size=9)
    ws.cell(row=i, column=1).border = thin_b
    ws.cell(row=i, column=1).alignment = left_a
    for col in range(2, 9):
        ws.cell(row=i, column=col).border = thin_b
        ws.cell(row=i, column=col).alignment = center
    ws.row_dimensions[i].height = 16

ws.row_dimensions[14].height = 6

# 15~16행 메인 헤더
for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
    ws.merge_cells(f'{col_letter}15:{col_letter}16')

for col, label in [(1,'판넬'),(2,'축/유닛명'),(3,'드라이버/장치 모델'),(4,'출력\n(W)'),(5,'정격전류\n(A)'),(6,'상수'),(7,'R상\n할당'),(8,'S상\n할당'),(9,'T상\n할당'),(10,'설계전류\n(A)')]:
    c = ws.cell(row=15, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b

ws.merge_cells('K15:N15')
ws['K15'] = '전원 구분 (체크)'
ws['K15'].font = mfont(bold=True, size=9, color='FFFFFF')
ws['K15'].fill = mfill(C_SUBHDR)
ws['K15'].alignment = center
ws['K15'].border = thin_b

for col, label in [(11,'AC 3상\n220V'),(12,'AC 단상\n220V'),(13,'DC 24V'),(14,'기타')]:
    c = ws.cell(row=16, column=col, value=label)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_SUBHDR)
    c.alignment = center
    c.border = thin_b

for col_letter, label in [('O','차단기\n선정 (A)'),('P','제조사/\n모델'),('Q','비고 / 수식 설명')]:
    ws.merge_cells(f'{col_letter}15:{col_letter}16')
    ws[f'{col_letter}15'] = label
    ws[f'{col_letter}15'].font = mfont(bold=True, size=9, color='FFFFFF')
    ws[f'{col_letter}15'].fill = mfill(C_HEADER)
    ws[f'{col_letter}15'].alignment = center if col_letter != 'Q' else left_a
    ws[f'{col_letter}15'].border = thin_b

ws.row_dimensions[15].height = 20
ws.row_dimensions[16].height = 24

ROW = 17

def data_row(ws, r, panel, axis, model, watt, rated_i, phase,
             r_flag, s_flag, t_flag, op_i,
             ac3=False, ac1=False, dc=False, misc=False,
             breaker='', mfr='', note='', fill_color=None):
    fill = mfill(fill_color) if fill_color else None
    vals = [panel, axis, model, watt, rated_i, phase,
            r_flag, s_flag, t_flag, op_i,
            '●' if ac3 else '', '●' if ac1 else '',
            '●' if dc else '', '●' if misc else '',
            breaker, mfr, note]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = mfont(size=9)
        c.alignment = center if i not in (1,2,3,17) else (left_a if i in (2,3,17) else center)
        c.border = thin_b
        if fill: c.fill = fill
        if i == 1: c.font = mfont(bold=True, size=9, color=C_TITLE)
    ws.row_dimensions[r].height = 17

def section_header(ws, r, label, color):
    ws.merge_cells(f'A{r}:Q{r}')
    c = ws[f'A{r}']
    c.value = label
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(color)
    c.alignment = left_a
    c.border = thin_b
    ws.row_dimensions[r].height = 16

def vlookup_e(r):
    return f'=IFERROR(VLOOKUP(C{r},파나소닉A6규격!$A:$C,3,FALSE),IFERROR(VLOOKUP(C{r},파스텍EziServo규격!$A:$C,3,FALSE),""))'

def set_e_j(ws, r, util, fill_color):
    fill = mfill(fill_color)
    ws[f'E{r}'] = vlookup_e(r)
    ws[f'E{r}'].font = mfont(size=9, color='1F3864')
    ws[f'E{r}'].alignment = center
    ws[f'E{r}'].border = thin_b
    ws[f'E{r}'].fill = fill
    ws[f'J{r}'] = f'=IF(E{r}="","",ROUND(E{r}*{util},3))'
    ws[f'J{r}'].font = mfont(size=9, color='375623')
    ws[f'J{r}'].alignment = center
    ws[f'J{r}'].border = thin_b
    ws[f'J{r}'].fill = fill

# ============================================================
# 섹션 1: 파나소닉 A6 서보드라이버
# ============================================================
section_header(ws, ROW, '■ 파나소닉 A6 서보드라이버 (AC 3상 220V)', C_PANASONIC)
ROW += 1
PAN_COLOR = 'EBF3FB'

# (판넬, 축명, 드라이버모델, 출력W, R, S, T, 수용률, 비고)
# 전류값: VLOOKUP → 파나소닉A6규격 시트에서 자동조회 (카탈로그 기준)
# 설계전류 = 정격전류 × 수용률 0.7
panasonic_data = [
    # ── 판넬 1 (13축) ──
    ('판넬1', 'LO_Y1',          'MDDLN55BL', 1000, 1,1,1, 0.7, ''),
    ('판넬1', 'LO_X1',          'MCDLN35BL',  750, 1,1,1, 0.7, ''),
    ('판넬1', 'LO_Z1',          'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬1', 'ULO_Y1',         'MCDLN35BL',  750, 1,1,1, 0.7, ''),
    ('판넬1', 'ULO_X1',         'MCDLN35BL',  750, 1,1,1, 0.7, ''),
    ('판넬1', 'ULO_Z1',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬1', 'NG_EV',          'MCDLN35BE',  750, 1,1,1, 0.7, ''),
    ('판넬1', 'NG_FD',          'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬1', 'NG_FD_Z',        'MBDLN25BE',  400, 1,1,1, 0.7, ''),
    ('판넬1', 'NG_FD_ALIGN',    'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬1', 'NG_FD_INDEX',    'MBDLN25BE',  400, 1,1,1, 0.7, ''),
    ('판넬1', 'JIG_ULO_Y',      'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬1', 'JIG_ULO_Z',      'MADLN05BE',   50, 1,1,1, 0.7, ''),
    # ── 판넬 2 (13축) ──
    ('판넬2', 'GOOD_EV',         'MCDLN35BE',  750, 1,1,1, 0.7, ''),
    ('판넬2', 'GOOD_FD',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬2', 'GOOD_FD_Z',       'MBDLN25BE',  400, 1,1,1, 0.7, ''),
    ('판넬2', 'GOOD_FD_ALIGN',   'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬2', 'GOOD_FD_INDEX',   'MBDLN25BE',  400, 1,1,1, 0.7, ''),
    ('판넬2', 'TURN_Y',          'MBDLN25BE',  400, 1,1,1, 0.7, ''),
    ('판넬2', 'TURN_GRIPPER_Y1', 'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬2', 'TURN_GRIPPER_X1', 'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬2', 'TURN_GRIPPER_Z1', 'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬2', 'TURN_GRIPPER_Y2', 'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬2', 'TURN_GRIPPER_X2', 'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬2', 'TURN_GRIPPER_Z2', 'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬2', 'BOT_MAC_Y',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    # ── 판넬 3 (21축) ──
    ('판넬3', 'BOT_MAC_X',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'BOT_MAC_Z',       'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'BOT_MIC_Y',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'BOT_MIC_X',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'BOT_MIC_Z',       'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'LO_ATC_X',        'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'LO_ATC_Z',        'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_X1',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_Y1',         'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_T1',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_X2',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_Y2',         'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_T2',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_X3',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_Y3',         'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬3', 'SIDE_T3',         'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'LO_EV',           'MCDLN35BE',  750, 1,1,1, 0.7, ''),
    ('판넬3', 'LO_FD',           'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'LO_FD_Z',         'MBDLN25BE',  400, 1,1,1, 0.7, ''),
    ('판넬3', 'LO_FD_ALIGN',     'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬3', 'LO_FD_INDEX',     'MBDLN25BE',  400, 1,1,1, 0.7, ''),
    # ── 판넬 4 (17축) ──
    ('판넬4', 'TOP_MAC_Y',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'TOP_MAC_X',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'TOP_MAC_Z',       'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬4', 'TOP_MIC_Y',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'TOP_MIC_X',       'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'TOP_MIC_Z',       'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬4', 'INSPEC_CONV1',    'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'INSPEC_CONV2',    'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'JIG_PLATE_CONV1', 'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'JIG_PLATE_CONV2', 'MADLN15BE',  200, 1,1,1, 0.7, ''),
    ('판넬4', 'LO_JIG_TRANS_Y',  'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬4', 'LO_JIG_TRANS_Z',  'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬4', 'TURN_PLATE',      'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬4', 'INSEPC_Y',        'MDDLN55BL', 1000, 1,1,1, 0.7, ''),
    ('판넬4', 'INSPEC_X',        'MCDLN35BL',  750, 1,1,1, 0.7, ''),
    ('판넬4', 'INSPEC_Z1',       'MADLN05BE',   50, 1,1,1, 0.7, ''),
    ('판넬4', 'INSPEC_Z2',       'MADLN05BE',   50, 1,1,1, 0.7, ''),
]

pan_start = ROW
for d in panasonic_data:
    panel, axis, model, watt, r, s, t, util, note = d
    data_row(ws, ROW, panel, axis, model, watt, 0, 3, r, s, t, 0,
             ac3=True, note=note, fill_color=PAN_COLOR)
    set_e_j(ws, ROW, util, PAN_COLOR)
    ROW += 1

pan_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  파나소닉 A6 소계  (총 64축)'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{pan_start}:E{pan_end})'
ws[f'E{ROW}'].number_format = '0.000'
ws[f'J{ROW}'] = f'=SUM(J{pan_start}:J{pan_end})'
ws[f'J{ROW}'].number_format = '0.000'
for col in [4,5,6,7,8,9,10,11,12,13,14,15,16]:
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = f'정격전류 합계={pan_end-pan_start+1}축 / 설계전류=정격×0.5 (SUM 수식)'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
pan_sum_row = ROW
ROW += 2

# ============================================================
# 섹션 2: 파스텍 EZS2 스텝서보드라이버 (DC 24V 입력)
# ============================================================
section_header(ws, ROW, '■ 파스텍 EZS2 스텝서보드라이버 (DC 24V 입력) — DC 전원장치 용량 계산용', C_FASTECH)
ROW += 1
FAS_COLOR = 'FEF4E8'

# (판넬, 축명, 모델, W, R, S, T, util, 비고)
fastech_data = [
    # 판넬1
    ('판넬1', 'LO_R1',          'EZS2-EC-42M',  0, 0,0,0, 1.0, 'LO 회전축 — DC 24V 1.2A'),
    ('판넬1', 'ULO_R2',         'EZS2-EC-42M',  0, 0,0,0, 1.0, 'ULO 회전축 — DC 24V 1.2A'),
    # 판넬2
    ('판넬2', 'TURN_R1',        'EZS2-EC-42M',  0, 0,0,0, 1.0, 'TURN 회전1'),
    ('판넬2', 'TURN_R2',        'EZS2-EC-42M',  0, 0,0,0, 1.0, 'TURN 회전2'),
    ('판넬2', 'TURN_GRIPPER_R1','EZS2-EC-42XL', 0, 0,0,0, 1.0, 'TURN 그리퍼R1'),
    ('판넬2', 'TURN_GRIPPER_R2','EZS2-EC-42XL', 0, 0,0,0, 1.0, 'TURN 그리퍼R2'),
    # 판넬4
    ('판넬4', 'INSPEC_R1',      'EZS2-EC-42M',  0, 0,0,0, 1.0, 'INSPEC 회전1'),
    ('판넬4', 'INSPEC_R2',      'EZS2-EC-42M',  0, 0,0,0, 1.0, 'INSPEC 회전2'),
]

fas_start = ROW
for d in fastech_data:
    panel, axis, model, watt, r, s, t, util, note = d
    data_row(ws, ROW, panel, axis, model, watt, 0, 1, r, s, t, 0,
             ac3=False, ac1=False, dc=True, note=note, fill_color=FAS_COLOR)
    set_e_j(ws, ROW, util, FAS_COLOR)
    ROW += 1

fas_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  EZS2 스텝서보 소계  (총 8축)'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{fas_start}:E{fas_end})'
ws[f'J{ROW}'] = f'=SUM(J{fas_start}:J{fas_end})'
for col in range(4, 17):
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = 'EZS2: DC 24V 입력 — DC 전원장치(SMPS) 용량에 반영'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
fas_sum_row = ROW
ROW += 2

# ============================================================
# 섹션 3: DC 24V 부하
# ============================================================
section_header(ws, ROW, '■ DC 24V 부하 (서보I/O·Brake·센서·솔레노이드·I/O모듈·EZI-SPEED)', C_DC)
ROW += 1
DC_COLOR = 'EDF7E3'

# (panel, axis, model, watt, rated_i, ph, r, s, t, op_i, ac3, ac1, dc, note)
# rated_i 단위: A (mA에서 변환)
dc_data = [
    # 1P24/1N24 — 서보 컨트롤러 전원
    ('공통-1P', '서보I/O (A6B)',    'A6B I/O',          0, 0.64,  1, 0,0,0, 0.64,  False, False, True, '64축×10mA=640mA'),
    ('공통-1P', '서보Brake',        'Brake',            0, 9.0,   1, 0,0,0, 9.0,   False, False, True, '18개×500mA=9000mA'),
    ('공통-1P', '리밋센서',         '리밋센서',         0, 0.64,  1, 0,0,0, 0.64,  False, False, True, '64개×10mA=640mA'),
    ('공통-1P', '호밍센서',         '호밍센서',         0, 0.06,  1, 0,0,0, 0.06,  False, False, True, '6개×10mA=60mA'),
    # 2P24/2N24 — I/O 전원
    ('공통-2P', '도어',             '도어스위치',       0, 3.0,   1, 0,0,0, 3.0,   False, False, True, '10개×300mA=3000mA'),
    ('공통-2P', 'I/O 모듈',         '파스텍 IO단자대',  0, 2.34,  1, 0,0,0, 2.34,  False, False, True, '18개×130mA=2340mA'),
    ('공통-2P', '입력부하 (센서)',   '포토/근접센서',    0, 2.5,   1, 0,0,0, 2.5,   False, False, True, '250개×10mA=2500mA'),
    ('공통-2P', '출력부하',         '릴레이·솔레노이드',0, 5.4,   1, 0,0,0, 5.4,   False, False, True, '180개×30mA=5400mA'),
    # 3P24/3N24 — EZI-SPEED 구동용
    ('공통-3P', 'LO_EV_TOP',        'EZI-SPEED60H30CR100P', 0, 1.0, 1, 0,0,0, 1.0, False, False, True, 'LO이젝터 상부'),
    ('공통-3P', 'LO_EV_BOT',        'EZI-SPEED60H30CR100P', 0, 1.0, 1, 0,0,0, 1.0, False, False, True, 'LO이젝터 하부'),
    ('공통-3P', 'GOOD_EV_TOP',      'EZI-SPEED60H30CR100P', 0, 1.0, 1, 0,0,0, 1.0, False, False, True, 'GOOD이젝터 상부'),
    ('공통-3P', 'GOOD_EV_BOT',      'EZI-SPEED60H30CR100P', 0, 1.0, 1, 0,0,0, 1.0, False, False, True, 'GOOD이젝터 하부'),
    ('공통-3P', 'NG_EV_TOP',        '(모델 확인 필요)',      0, 1.0, 1, 0,0,0, 1.0, False, False, True, 'NG이젝터 상부 ← 모델 미지정'),
    ('공통-3P', 'NG_EV_BOT',        '(모델 확인 필요)',      0, 1.0, 1, 0,0,0, 1.0, False, False, True, 'NG이젝터 하부 ← 모델 미지정'),
]
dc_start = ROW
for d in dc_data:
    panel, axis, model, watt, ri, ph, r, s, t, oi, ac3, ac1, dc, note = d
    data_row(ws, ROW, panel, axis, model, watt, ri, ph, r, s, t, oi,
             ac3=ac3, ac1=ac1, dc=dc, note=note, fill_color=DC_COLOR)
    ROW += 1

dc_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  DC 24V 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{dc_start}:E{dc_end})'
ws[f'J{ROW}'] = f'=SUM(J{dc_start}:J{dc_end})'
for col in range(4, 17):
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = 'DC 전원장치 선정: 합계×1.25 이상. 1P총=10.34A, 2P총=13.24A, 3P총=6.0A'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
dc_sum_row = ROW
ROW += 2

# ============================================================
# 섹션 4: 기타 부하 (UPS, 조명CTRL, SMPS)
# ============================================================
section_header(ws, ROW, '■ 기타 부하 (UPS · 조명CTRL · SMPS) — AC 단상 220V', C_MISC)
ROW += 1
MISC_COLOR = 'F3EBF9'

# UPS: BX950MI-GR 1.5A, BX1600MI-GR 4.5A×3
# 조명CTRL: POD-22024-4 0.5A×10 = 5.0A (10A MCB)
# SMPS 24V: PRO ECO 480W 2.4A×2, 1.2A×1
misc_data = [
    ('공통', 'UPS',          'BX950MI-GR',            0, 1.5, 1, 0,0,0, 1.5, False, True, False, 'AC단상 220V — 1대'),
    ('공통', 'UPS',          'BX1600MI-GR',           0, 4.5, 1, 0,0,0, 4.5, False, True, False, 'AC단상 220V — 3대 각각 별도 차단기'),
    ('공통', '조명CTRL',     'POD-22024-4-PEI-LKCB',  0, 5.0, 1, 0,1,0, 5.0, False, True, False, '10개×500mA=5000mA → 10A MCB'),
    ('공통', 'SMPS 24V 20A', 'PRO ECO 480W 24V 20A',  0, 2.4, 1, 0,0,1, 2.4, False, True, False, '2대×2400mA + 1대×1200mA'),
    ('공통', 'SMPS 24V 10A', 'PRO ECO 480W 24V 10A',  0, 1.2, 1, 1,0,0, 1.2, False, True, False, '1대×1200mA'),
    ('공통', 'HMI/PC',       '(직접 입력)',            0, 0.0, 1, 0,0,0, 0.0, False, True, False, '← 실측값 입력'),
    ('공통', '기타',          '(직접 입력)',            0, 0.0, 1, 0,0,0, 0.0, False, False, False, '← 실측값 입력'),
]
misc_start = ROW
for d in misc_data:
    panel, axis, model, watt, ri, ph, r, s, t, oi, ac3, ac1, dc, note = d
    data_row(ws, ROW, panel, axis, model, watt, ri, ph, r, s, t, oi,
             ac3=ac3, ac1=ac1, dc=dc, note=note, fill_color=MISC_COLOR)
    ROW += 1

misc_end = ROW - 1
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  기타 부하 소계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9)
ws[f'A{ROW}'].fill = mfill(C_SUM)
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
ws[f'E{ROW}'] = f'=SUM(E{misc_start}:E{misc_end})'
ws[f'J{ROW}'] = f'=SUM(J{misc_start}:J{misc_end})'
for col in range(4, 17):
    ws.cell(row=ROW, column=col).fill = mfill(C_SUM)
    ws.cell(row=ROW, column=col).border = thin_b
    ws.cell(row=ROW, column=col).font = mfont(bold=True, size=9)
    ws.cell(row=ROW, column=col).alignment = center
ws[f'Q{ROW}'] = '기타: 단상 부하 상 편중 주의'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
misc_sum_row = ROW
ROW += 2

# ============================================================
# 전체 합계
# ============================================================
section_header(ws, ROW, '■ 전체 합계 및 차단기 선정', C_TITLE)
ROW += 1

TOTAL_ROW = ROW
ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  전체 정격전류 합계 (AC 3상 서보 기준)'
ws[f'A{ROW}'].font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill('2F5496')
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b

ws[f'E{ROW}'] = f'=E{pan_sum_row}+E{fas_sum_row}+E{dc_sum_row}+E{misc_sum_row}'
ws[f'E{ROW}'].font = mfont(bold=True, size=11)
ws[f'E{ROW}'].number_format = '0.00'
ws[f'E{ROW}'].alignment = center
ws[f'E{ROW}'].border = thin_b
ws[f'E{ROW}'].fill = mfill('FFFF99')

ws[f'F{ROW}'] = f'=E{ROW}*1.25'
ws[f'F{ROW}'].font = mfont(bold=True, size=10)
ws[f'F{ROW}'].number_format = '0.00'
ws[f'F{ROW}'].alignment = center
ws[f'F{ROW}'].border = thin_b
ws[f'F{ROW}'].fill = mfill('FFE699')

ws[f'G{ROW}'] = '×1.25 (안전계수)'
ws[f'G{ROW}'].font = mfont(size=9, color='555555')
ws[f'G{ROW}'].border = thin_b

ws[f'O{ROW}'] = '→ 차단기 직접 선정'
ws[f'O{ROW}'].font = mfont(bold=True, size=9)
ws[f'O{ROW}'].border = thin_b
ws[f'O{ROW}'].alignment = center

ws[f'Q{ROW}'] = '안전계수 1.25 적용 후 규격 차단기 선정 (Sheet4 참조)'
ws[f'Q{ROW}'].font = mfont(size=8, color=C_TITLE)
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 20
ROW += 1

ws.merge_cells(f'A{ROW}:C{ROW}')
ws[f'A{ROW}'] = '  R/S/T 상별 전류 합계'
ws[f'A{ROW}'].font = mfont(bold=True, size=9, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill('833C00')
ws[f'A{ROW}'].alignment = left_a
ws[f'A{ROW}'].border = thin_b
for col, label in [(7,'R'),(8,'S'),(9,'T')]:
    c = ws.cell(row=ROW, column=col, value=f'→ {label}상 합산 입력')
    c.font = mfont(size=9, color='555555')
    c.border = thin_b
    c.alignment = center
    c.fill = mfill('F2DCDB')
ws[f'Q{ROW}'] = '각 상별 전류: 드라이버별 R/S/T 할당 × 정격전류 합산 (수동 확인)'
ws[f'Q{ROW}'].font = mfont(size=8, color='555555')
ws[f'Q{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 17
ROW += 2

# 수식 설명
ws.merge_cells(f'A{ROW}:Q{ROW}')
ws[f'A{ROW}'].value = '▶ 주요 수식 설명'
ws[f'A{ROW}'].font = mfont(bold=True, size=10, color='FFFFFF')
ws[f'A{ROW}'].fill = mfill('375623')
ws[f'A{ROW}'].alignment = center
ws[f'A{ROW}'].border = thin_b
ws.row_dimensions[ROW].height = 16
ROW += 1

formulas = [
    ('AC 3상 정격전류', '카탈로그 정격입력전류 사용 (VLOOKUP)', '예) MADLN05BE=0.3A, MCDLN35BL=2.4A', 'Sheet2 파나소닉A6규격 시트 참조'),
    ('설계전류',        'I_설계 = I_정격 × 수용률 0.7',       '→ J열 자동계산 (실측 수용률로 교체 가능)', '파스텍 EZS2/EZI-SPEED: 수용률 1.0 (정격=설계)'),
    ('VLOOKUP 수식',    'E열 = VLOOKUP(C열, 규격시트, 3, FALSE)', '→ C열 모델 코드 정확히 입력 시 자동조회', '파나소닉A6규격 / 파스텍EziServo규격 시트'),
    ('안전계수',        'I_설계_최종 = I_합계 × 1.25',        '→ 이 값 이상의 차단기 선정',               'NEC/IEC 기준 1.25배 (연속운전 고려)'),
    ('차단기 선정',     '정격 > I_설계_최종, 단락전류 ≥ 현장', '→ Sheet4 차단기 규격 참조표 확인',         'MCCB: 산업용, MCB: 소용량'),
]
for row_data in formulas:
    label, formula, example, note = row_data
    ws.merge_cells(f'A{ROW}:B{ROW}')
    ws[f'A{ROW}'].value = label
    ws[f'A{ROW}'].font = mfont(bold=True, size=9)
    ws[f'A{ROW}'].fill = mfill(C_FORMULA)
    ws[f'A{ROW}'].alignment = left_a
    ws[f'A{ROW}'].border = thin_b
    ws.merge_cells(f'C{ROW}:F{ROW}')
    ws[f'C{ROW}'].value = formula
    ws[f'C{ROW}'].font = Font(name='Courier New', bold=True, size=9, color='1F3864')
    ws[f'C{ROW}'].fill = mfill(C_FORMULA)
    ws[f'C{ROW}'].alignment = left_a
    ws[f'C{ROW}'].border = thin_b
    ws.merge_cells(f'G{ROW}:J{ROW}')
    ws[f'G{ROW}'].value = example
    ws[f'G{ROW}'].font = mfont(size=9, color='375623')
    ws[f'G{ROW}'].fill = mfill('EBF3E8')
    ws[f'G{ROW}'].alignment = left_a
    ws[f'G{ROW}'].border = thin_b
    ws.merge_cells(f'K{ROW}:Q{ROW}')
    ws[f'K{ROW}'].value = note
    ws[f'K{ROW}'].font = mfont(size=9, color='555555')
    ws[f'K{ROW}'].fill = mfill(C_FORMULA)
    ws[f'K{ROW}'].alignment = left_a
    ws[f'K{ROW}'].border = thin_b
    ws.row_dimensions[ROW].height = 16
    ROW += 1

# 상단 요약 수식
ws['B4'].value = f'=E{pan_sum_row}'
ws['B5'].value = f'=E{fas_sum_row}'
ws['B6'].value = f'=E{dc_sum_row}'
ws['B7'].value = f'=E{misc_sum_row}'
for r in range(4, 8):
    ws.cell(row=r, column=2).number_format = '0.0'
    ws.cell(row=r, column=3).value = f'=B{r}*1.25'
    ws.cell(row=r, column=3).number_format = '0.0'
    ws.cell(row=r, column=4).value = '→ Sheet4 참조'
    ws.cell(row=r, column=4).font = mfont(size=9)

# ============================================================
# SHEET 2: 파나소닉A6규격 (VLOOKUP 소스)
# VA 기준: MADL=500VA, MBDL=900VA, MCDL=1300VA, MDDL=2300VA
# I = VA / (√3 × 220 × 0.95) = VA / 362.1
# ============================================================
ws2 = wb.create_sheet('파나소닉A6규격')
ws2.merge_cells('A1:D1')
ws2['A1'] = '파나소닉 A6 서보드라이버 정격입력전류 (VLOOKUP 소스) — 카탈로그 기준값'
ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws2['A1'].alignment = center
ws2['A1'].fill = mfill('D6E4F0')
ws2.row_dimensions[1].height = 22

for j, h in enumerate(['드라이버 모델 (A열)','피상전력 VA (B열)','정격입력전류 A (C열)','비고'], 1):
    c = ws2.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws2.row_dimensions[2].height = 20

# 카탈로그 정격입력전류 (3상 AC200V 기준)
pan_lookup = [
    ('MADLN05BE',   50,  0.3),
    ('MADLN10BE',  100,  0.5),
    ('MADLN15BE',  200,  1.0),
    ('MBDLN25BE',  400,  1.5),
    ('MCDLN35BL',  750,  2.4),
    ('MCDLN35BE',  750,  2.4),
    ('MDDLN55BL', 1000,  3.4),
    ('MDDLN55BE', 1000,  3.4),
    ('MDDLN85BL', 1500,  5.5),
]
for i, (model, watt, curr) in enumerate(pan_lookup, 3):
    fill = mfill('EBF3FB') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate([model, watt, curr, f'카탈로그 정격전류 {curr}A'], 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j > 1 else left_a
        c.border = thin_b
        c.fill = fill
    ws2.row_dimensions[i].height = 16

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 28

note_r = len(pan_lookup) + 4
ws2.merge_cells(f'A{note_r}:D{note_r}')
ws2[f'A{note_r}'] = '※ A열 모델명과 메인시트 C열 코드 정확히 일치 필요  |  카탈로그 정격입력전류 기준 (3상 AC200V)  |  모델 추가 시 이 시트에 행 추가'
ws2[f'A{note_r}'].font = mfont(size=8, color='555555')
ws2[f'A{note_r}'].fill = mfill(C_FORMULA)
ws2[f'A{note_r}'].alignment = left_a

# ============================================================
# SHEET 3: 파스텍EziServo규격 (VLOOKUP 소스)
# ============================================================
ws3 = wb.create_sheet('파스텍EziServo규격')
ws3.merge_cells('A1:C1')
ws3['A1'] = '파스텍 EZS2/EziServo 정격전류 (VLOOKUP 소스)'
ws3['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws3['A1'].alignment = center
ws3['A1'].fill = mfill('FCE4D6')
ws3.row_dimensions[1].height = 22

for j, h in enumerate(['드라이버 모델 (A열)','정격출력 W (B열)','정격전류 A (C열)'], 1):
    c = ws3.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws3.row_dimensions[2].height = 20

fas_lookup = [
    ('EZS2-EC-42M',           0,   1.2),
    ('EZS2-EC-42XL',          0,   1.2),
    ('EZI-SPEED60H30CR100P',  0,   1.0),
    ('Ezi-EC-42XL-A',         0,   2.0),
    ('Ezi-EC-42M-A',          0,   1.5),
]
for i, (model, watt, curr) in enumerate(fas_lookup, 3):
    fill = mfill('FEF4E8') if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate([model, watt, curr], 1):
        c = ws3.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j > 1 else left_a
        c.border = thin_b
        c.fill = fill
    ws3.row_dimensions[i].height = 16

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 20

# ============================================================
# SHEET 4: 차단기 규격 참조표
# ============================================================
ws4 = wb.create_sheet('차단기 규격 참조표')
ws4.merge_cells('A1:G1')
ws4['A1'] = '차단기 규격 참조표 (MCCB / MCB — AC 220V 3상/단상)'
ws4['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color=C_TITLE)
ws4['A1'].alignment = center
ws4['A1'].fill = mfill('D9E2F3')
ws4.row_dimensions[1].height = 22

for j, h in enumerate(['정격전류(A)','적용 부하전류 범위','제조사','모델 3P (MCCB)','모델 1P (MCB)','특이사항','비고'], 1):
    c = ws4.cell(row=2, column=j, value=h)
    c.font = mfont(bold=True, size=9, color='FFFFFF')
    c.fill = mfill(C_HEADER)
    c.alignment = center
    c.border = thin_b
ws4.row_dimensions[2].height = 22

breaker_data = [
    (3,'~2.4A','LS/Chint','ABN3c / NXB-63','BKN1-1P C3','조명·소형제어',''),
    (5,'2.4~4A','LS/Chint','ABN5c / NXB-63','BKN1-1P C5','소형I/O',''),
    (10,'4~8A','LS/Chint','ABN10c / NXB-63','BKN1-1P C10','소형서보 1~2축',''),
    (15,'8~12A','LS/Chint','ABN15c / NXB-63','BKN1-1P C15','',''),
    (20,'12~16A','LS/Chint','ABN20c / NXB-63','BKN1-1P C20','중형서보',''),
    (25,'16~20A','LS/Chint','ABN25c / NXB-63','BKN1-1P C25','',''),
    (30,'20~24A','LS/Chint','ABN30c / NXB-63','-','다축서보 판넬',''),
    (40,'24~32A','LS/Chint','ABN40c','-','',''),
    (50,'32~40A','LS/Chint','ABN50c','-','',''),
    (63,'40~50A','LS/Chint','ABN63c','-','대형 판넬',''),
    (75,'50~60A','LS/Chint','ABN75c','-','메인 차단기',''),
    (100,'60~80A','LS/Chint','ABN100c','-','',''),
    (125,'80~100A','LS','ABN125c','-','대형 메인',''),
]
C_BREAKER = 'F4F9C6'
for i, row in enumerate(breaker_data, 3):
    fill = mfill(C_BREAKER) if i % 2 == 1 else mfill('FFFFFF')
    for j, val in enumerate(row, 1):
        c = ws4.cell(row=i, column=j, value=val)
        c.font = mfont(size=9)
        c.alignment = center if j != 6 else left_a
        c.border = thin_b
        c.fill = fill
        if j == 1:
            c.fill = mfill(C_SUM)
            c.font = mfont(bold=True, size=9)
    ws4.row_dimensions[i].height = 16

for col, w in {'A':14,'B':18,'C':12,'D':20,'E':18,'F':20,'G':14}.items():
    ws4.column_dimensions[col].width = w

# ── 저장 ──
out_path = 'C:/MES/wta-agents/workspaces/control-agent/대구텍_F2_1_판넬별_차단기선정표_v3.xlsx'
wb.save(out_path)
print('저장 완료:', out_path)
print(f'파나소닉 {len(panasonic_data)}축 / EZS2 {len(fastech_data)}축 / DC {len(dc_data)}항목 / 기타 {len(misc_data)}항목')
