"""
wmx_license 테이블 생성 및 WMX2/WMX3 라이센스 데이터 삽입 스크립트
실행: py db-write-wmx.py
"""
import os
import sys
import psycopg2
import openpyxl
from datetime import datetime, date

XLSX_PATH = r"C:\MES\wta-agents\data\motion-data\WMX2 라이센스 관리대장.xlsx"

def load_mes_password():
    env_path = "C:/MES/backend/.env"
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.startswith("DB_PASSWORD="):
                    return line.strip().split("=", 1)[1]
    return None

DDL = """
CREATE TABLE IF NOT EXISTS hardware.wmx_license (
    id              SERIAL PRIMARY KEY,
    version         VARCHAR(10) NOT NULL CHECK (version IN ('WMX2', 'WMX3')),
    company_name    VARCHAR(200),
    project_number  VARCHAR(100),
    equipment_name  VARCHAR(200),
    unit_number     VARCHAR(50),
    product_part_number  VARCHAR(100),
    serial_number        VARCHAR(100),
    pik_code        VARCHAR(200),
    dongle          VARCHAR(100),
    rtx_pac         VARCHAR(100),
    license_code    TEXT,
    computer_id     VARCHAR(200),
    dongle_id       VARCHAR(100),
    update_date     DATE,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    updated_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_wmx_license_version ON hardware.wmx_license (version);
CREATE INDEX IF NOT EXISTS idx_wmx_license_serial  ON hardware.wmx_license (serial_number);
CREATE INDEX IF NOT EXISTS idx_wmx_license_company ON hardware.wmx_license (company_name);
"""

TRIGGER_FUNC = """
CREATE OR REPLACE FUNCTION hardware.update_wmx_license_updated_at()
RETURNS TRIGGER AS $$
BEGIN
    NEW.updated_at = NOW();
    RETURN NEW;
END;
$$ LANGUAGE plpgsql;
"""

TRIGGER = """
DO $$
BEGIN
    IF NOT EXISTS (
        SELECT 1 FROM pg_trigger WHERE tgname = 'trg_wmx_license_updated_at'
    ) THEN
        CREATE TRIGGER trg_wmx_license_updated_at
            BEFORE UPDATE ON hardware.wmx_license
            FOR EACH ROW EXECUTE FUNCTION hardware.update_wmx_license_updated_at();
    END IF;
END$$;
"""

INSERT_SQL = """
INSERT INTO hardware.wmx_license (
    version, company_name, project_number, equipment_name, unit_number,
    product_part_number, serial_number, pik_code, dongle, rtx_pac,
    license_code, computer_id, dongle_id, update_date
) VALUES (
    %(version)s, %(company_name)s, %(project_number)s, %(equipment_name)s, %(unit_number)s,
    %(product_part_number)s, %(serial_number)s, %(pik_code)s, %(dongle)s, %(rtx_pac)s,
    %(license_code)s, %(computer_id)s, %(dongle_id)s, %(update_date)s
)
"""

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != 'None' else None

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def load_wmx2_rows(ws):
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # 3행부터 데이터
    records = []
    for row in rows:
        # col[0]=인덱스, [1]=업체명, [2]=프로젝트번호, [3]=장비명, [4]=호기
        # [5]=Product Part Number, [6]=Serial Number, [7]=PIK Code, [8]=DONGLE
        # [9]=RTX PAC#, [10]=WMX2 License Code, [11]=Computer ID, [12]=Update Date
        # [14]=WMX License Code (추가), [15]=Dongle ID
        if not any(row[1:13]):
            continue
        records.append({
            "version": "WMX2",
            "company_name": clean(row[1]),
            "project_number": clean(row[2]),
            "equipment_name": clean(row[3]),
            "unit_number": clean(row[4]),
            "product_part_number": clean(row[5]),
            "serial_number": clean(row[6]),
            "pik_code": clean(row[7]),
            "dongle": clean(row[8]),
            "rtx_pac": clean(row[9]),
            "license_code": clean(row[10]) or clean(row[14] if len(row) > 14 else None),
            "computer_id": clean(row[11]),
            "dongle_id": clean(row[15]) if len(row) > 15 else None,
            "update_date": parse_date(row[12]),
        })
    return records

def load_wmx3_rows(ws):
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    records = []
    for row in rows:
        # col[0]=인덱스(수식), [1]=업체명, [2]=프로젝트번호, [3]=장비명, [4]=호기
        # [5]=Product Part Number, [6]=Serial Number, [7]=PIK Code, [8]=DONGLE
        # [9]=RTX PAC#, [10]=WMX3 License Code, [11]=Computer ID, [12]=Dongle ID
        if not any(row[1:12]):
            continue
        records.append({
            "version": "WMX3",
            "company_name": clean(row[1]),
            "project_number": clean(row[2]),
            "equipment_name": clean(row[3]),
            "unit_number": clean(row[4]),
            "product_part_number": clean(row[5]),
            "serial_number": clean(row[6]),
            "pik_code": clean(row[7]),
            "dongle": clean(row[8]),
            "rtx_pac": clean(row[9]),
            "license_code": clean(row[10]),
            "computer_id": clean(row[11]),
            "dongle_id": clean(row[12]),
            "update_date": None,
        })
    return records

def main():
    password = load_mes_password()
    if not password:
        print("[오류] MES DB 비밀번호를 찾을 수 없습니다.")
        sys.exit(1)

    conn = psycopg2.connect(
        host="localhost", port=55432, user="postgres",
        password=password, dbname="postgres"
    )
    conn.autocommit = False

    try:
        cur = conn.cursor()

        # 1단계: 테이블 생성
        print("[1단계] 테이블 생성 중...")
        cur.execute(DDL)
        cur.execute(TRIGGER_FUNC)
        cur.execute(TRIGGER)
        conn.commit()
        print("  → hardware.wmx_license 테이블 생성 완료")

        # 2단계: xlsx 읽기
        print("[2단계] xlsx 파일 로드 중...")
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        wmx2_records = load_wmx2_rows(wb["WMX2"])
        wmx3_records = load_wmx3_rows(wb["WMX3"])
        print(f"  → WMX2: {len(wmx2_records)}건, WMX3: {len(wmx3_records)}건 파싱 완료")

        # 3단계: 데이터 삽입
        print("[3단계] 데이터 삽입 중...")
        cur.execute("TRUNCATE hardware.wmx_license RESTART IDENTITY")
        for rec in wmx2_records + wmx3_records:
            cur.execute(INSERT_SQL, rec)
        conn.commit()
        print(f"  → 총 {len(wmx2_records) + len(wmx3_records)}건 삽입 완료")

        # 검증
        cur.execute("SELECT version, COUNT(*) FROM hardware.wmx_license GROUP BY version ORDER BY version")
        rows = cur.fetchall()
        for v, c in rows:
            print(f"  → {v}: {c}건")

    except Exception as e:
        conn.rollback()
        print(f"[오류] {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()

if __name__ == "__main__":
    main()
