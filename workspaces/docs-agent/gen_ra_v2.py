"""몰디노 PVD 로딩 장비 RA 평가표 v2 — WTA 실제 장비 구조 기반"""
import sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment

SRC = r"C:\MES\wta-agents\workspaces\MAX\safety_check.xlsx"
DST = r"C:\MES\wta-agents\reports\MAX\moldino-pvd-ra.xlsx"

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb.worksheets[1]

ws['F4'] = '(주)윈텍오토메이션'
ws['F6'] = '몰디노(Moldino) PVD 코팅 로딩 장비 (HAM-PVD-L)'

# WTA PVD Loading 장비 실제 구조 기반 리스크 10건
# 출처: 1._User_Manual_PVD_Loading.md, 1.PVD_Loading_Automation_1_Manual.md
risks = [
    (
        "팔레트 공급/취출\n(작업 도어 통한\n팔레트 교체)",
        "작업 도어를 열고 코팅 팔레트 공급·취출 시,\n서보축 동작 영역에 팔레트를 밀어 넣다가\n서보축·회전 적재 장치에 손·팔이 협착될 위험.\n(매뉴얼 §2.3 협착/말림 위험 해당)",
        "협착\n(挟まれ)",
        3, 3, 4,
        "①작업 도어 인터록 정상 작동 확인 (매 교대)\n②팔레트 교체 시 일시정지(PAUSE) 선행 확인\n③작업 도어 열림 시 서보축 자동 정지 기능 검증\n④팔레트 가이드레일 설치로 수작업 최소화",
        3, 1, 4
    ),
    (
        "회전 적재 장치\n구동 중 접근",
        "인서트 적재용 회전 장치 구동 중 회전부에\n손·의복·장갑이 말려 들어감.\n벨트 구동부에서 압착·절단 사고 가능.\n(매뉴얼 §2.5 구동기 위험 해당)",
        "말림/절단\n(巻込まれ)",
        4, 2, 3,
        "①회전부 안전 커버(가드) 설치 및 인터록\n②구동 중 안전 도어 개방 시 즉시 정지 확인\n③헐거운 의복·장갑 착용 금지 교육\n④정기 점검 시 커버 체결 상태 확인",
        4, 1, 1
    ),
    (
        "서보축 원점복귀\n/티칭 작업",
        "원점복귀 시 다축 서보가 동시 이동하여\n예기치 않은 축 동작으로 작업자 충돌·끼임.\n원점 데이터 손실 시 비정상 궤적 이동 위험.\n(매뉴얼 §2.5 서보축 위험 해당)",
        "충돌/끼임\n(衝突/挟まれ)",
        4, 2, 3,
        "①원점복귀 시 안전 도어 닫힘 필수 (인터록)\n②티칭 작업은 WTA 엔지니어 또는 교육자만\n③저속 모드(Jog) 필수, 2인 1조 작업\n④원점 데이터 백업 주기적 실시",
        4, 1, 1
    ),
    (
        "비전 카메라 시스템\n(NoseR/CB 검사,\n트래킹 패턴)",
        "비전 시스템 광 센서 및 조명의 레이저/LED\n광선이 작업자 눈에 직접 또는 반사로 입사.\n장시간 노출 시 시력 손상 가능.\n(매뉴얼 §2.5 시각 광선 해당)",
        "시력손상\n(視覚障害)",
        2, 2, 3,
        "①비전 세팅(§11.9) 작업 시 보호 안경 착용\n②광선 방향 조정 시 직접 주시 금지 교육\n③반사 방지 커버 설치\n④광원 장치 임의 조작 금지 표시 부착",
        2, 1, 3
    ),
    (
        "전원부 점검/보수\n(컨트롤 판넬)",
        "400V/30A/3Phase 고압 전원부 접촉 시 감전.\n전원 차단 후에도 잔류 전원 존재(UPS 1분,\n잔류전압 3분 이상 방치 필요).\n(매뉴얼 §2.5 전기 위험, §3.3 참고)",
        "감전\n(感電)",
        5, 2, 2,
        "①전원 차단 후 3분 이상 대기 (잔류전압 방전)\n②UPS 완전 차단 확인 후 작업 개시\n③LOTO(잠금/표찰) 절차 적용\n④전기 전문 인력만 작업 가능\n⑤컨트롤 판넬 커버 항시 닫힘 유지",
        5, 1, 1
    ),
    (
        "공압 장치 점검\n(Ø12/0.5Mpa)",
        "공압 호스·라인 결합부 이탈 시 0.5Mpa\n압축공기 분출로 피부·눈 손상.\n씰링 불량 시 예고 없이 분출 가능.\n(매뉴얼 §2.5 공압 해당)",
        "비래/파열\n(飛来/破裂)",
        2, 2, 2,
        "①공압 전문 지식자만 점검 작업 수행\n②호스·라인 결합부 정기 점검(씰링·외관)\n③점검 전 공압 차단 후 잔압 배출\n④보호 안경 착용 의무",
        2, 1, 2
    ),
    (
        "호퍼 공급부/\n스페이서 공급부\n소재 보충 작업",
        "호퍼·스페이서 공급부 소재 보충 시\n공급 기구 동작에 의한 손가락 끼임.\n자동 운전 중 보충 시도 시 위험 증가.\n(설비 구성: §8.4 스페이서, §8.5 호퍼)",
        "끼임\n(挟まれ)",
        3, 2, 4,
        "①소재 보충 시 일시정지(PAUSE) 선행 필수\n②공급부 투입구 안전 가이드 설치\n③자동 운전 중 투입구 접근 금지 표시\n④소모품 점검(§10.6) 주기 준수",
        3, 1, 4
    ),
    (
        "안전 도어 내\n내부 보수 작업",
        "안전 도어 인터록 해제 후 내부 보수 시\n수동 축 이동 조작 중 서보축·기구 돌출부에\n신체 충돌·끼임. 도어 미닫힘 상태로 운전\n재개 시 내부 작업자 위험.\n(매뉴얼 §2.3 충돌 위험, §2.4 도어 인터락)",
        "충돌/끼임\n(衝突/挟まれ)",
        4, 2, 2,
        "①보수 시 비상정지(E-STOP) 선행 후 작업\n②2인 1조 작업 (감시자 배치)\n③도어 닫힘 확인 → START 2회 눌러 재개\n(의도치 않은 재기동 방지)\n④안전 도어 인터록 월 1회 기능 점검",
        4, 1, 1
    ),
    (
        "초경 인서트\n취급/정렬\n(막대 적재 작업)",
        "초경합금 인서트 날부(에지) 접촉 시\n손·손가락 절상. 막대(Rod)에 인서트 적재 시\n반복 취급으로 누적 절상 위험.\n(설비 기능: §11.1 팔레트 관리, §11.2 막대 관리)",
        "절상\n(切傷)",
        2, 3, 4,
        "①절단방지 장갑(EN388 레벨5) 착용 의무\n②자동 적재 모드 활용으로 수작업 최소화\n③막대·팔레트 교체 시 전용 지그 사용\n④작업 전 안전교육(인서트 취급 요령)",
        2, 1, 4
    ),
    (
        "설비 내부 분진\n(코팅 잔류물/\n초경 미세분진)",
        "초경합금 가공·적재 과정 발생 미세분진 흡입.\n설비 내부 청소 시 코팅 잔류물(금속 분진) 비산.\n통풍 불량 시 농도 상승.\n(매뉴얼 §2.5 분진,증기,연기 해당)",
        "직업성질환\n(粉塵曝露)",
        3, 2, 2,
        "①설비 반드시 통풍 양호 장소에서 가동\n②방진/방습 도어 항시 밀폐 상태 유지\n③내부 청소 시 분진 마스크(KF94↑) 착용\n④국소 배기장치(집진) 가동 확인",
        3, 1, 2
    ),
]

col_map = {
    'work': 3, 'hazard': 11, 'accident': 23,
    'S': 25, 'P': 27, 'F': 29,
    'fix': 35,
    'S2': 51, 'P2': 53, 'F2': 55,
}

font_jp = Font(name='맑은 고딕', size=9)
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center')

for i, r in enumerate(risks):
    row = 11 + i
    work, hazard, accident, s, p, f, fix, s2, p2, f2 = r
    for col, val in [(col_map['work'], work), (col_map['hazard'], hazard),
                     (col_map['accident'], accident), (col_map['fix'], fix)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font_jp
        c.alignment = wrap
    for col, val in [(col_map['S'], s), (col_map['P'], p), (col_map['F'], f),
                     (col_map['S2'], s2), (col_map['P2'], p2), (col_map['F2'], f2)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font_jp
        c.alignment = center

wb.save(DST)
print(f"OK: {DST}")

print("\n=== WTA PVD 로딩 장비 리스크 어세스먼트 (HAM-PVD-L) ===")
print(f"{'No':>3} {'작업명':<22} {'유형':<10} {'S':>2} {'P':>2} {'F':>2} {'합':>3} {'Lv':>4}  →  {'S2':>2} {'P2':>2} {'F2':>2} {'합':>3} {'Lv':>4}")
print("-" * 90)
for i, r in enumerate(risks, 1):
    _, _, _, s, p, f, _, s2, p2, f2 = r
    total = s * p * f
    total2 = s2 * p2 * f2
    lv = 'I' if total <= 4 else ('II' if total <= 18 else ('III' if total <= 27 else 'IV'))
    lv2 = 'I' if total2 <= 4 else ('II' if total2 <= 18 else ('III' if total2 <= 27 else 'IV'))
    name = r[0].replace('\n', ' ')[:21]
    acc = r[2].split('\n')[0][:9]
    print(f"{i:>3} {name:<22} {acc:<10} {s:>2} {p:>2} {f:>2} {total:>3}  {lv:>4}  →  {s2:>2} {p2:>2} {f2:>2} {total2:>3}  {lv2:>4}")

print(f"\n시정 전: III {sum(1 for r in risks if r[3]*r[4]*r[5]>18)}건, II {sum(1 for r in risks if 5<=r[3]*r[4]*r[5]<=18)}건, I {sum(1 for r in risks if r[3]*r[4]*r[5]<=4)}건")
print(f"시정 후: III {sum(1 for r in risks if r[7]*r[8]*r[9]>18)}건, II {sum(1 for r in risks if 5<=r[7]*r[8]*r[9]<=18)}건, I {sum(1 for r in risks if r[7]*r[8]*r[9]<=4)}건")
