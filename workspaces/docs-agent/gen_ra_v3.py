"""몰디노 PVD 로딩 장비 RA v3 — sales-agent 사양 반영"""
import sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment

SRC = r"C:\MES\wta-agents\workspaces\MAX\safety_check.xlsx"
DST = r"C:\MES\wta-agents\reports\MAX\moldino-pvd-ra.xlsx"

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb.worksheets[1]

# 헤더 정보 — sales-agent 사양 반영
ws['F4'] = '(주)윈텍오토메이션'
ws['F6'] = 'PVD 로딩 장비 (PVD-L #1) / JPJMOHS20001\n3,000×2,300×2,100mm / PC 기반 제어 / 몰디노(Moldino) 납품'

# WTA PVD-L #1 실제 사양 + 매뉴얼 기반 리스크 10건
risks = [
    (
        "팔레트 공급부\n(Stacker 20단\n공급/교체)",
        "Stacker type 20단 적재 팔레트 공급부에서\n팔레트 교체·보충 시 Stacker 승강 기구에\n손·팔 협착 위험. 작업 도어 인터록 미확인\n상태로 접근 시 자동 승강 동작 가능.\n(매뉴얼 §2.3 협착/말림, §2.4 도어 인터락)",
        "협착\n(挟まれ)",
        3, 3, 4,
        "①작업 도어 인터록 정상 동작 확인 (매 교대)\n②팔레트 교체 시 PAUSE 선행 필수\n③Stacker 승강부 안전 가이드 설치\n④교체 절차 표준화 (SOP 게시)",
        3, 1, 4
    ),
    (
        "XYZ 직교로봇\n동작 영역 접근\n(3jaw TOOL)",
        "메인 XYZ 직교로봇(3jaw TOOL) 동작범위 내\n작업자 진입 시 로봇 암·3jaw에 의한 충돌·끼임.\n제품 픽업/이송 중 고속 동작으로 심각한 충격.\nPass Line 1,070mm 높이로 상체 접근 용이.\n(매뉴얼 §2.3 충돌 위험, §2.5 서보축)",
        "충돌/끼임\n(衝突/挟まれ)",
        4, 2, 3,
        "①안전 펜스 + 안전 도어 인터록 완비\n②라이트커튼 설치 (카테고리4)\n③티칭/점검 시 저속 Jog 모드 + 2인 1조\n④로봇 동작범위 바닥 표시(황색 라인)",
        4, 1, 1
    ),
    (
        "회전 적재 장치\n/센터링 기구\n구동 중 접근",
        "인서트 적재용 회전 장치 및 지그 센터링 기구\n구동 중 회전부에 손·의복 말림.\n3종 지그 교체 시 회전부 잔류 동작 위험.\n(매뉴얼 §2.5 구동기 — 벨트·구동부 압착/절단)",
        "말림/절단\n(巻込まれ)",
        4, 2, 3,
        "①회전부 안전 커버(가드) + 인터록 설치\n②지그 교체 시 E-STOP 선행 후 작업\n③헐거운 의복·장갑 착용 금지 교육\n④커버 체결 상태 정기 점검 (주 1회)",
        4, 1, 1
    ),
    (
        "비전 시스템\n(자동 align\n카메라/조명)",
        "비전 시스템 자동 align 카메라 광원(LED/레이저)\n광선이 작업자 눈에 직접·반사 입사.\nNoseR/CB 검사, 트래킹 패턴 세팅 중 노출.\n(매뉴얼 §2.5 시각 광선 — 직접·반사 주시 금지)",
        "시력손상\n(視覚障害)",
        2, 2, 3,
        "①비전 세팅 작업 시 보호 안경 착용 필수\n②광원 직접 주시 금지 교육 (입사 시)\n③반사 방지 커버 설치\n④광원 장치 임의 조작 금지 표지 부착",
        2, 1, 3
    ),
    (
        "전원부 점검/보수\n(PC 기반 제어부\nUPS 포함)",
        "PC 기반 제어 시스템 전원부(400V/3Phase) 접촉 감전.\nUPS 동작으로 메인 차단 후에도 PC 전원 유지(1분).\n잔류 전압 3분 이상 방치 필요.\n(매뉴얼 §2.5 전기 위험, §3.3 전원 끄기 절차)",
        "감전\n(感電)",
        5, 2, 2,
        "①전원 차단 후 3분 이상 대기(잔류전압 방전)\n②UPS 완전 차단 확인 후 작업 개시\n③LOTO(잠금/표찰) 절차 적용\n④전기 전문 인력(유자격자)만 작업\n⑤컨트롤 판넬 커버 항시 닫힘 유지",
        5, 1, 1
    ),
    (
        "Rod 공급장치\n(전장 265mm,\nØ1.8~Ø5)",
        "Rod 공급장치에서 막대(전장 265mm) 공급·교체 시\n공급 기구 동작에 의한 손가락 끼임.\nRod 선단이 날카로워 찔림·절상 위험 병존.\n(설비 구성: Rod 공급장치, §11.2 막대 관리)",
        "끼임/절상\n(挟まれ/切傷)",
        3, 2, 4,
        "①Rod 보충 시 PAUSE 선행 필수\n②공급부 투입구 안전 가이드 설치\n③보호 장갑 착용 (끼임+절상 복합 대응)\n④자동 운전 중 투입구 접근 금지 표시",
        3, 1, 4
    ),
    (
        "초경 인서트 취급\n(1~50g, Ø4~Ø25,\n높이 2~6.35mm)",
        "초경합금 인서트(1~50g, Ø4~25mm) 날부 접촉 시\n손·손가락 절상. 다양한 사이즈 혼용 취급으로\n그립 미스 시 낙하+절상 복합 위험.\n(매뉴얼 §11.1 팔레트 관리, 모델별 패턴 설정)",
        "절상\n(切傷)",
        2, 3, 4,
        "①절단방지 장갑(EN388 레벨5) 착용 의무\n②제품 사이즈별 전용 지그(3종) 정확 선택\n③자동 적재 모드 최대 활용(수작업 최소화)\n④작업 전 안전교육(인서트 취급 요령)",
        2, 1, 4
    ),
    (
        "컨베이어 언로딩부\n/Spacer 회수부\n동작",
        "완제품 컨베이어 이송 및 Spacer 회수 동작 중\n벨트·롤러 회전부에 손·의복 끼임.\nSpacer 회수 기구 접근 시 끼임 위험.\n(매뉴얼 §2.5 구동기, 언로딩부 컨베이어)",
        "끼임\n(巻込まれ)",
        3, 2, 3,
        "①컨베이어 회전부 커버 가드 설치\n②비상정지 스위치 언로딩부 인접 배치\n③Spacer 회수 시 자동 정지 연동\n④정기 점검 시 커버·인터록 확인",
        3, 1, 3
    ),
    (
        "안전 도어 내\n내부 보수 작업\n(3종 지그 교체 포함)",
        "안전 도어 인터록 해제 후 내부 보수 시\n서보축·XYZ로봇·기구 돌출부에 충돌·끼임.\n3종 지그 교체 작업 시 17인치 HMI 조작 병행으로\n주의력 분산. 도어 미닫힘 상태 재기동 위험.\n(매뉴얼 §2.3 충돌, §2.4 도어 인터락)",
        "충돌/끼임\n(衝突/挟まれ)",
        4, 2, 2,
        "①보수 시 E-STOP 선행 후 작업\n②2인 1조 작업 (감시자 배치 필수)\n③도어 닫힘 확인 → START 2회 눌러 재개\n④안전 도어 인터록 월 1회 기능 점검\n⑤지그 교체 SOP 표준화 (3종별)",
        4, 1, 1
    ),
    (
        "설비 내부 분진\n(초경 미세분진/\n코팅 잔류물)",
        "초경합금 인서트 취급·적재 시 발생 미세분진.\n설비 내부 청소 시 코팅 잔류물(금속 분진) 비산.\n3,000×2,300mm 대형 설비로 내부 청소 범위 광범위.\n(매뉴얼 §2.5 분진/증기/연기 — 통풍 양호 필수)",
        "직업성질환\n(粉塵曝露)",
        3, 2, 2,
        "①설비 통풍 양호 장소 설치 (환기 확보)\n②방진/방습 도어 항시 밀폐 상태 유지\n③내부 청소 시 분진 마스크(KF94↑) 착용\n④국소 배기장치(집진기) 가동 확인",
        3, 1, 2
    ),
]

col_map = {
    'work': 3, 'hazard': 11, 'accident': 23,
    'S': 25, 'P': 27, 'F': 29,
    'fix': 35,
    'S2': 51, 'P2': 53, 'F2': 55,
}

font = Font(name='맑은 고딕', size=9)
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center')

for i, r in enumerate(risks):
    row = 11 + i
    work, hazard, accident, s, p, f, fix, s2, p2, f2 = r
    for col, val in [(col_map['work'], work), (col_map['hazard'], hazard),
                     (col_map['accident'], accident), (col_map['fix'], fix)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font; c.alignment = wrap
    for col, val in [(col_map['S'], s), (col_map['P'], p), (col_map['F'], f),
                     (col_map['S2'], s2), (col_map['P2'], p2), (col_map['F2'], f2)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = font; c.alignment = center

wb.save(DST)
print(f"OK: {DST}")

print("\n=== PVD-L #1 (JPJMOHS20001) 몰디노 리스크 어세스먼트 ===")
print(f"{'No':>3} {'작업명':<24} {'유형':<8} {'S':>2} {'P':>2} {'F':>2} {'합':>3} {'Lv':>4}  →  {'S2':>2} {'P2':>2} {'F2':>2} {'합':>3} {'Lv':>4}")
print("-" * 92)
for i, r in enumerate(risks, 1):
    _, _, _, s, p, f, _, s2, p2, f2 = r
    total, total2 = s*p*f, s2*p2*f2
    lv = 'I' if total<=4 else ('II' if total<=18 else ('III' if total<=27 else 'IV'))
    lv2 = 'I' if total2<=4 else ('II' if total2<=18 else ('III' if total2<=27 else 'IV'))
    name = r[0].replace('\n',' ')[:23]
    acc = r[2].split('\n')[0][:7]
    print(f"{i:>3} {name:<24} {acc:<8} {s:>2} {p:>2} {f:>2} {total:>3}  {lv:>4}  →  {s2:>2} {p2:>2} {f2:>2} {total2:>3}  {lv2:>4}")

iii_pre = sum(1 for r in risks if r[3]*r[4]*r[5]>18)
iv_pre = sum(1 for r in risks if r[3]*r[4]*r[5]>=28)
print(f"\n시정 전: IV {iv_pre}건, III {iii_pre-iv_pre}건, II {sum(1 for r in risks if 5<=r[3]*r[4]*r[5]<=18)}건")
print(f"시정 후: I {sum(1 for r in risks if r[7]*r[8]*r[9]<=4)}건, II {sum(1 for r in risks if 5<=r[7]*r[8]*r[9]<=18)}건 (전 건 II 이하)")
