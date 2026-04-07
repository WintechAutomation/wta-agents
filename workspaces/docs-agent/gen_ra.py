"""몰디노 PVD 로딩 장비 리스크 어세스먼트 평가표 생성"""
import sys, io, shutil, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

SRC = r"C:\MES\wta-agents\workspaces\MAX\safety_check.xlsx"
DST = r"C:\MES\wta-agents\reports\MAX\moldino-pvd-ra.xlsx"

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb.worksheets[1]  # ⑭メーカーリスクアセスメント表

# 회사명, 장비명 등 헤더 정보
ws['F4'] = '(주)윈텍오토메이션'
ws['F6'] = '몰디노(Moldino) PVD 코팅 로딩 장비'

# 리스크 평가 데이터 10건
# (작업명, 위험요인, 사고유형, S, P, F, 시정내용, 시정후S, 시정후P, 시정후F)
risks = [
    (
        "인서트 정렬\n(하드링 작업)",
        "초경합금 인서트 날부 취급 시 손·손가락 베임 위험.\n날카로운 절삭 에지가 피부를 절단할 수 있음.",
        "절상\n(切傷)",
        2, 3, 4,
        "①절단방지 장갑(EN388 레벨5) 착용 의무화\n②자동 정렬 피더 적용으로 수작업 최소화\n③작업 전 안전교육 실시",
        2, 1, 4
    ),
    (
        "로봇 로딩/\n언로딩 동작",
        "로봇 암 동작범위 내 작업자 진입 시 로봇에 의한 끼임·충돌 위험.\n티칭/점검 중 예기치 않은 로봇 기동 가능.",
        "끼임\n(挟まれ)",
        4, 2, 3,
        "①안전펜스 설치 + 도어 인터록\n②라이트커튼 설치(카테고리4)\n③티칭 시 저속 모드 필수 + 2인 1조 작업",
        4, 1, 1
    ),
    (
        "지그 세팅/\n교체 작업",
        "중량(10~20kg) 코팅 지그 운반·세팅 중 낙하로 발·손 타박·골절.\n지그 고정 불량 시 공정 중 낙하 가능.",
        "낙하/전도\n(落下/転倒)",
        3, 2, 3,
        "①지그 리프트 보조장치 도입\n②고정 클램프 체결 상태 더블체크 절차\n③안전화(강철 선심) 착용 의무",
        3, 1, 3
    ),
    (
        "PVD 챔버\n투입/인출",
        "고온(200~500℃) 챔버 외벽·도어 접촉 시 화상.\n공정 직후 챔버 개방 시 복사열에 의한 화상 위험.",
        "화상\n(火傷)",
        3, 3, 3,
        "①내열 장갑 + 내열 앞치마 착용 필수\n②차단막(열 실드) 설치\n③챔버 표면 온도 표시등(50℃↑ 경고등)\n④냉각 대기시간 준수 절차 수립",
        3, 1, 3
    ),
    (
        "전원부 보수/\n점검 작업",
        "고압(200~400V) 전원부 접촉 시 감전 사망 위험.\n콘덴서 잔류전하에 의한 지연 감전 가능.",
        "감전\n(感電)",
        5, 2, 2,
        "①LOTO(잠금/표찰) 절차 100% 적용\n②절연 공구 사용 필수\n③잔류전하 방전 확인 후 작업 개시\n④유자격자(전기 담당) 작업 한정",
        5, 1, 1
    ),
    (
        "이송 컨베이어\n가동",
        "벨트·롤러 회전부에 손·의복 끼임.\n비상정지 미작동 시 중상 가능.",
        "끼임\n(巻込まれ)",
        3, 2, 4,
        "①회전부 커버 가드 설치\n②비상정지 스위치 동선별 배치\n③정기 점검 시 커버 인터록 확인",
        3, 1, 4
    ),
    (
        "진공펌프\n운전",
        "진공펌프 가동 시 지속적 소음(85dB↑) 노출.\n장기간 노출에 의한 소음성 난청 위험.",
        "직업성질환\n(職業性疾患)",
        2, 3, 4,
        "①방음 커버(인클로저) 설치\n②귀마개/이어머프 착용 의무\n③연간 청력 검사 실시",
        2, 1, 4
    ),
    (
        "챔버 가스\n공급/배기",
        "PVD 공정 가스(N₂/Ar/반응가스) 누출 시 산소결핍·질식.\n밀폐 공간 내 가스 축적 시 의식 상실 가능.",
        "질식/중독\n(窒息/中毒)",
        5, 2, 2,
        "①가스 감지기 설치(O₂, 유해가스 동시 모니터링)\n②강제 환기팬 상시 가동\n③비상 환기 시스템 연동\n④가스 취급 자격자 지정",
        5, 1, 1
    ),
    (
        "코팅 완제품\n언로딩",
        "고온(100~200℃) 코팅 직후 인서트·트레이 접촉 화상.\n냉각 미완료 제품 조기 언로딩 시 발생.",
        "화상\n(火傷)",
        2, 3, 4,
        "①냉각 대기시간 설정(인터록 연동)\n②내열 장갑(200℃ 이상 내열) 착용\n③온도 센서 연동 경고 시스템",
        2, 1, 4
    ),
    (
        "장비 청소/\n정기 보수",
        "PVD 코팅 잔류 분진(금속 미세분진) 흡입.\n장기 노출 시 진폐증·호흡기 장해 위험.",
        "직업성질환\n(粉塵曝露)",
        3, 2, 2,
        "①분진 마스크(KF94↑) 착용 의무\n②국소 배기장치(집진기) 가동\n③MSDS 비치 및 교육 실시",
        3, 1, 2
    ),
]

# 열 매핑 (openpyxl은 1-indexed)
# B=2, C=3, K=11, W=23, Y=25, AA=27, AC=29, AI=35, AY=51, BA=53, BC=55
col_map = {
    'work': 3,      # C: 작업명
    'hazard': 11,    # K: 위험요인
    'accident': 23,  # W: 사고유형
    'S': 25,         # Y: S
    'P': 27,         # AA: P
    'F': 29,         # AC: F
    'fix': 35,       # AI: 시정내용
    'S2': 51,        # AY: 시정후 S
    'P2': 53,        # BA: 시정후 P
    'F2': 55,        # BC: 시정후 F
}

from openpyxl.styles import Font, Alignment

font_jp = Font(name='맑은 고딕', size=9)
wrap = Alignment(wrap_text=True, vertical='center')

for i, r in enumerate(risks):
    row = 11 + i  # 11~20
    work, hazard, accident, s, p, f, fix, s2, p2, f2 = r
    ws.cell(row=row, column=col_map['work'], value=work).font = font_jp
    ws.cell(row=row, column=col_map['work']).alignment = wrap
    ws.cell(row=row, column=col_map['hazard'], value=hazard).font = font_jp
    ws.cell(row=row, column=col_map['hazard']).alignment = wrap
    ws.cell(row=row, column=col_map['accident'], value=accident).font = font_jp
    ws.cell(row=row, column=col_map['accident']).alignment = wrap
    ws.cell(row=row, column=col_map['S'], value=s).font = font_jp
    ws.cell(row=row, column=col_map['S']).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=col_map['P'], value=p).font = font_jp
    ws.cell(row=row, column=col_map['P']).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=col_map['F'], value=f).font = font_jp
    ws.cell(row=row, column=col_map['F']).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=col_map['fix'], value=fix).font = font_jp
    ws.cell(row=row, column=col_map['fix']).alignment = wrap
    ws.cell(row=row, column=col_map['S2'], value=s2).font = font_jp
    ws.cell(row=row, column=col_map['S2']).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=col_map['P2'], value=p2).font = font_jp
    ws.cell(row=row, column=col_map['P2']).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=col_map['F2'], value=f2).font = font_jp
    ws.cell(row=row, column=col_map['F2']).alignment = Alignment(horizontal='center', vertical='center')

wb.save(DST)
print(f"OK: {DST}")

# 검증 출력
print("\n=== 리스크 평가 요약 ===")
print(f"{'No':>3} {'작업명':<16} {'S':>2} {'P':>2} {'F':>2} {'합':>3} {'Lv':>4}  │  {'시정후S':>2} {'P':>2} {'F':>2} {'합':>3} {'Lv':>4}")
print("-" * 80)
for i, r in enumerate(risks, 1):
    _, _, _, s, p, f, _, s2, p2, f2 = r
    total = s * p * f
    total2 = s2 * p2 * f2
    lv = 'I' if total <= 4 else ('II' if total <= 18 else ('III' if total <= 27 else 'IV'))
    lv2 = 'I' if total2 <= 4 else ('II' if total2 <= 18 else ('III' if total2 <= 27 else 'IV'))
    name = r[0].replace('\n', ' ')[:15]
    print(f"{i:>3} {name:<16} {s:>2} {p:>2} {f:>2} {total:>3}  {lv:>4}  │  {s2:>6} {p2:>2} {f2:>2} {total2:>3}  {lv2:>4}")
