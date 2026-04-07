"""견적서 양식 Excel 생성 스크립트 — cs-wta.com 견적서 구조 기반"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = r"C:\MES\wta-agents\reports\견적서양식.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "견적서"

# ── 스타일 정의 ──
thin = Side(style="thin")
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
border_bottom = Border(bottom=Side(style="medium"))

title_font = Font(name="맑은 고딕", size=20, bold=True)
header_font = Font(name="맑은 고딕", size=11, bold=True)
normal_font = Font(name="맑은 고딕", size=10)
small_font = Font(name="맑은 고딕", size=9)
white_bold = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")

# ── 열 너비 설정 ──
col_widths = {
    "A": 5,   # No
    "B": 22,  # 품명
    "C": 20,  # 규격/모델
    "D": 8,   # 단위
    "E": 10,  # 수량
    "F": 14,  # 단가
    "G": 14,  # 금액
    "H": 14,  # 비고
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ── 행 1~2: 회사명 + 제목 ──
ws.merge_cells("A1:H1")
ws["A1"] = "(주)윈텍오토메이션"
ws["A1"].font = Font(name="맑은 고딕", size=12, bold=True, color="4472C4")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:H2")
ws["A2"] = "견 적 서"
ws["A2"].font = title_font
ws["A2"].alignment = center
ws.row_dimensions[2].height = 40

# ── 행 3: 빈줄 ──
ws.row_dimensions[3].height = 10

# ── 행 4~8: 견적 정보 (좌측) + 공급자 정보 (우측) ──
info_labels_left = [
    ("견적번호", "CS-XXXX-XXX"),
    ("견적일자", "2026-01-01"),
    ("유효기간", "견적일로부터 30일"),
    ("담당자", ""),
    ("연락처", ""),
]

info_labels_right = [
    ("상호", "(주)윈텍오토메이션"),
    ("대표", ""),
    ("사업자등록번호", ""),
    ("주소", "경남 창원시"),
    ("TEL / FAX", ""),
]

row = 4
for i, ((ll, lv), (rl, rv)) in enumerate(zip(info_labels_left, info_labels_right)):
    r = row + i
    ws.row_dimensions[r].height = 22
    # 좌측
    ws[f"A{r}"] = ll
    ws[f"A{r}"].font = header_font
    ws[f"A{r}"].fill = light_blue_fill
    ws[f"A{r}"].border = border_all
    ws[f"A{r}"].alignment = center
    ws.merge_cells(f"B{r}:C{r}")
    ws[f"B{r}"] = lv
    ws[f"B{r}"].font = normal_font
    ws[f"B{r}"].border = border_all
    ws[f"B{r}"].alignment = left_align
    # 우측
    ws[f"E{r}"] = rl
    ws[f"E{r}"].font = header_font
    ws[f"E{r}"].fill = light_blue_fill
    ws[f"E{r}"].border = border_all
    ws[f"E{r}"].alignment = center
    ws.merge_cells(f"F{r}:H{r}")
    ws[f"F{r}"] = rv
    ws[f"F{r}"].font = normal_font
    ws[f"F{r}"].border = border_all
    ws[f"F{r}"].alignment = left_align

# D열 중간 비움
for i in range(5):
    ws[f"D{row + i}"].border = Border()

# ── 행 9: 수신 ──
row = 9
ws.row_dimensions[row].height = 10

# ── 행 10: 수신처 ──
row = 10
ws.row_dimensions[row].height = 25
ws.merge_cells(f"A{row}:B{row}")
ws[f"A{row}"] = "수신: "
ws[f"A{row}"].font = header_font
ws[f"A{row}"].alignment = left_align
ws.merge_cells(f"C{row}:H{row}")
ws[f"C{row}"] = "(고객사명)  귀중"
ws[f"C{row}"].font = normal_font
ws[f"C{row}"].alignment = left_align
ws[f"C{row}"].border = border_bottom

# ── 행 11: 건명 ──
row = 11
ws.row_dimensions[row].height = 25
ws.merge_cells(f"A{row}:B{row}")
ws[f"A{row}"] = "건명: "
ws[f"A{row}"].font = header_font
ws[f"A{row}"].alignment = left_align
ws.merge_cells(f"C{row}:H{row}")
ws[f"C{row}"] = ""
ws[f"C{row}"].font = normal_font
ws[f"C{row}"].alignment = left_align
ws[f"C{row}"].border = border_bottom

# ── 행 12: 빈줄 ──
ws.row_dimensions[12].height = 10

# ── 행 13: 합계 박스 ──
row = 13
ws.row_dimensions[row].height = 30
ws.merge_cells(f"A{row}:B{row}")
ws[f"A{row}"] = "합계 금액 (VAT 별도)"
ws[f"A{row}"].font = header_font
ws[f"A{row}"].fill = blue_fill
ws[f"A{row}"].font = white_bold
ws[f"A{row}"].alignment = center
ws[f"A{row}"].border = border_all
ws.merge_cells(f"C{row}:H{row}")
ws[f"C{row}"] = "￦ 0"
ws[f"C{row}"].font = Font(name="맑은 고딕", size=14, bold=True)
ws[f"C{row}"].alignment = Alignment(horizontal="right", vertical="center")
ws[f"C{row}"].border = border_all
ws[f"C{row}"].fill = light_gray_fill

# ── 행 14: 빈줄 ──
ws.row_dimensions[14].height = 8

# ── 행 15: 품목 테이블 헤더 ──
row = 15
headers = ["No", "품명", "규격/모델", "단위", "수량", "단가", "금액", "비고"]
ws.row_dimensions[row].height = 28
for ci, h in enumerate(headers, 1):
    col = get_column_letter(ci)
    cell = ws[f"{col}{row}"]
    cell.value = h
    cell.font = white_bold
    cell.fill = blue_fill
    cell.alignment = center
    cell.border = border_all

# ── 행 16~25: 품목 입력 행 (10행) ──
for i in range(10):
    r = row + 1 + i
    ws.row_dimensions[r].height = 24
    for ci in range(1, 9):
        col = get_column_letter(ci)
        cell = ws[f"{col}{r}"]
        cell.font = normal_font
        cell.border = border_all
        if ci == 1:  # No
            cell.value = i + 1
            cell.alignment = center
        elif ci in (5, 6, 7):  # 수량, 단가, 금액
            cell.alignment = right_align
            cell.number_format = '#,##0'
        elif ci == 4:  # 단위
            cell.alignment = center
            if i == 0:
                cell.value = "EA"
        else:
            cell.alignment = left_align
    # 금액 수식: 수량 × 단가
    ws[f"G{r}"] = f"=IF(E{r}*F{r}=0,\"\",E{r}*F{r})"
    ws[f"G{r}"].number_format = '#,##0'

# ── 소계/부가세/합계 행 ──
sum_start = row + 1
sum_end = row + 10

r = row + 11  # 26
ws.row_dimensions[r].height = 26
ws.merge_cells(f"A{r}:F{r}")
ws[f"A{r}"] = "공급가액 합계"
ws[f"A{r}"].font = header_font
ws[f"A{r}"].alignment = center
ws[f"A{r}"].fill = light_gray_fill
ws[f"A{r}"].border = border_all
ws[f"G{r}"] = f"=SUM(G{sum_start}:G{sum_end})"
ws[f"G{r}"].font = header_font
ws[f"G{r}"].number_format = '#,##0'
ws[f"G{r}"].alignment = right_align
ws[f"G{r}"].border = border_all
ws[f"G{r}"].fill = light_gray_fill
ws[f"H{r}"].border = border_all
ws[f"H{r}"].fill = light_gray_fill

r += 1  # 27
ws.row_dimensions[r].height = 26
ws.merge_cells(f"A{r}:F{r}")
ws[f"A{r}"] = "부가세 (10%)"
ws[f"A{r}"].font = header_font
ws[f"A{r}"].alignment = center
ws[f"A{r}"].fill = light_gray_fill
ws[f"A{r}"].border = border_all
ws[f"G{r}"] = f"=G{r-1}*0.1"
ws[f"G{r}"].font = header_font
ws[f"G{r}"].number_format = '#,##0'
ws[f"G{r}"].alignment = right_align
ws[f"G{r}"].border = border_all
ws[f"G{r}"].fill = light_gray_fill
ws[f"H{r}"].border = border_all
ws[f"H{r}"].fill = light_gray_fill

r += 1  # 28
ws.row_dimensions[r].height = 28
ws.merge_cells(f"A{r}:F{r}")
ws[f"A{r}"] = "총 합계 (VAT 포함)"
ws[f"A{r}"].font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
ws[f"A{r}"].alignment = center
ws[f"A{r}"].fill = blue_fill
ws[f"A{r}"].border = border_all
ws[f"G{r}"] = f"=G{r-2}+G{r-1}"
ws[f"G{r}"].font = Font(name="맑은 고딕", size=12, bold=True)
ws[f"G{r}"].number_format = '#,##0'
ws[f"G{r}"].alignment = right_align
ws[f"G{r}"].border = border_all
ws[f"H{r}"].border = border_all

# ── 비고/특기사항 ──
r += 2  # 30
ws.merge_cells(f"A{r}:H{r}")
ws[f"A{r}"] = "[ 비고 / 특기사항 ]"
ws[f"A{r}"].font = header_font

r += 1  # 31
for i in range(3):
    rr = r + i
    ws.row_dimensions[rr].height = 22
    ws.merge_cells(f"A{rr}:H{rr}")
    ws[f"A{rr}"].border = Border(bottom=Side(style="dotted", color="AAAAAA"))
    ws[f"A{rr}"].font = normal_font

# ── 납품/결제 조건 ──
r += 4  # 35
ws.merge_cells(f"A{r}:H{r}")
ws[f"A{r}"] = "[ 납품 및 결제 조건 ]"
ws[f"A{r}"].font = header_font

r += 1  # 36
conditions = [
    "1. 납품조건: 협의 후 결정",
    "2. 결제조건: 협의 후 결정",
    "3. 납기일: 발주 후 협의",
]
for i, cond in enumerate(conditions):
    rr = r + i
    ws.merge_cells(f"A{rr}:H{rr}")
    ws[f"A{rr}"] = cond
    ws[f"A{rr}"].font = small_font
    ws.row_dimensions[rr].height = 20

# ── 서명란 ──
r += len(conditions) + 1  # 40
ws.row_dimensions[r].height = 10

r += 1  # 41
ws.merge_cells(f"E{r}:H{r}")
ws[f"E{r}"] = "(주)윈텍오토메이션"
ws[f"E{r}"].font = Font(name="맑은 고딕", size=12, bold=True)
ws[f"E{r}"].alignment = Alignment(horizontal="right", vertical="center")

r += 1  # 42
ws.merge_cells(f"E{r}:H{r}")
ws[f"E{r}"] = "담당:                  (인)"
ws[f"E{r}"].font = normal_font
ws[f"E{r}"].alignment = Alignment(horizontal="right", vertical="center")

# ── 인쇄 설정 ──
ws.print_area = f"A1:H{r}"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.6
ws.page_margins.right = 0.6
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5

# ── 상단 합계 금액에 총합계 연결 ──
total_row = sum_end + 3  # 28 (총합계 행)
ws["C13"] = f'="￦ "&TEXT(G{total_row},"#,##0")'

# ── 저장 ──
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"견적서 양식 저장 완료: {OUTPUT_PATH}")
